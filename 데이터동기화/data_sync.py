from PyQt5.QtGui import QFont, QIcon, QStandardItemModel, QStandardItem, QTextBlock, QTextCursor, QPalette, QColor
from PyQt5.QtCore import Qt, QThread, QObject, pyqtSignal, QCoreApplication, QDate
from selenium.common.exceptions import SessionNotCreatedException
from openpyxl.utils.exceptions import InvalidFileException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoAlertPresentException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver import Edge
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from gspread.utils import rowcol_to_a1
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from datetime import datetime, date, timedelta
from PyQt5.QtWidgets import *
from webdriver_manager.microsoft import EdgeChromiumDriverManager
import chromedriver_autoinstaller
import random
import pandas as pd
import threading
import openpyxl
import datetime
import gspread
import json
import time
import glob
import csv
import sys
import os
import re

# 각 앱에 대한 메모장 경로
CREDENTIALS_FILES = {
    "노마셀": "knowmycell_credentials.txt",
    "러브슬라임": "loveslime_credentials.txt",
    "하엔": "haen_credentials.txt"
}

# 메모장에서 ID, PW를 읽는 함수
def read_credentials(app_name):
    file_path = CREDENTIALS_FILES.get(app_name)
    if not file_path or not os.path.exists(file_path):
        return "", ""
    with open(file_path, "r") as file:
        lines = file.readlines()
        if len(lines) >= 2:
            return lines[0].strip(), lines[1].strip()
    return "", ""

# ID와 PW를 메모장에 저장하는 함수
def save_credentials(app_name, id_text, pw_text):
    file_path = CREDENTIALS_FILES.get(app_name)
    if file_path:
        with open(file_path, "w") as file:
            file.write(f"{id_text}\n{pw_text}")

# 두 번째 새 창 클래스 (노마셀 또는 러브슬라임 창)
class NewWindow(QWidget):
    def __init__(self, app_name):
        super().__init__()
        self.setWindowTitle(f"{app_name} ID & PW Edit")

        self.app_name = app_name

        # ID, PW 읽어오기
        id_text, pw_text = read_credentials(app_name)

        # ID와 PW 입력 필드
        self.id_line_edit = QLineEdit()
        self.id_line_edit.setText(id_text)
        self.pw_line_edit = QLineEdit()
        self.pw_line_edit.setText(pw_text)

        # 저장 버튼
        self.save_button = QPushButton("Save")
        self.save_button.clicked.connect(self.save_credentials)

        # 레이아웃 설정
        layout = QVBoxLayout()
        layout.addWidget(QLabel(f"{app_name} ID:"))
        layout.addWidget(self.id_line_edit)
        layout.addWidget(QLabel(f"{app_name} PW:"))
        layout.addWidget(self.pw_line_edit)
        layout.addWidget(self.save_button)

        self.setLayout(layout)

    def save_credentials(self):
        # 입력된 ID와 PW 가져오기
        id_text = self.id_line_edit.text()
        pw_text = self.pw_line_edit.text()

        # 메모장에 저장
        save_credentials(self.app_name, id_text, pw_text)

        # 저장 성공 메시지 박스 표시
        QMessageBox.information(self, "Saved", f"{self.app_name} ID and PW have been saved successfully.")


class data_synchronization(QWidget):

    def __init__(self):
        super().__init__()
        self.UI초기화()
    
    def UI초기화(self):


        # "노마셀"과 "러브슬라임" 버튼 생성
        self.knowmycell_button = QPushButton("노마셀 계정정보", self)
        self.knowmycell_button.move(30, 320)
        self.knowmycell_button.clicked.connect(self.open_nomacell_window)
        self.knowmycell_button.setStyleSheet("QPushButton { background-color: green;}")

        self.loveslime_button = QPushButton("러브슬라임 계정정보", self)
        self.loveslime_button.move(30, 350)
        self.loveslime_button.clicked.connect(self.open_loveslime_window)
        self.loveslime_button.setStyleSheet("QPushButton { background-color: green;}")

        self.haen_button = QPushButton("하엔 계정정보", self)
        self.haen_button.move(30, 380)
        self.haen_button.clicked.connect(self.open_haen_window)
        self.haen_button.setStyleSheet("QPushButton { background-color: green;}")


# 윈도우창
        self.setWindowTitle("매출 데이터 동기화")
        self.setFixedSize(800, 600)

# 시작일/마감일 lable
        self.startdate = QLabel("시작일",self)
        self.startdate.move(50, 25)
        self.enddate = QLabel("마감일",self)
        self.enddate.move(200, 25)

# 날짜 input
        self.input_startday = QDateEdit(self)
        self.input_startday.setDate(QDate.currentDate())  # 기본값: 오늘 날짜
        self.input_startday.setCalendarPopup(True)  # 캘린더 팝업 사용
        self.input_startday.setGeometry(50, 50, 100,25)

        self.input_endday = QDateEdit(self)
        self.input_endday.setDate(QDate.currentDate())  # 기본값: 오늘 날짜
        self.input_endday.setCalendarPopup(True)  # 캘린더 팝업 사용
        self.input_endday.setGeometry(200, 50, 100, 25)

# 적용 & 날짜계산(오늘부터 며칠전인지)
        self.calc_button = QPushButton("적용", self)
        self.calc_button.setGeometry(330, 49, 50, 27)
        self.calc_button.clicked.connect(self.calculate_difference)
        
# lable
        self.txt_xdays_before = QLabel("n일 전 부터 n일 전 까지", self)
        self.txt_xdays_before.setGeometry(50, 100, 500, 25)
# 브랜드 comboBox

        self.brand_comboBox = QComboBox(self)
        self.brand_comboBox.move(30, 270)
        self.brand_comboBox.addItems(["러브슬라임", "노마셀", "하엔"])
        self.brand_comboBox.setStyleSheet("QComboBox { background-color: white; }")

# run 버튼
        self.ss_down = QPushButton("스스다운", self)
        self.ss_down.move(50, 400)
        self.ss_down.clicked.connect(self.run_ss_down)

        self.ss_write = QPushButton("스스입력", self)
        self.ss_write.move(150, 400)
        self.ss_write.clicked.connect(self.run_ss_write)

        self.coup_down = QPushButton("쿠팡다운", self)
        self.coup_down.move(50, 450)
        self.coup_down.clicked.connect(self.run_coup_down)

        self.coup_write = QPushButton("쿠팡입력", self)
        self.coup_write.move(150, 450)
        self.coup_write.clicked.connect(self.run_coup_write)

# 다운로드
# 다운로드폴더 버튼
        self.slt_folder = QPushButton('다운로드폴더',self)
        self.slt_folder.setGeometry(330,511,100,29)
        self.slt_folder.clicked.connect(self.folderopen)
        self.slt_folder.setStyleSheet(
            """
            QPushButton {
                background-color: white;
                border-radius: 1.5px;
                border-width: 1px;
                border-color: black;
                border-style: solid;
            }
            QPushButton:hover {
                background-color: rgb(120,120,120);
            }
            QPushButton:pressed {
                background-color: rgb(50, 50, 50);
            }
            """
        )

        # 다운로드폴더 설정저장 버튼
        self.saveButton = QPushButton('설정저장', self)
        self.saveButton.setGeometry(440,511,100,29)
        self.saveButton.clicked.connect(self.saveText)
        self.saveButton.setStyleSheet(
            """
            QPushButton {
                background-color: white;
                border-radius: 1.5px;
                border-width: 1px;
                border-color: black;
                border-style: solid;
            }
            QPushButton:hover {
                background-color: rgb(120,120,120);
            }
            QPushButton:pressed {
                background-color: rgb(50, 50, 50);
            }
            """
        )

         # 다운로드폴더 경로
        self.path_folder = QLineEdit(self)
        self.path_folder.setGeometry(80,511,240,27)
        self.path_folder.setStyleSheet(
                        "background-color: white;"
                        "border-radius: 1.5px;"
                        "border-width: 1px;"
                        "border-color: black;"
                        "border-style: solid;")  # 테두리 스타일 추가
        self.path_folder.setReadOnly(True)

        self.loadText()

# 함수 지정
        self.diff1 = None  # diff1 값을 저장할 인스턴스 변수
        self.diff2 = None  # diff2 값을 저장할 인스턴스 변수

        global download_folder
        download_folder = self.path_folder.text()

    
    def loveslime_credential(self):
        try:
            with open('loveslime_credentials.txt', 'r', encoding='utf-8') as f:
                lines = f.readlines()  # 파일의 모든 줄을 읽어 리스트로 저장

            if len(lines) >= 2:  # 첫 번째 줄과 두 번째 줄이 있는지 확인
                id_value = lines[0].strip()  # 첫 번째 줄 (ID)
                pw_value = lines[1].strip()  # 두 번째 줄 (PW)
                return id_value, pw_value  # ID와 PW 반환
            else:
                print("File does not contain enough lines for ID and PW.")
                return None, None  # ID와 PW를 찾을 수 없는 경우
        except FileNotFoundError:
            print("Cannot find login information.")
            return None, None
        
    def knowmycell_credential(self):
        try:
            with open('knowmycell_credentials.txt', 'r', encoding='utf-8') as f:
                lines = f.readlines()  # 파일의 모든 줄을 읽어 리스트로 저장

            if len(lines) >= 2:  # 첫 번째 줄과 두 번째 줄이 있는지 확인
                id_value = lines[0].strip()  # 첫 번째 줄 (ID)
                pw_value = lines[1].strip()  # 두 번째 줄 (PW)
                return id_value, pw_value  # ID와 PW 반환
            else:
                print("File does not contain enough lines for ID and PW.")
                return None, None  # ID와 PW를 찾을 수 없는 경우
        except FileNotFoundError:
            print("Cannot find login information.")
            return None, None
        
    def haen_credential(self):
        try:
            with open('haen_credentials.txt', 'r', encoding='utf-8') as f:
                lines = f.readlines()  # 파일의 모든 줄을 읽어 리스트로 저장

            if len(lines) >= 2:  # 첫 번째 줄과 두 번째 줄이 있는지 확인
                id_value = lines[0].strip()  # 첫 번째 줄 (ID)
                pw_value = lines[1].strip()  # 두 번째 줄 (PW)
                return id_value, pw_value  # ID와 PW 반환
            else:
                print("File does not contain enough lines for ID and PW.")
                return None, None  # ID와 PW를 찾을 수 없는 경우
        except FileNotFoundError:
            print("Cannot find login information.")
            return None, None

    def open_nomacell_window(self):
        self.nomacell_window = NewWindow("노마셀")
        self.nomacell_window.show()

    def open_loveslime_window(self):
        self.loveslime_window = NewWindow("러브슬라임")
        self.loveslime_window.show()

    def open_haen_window(self):
        self.haen_window = NewWindow("하엔")
        self.haen_window.show()

    def save_credentials(self):
        # 입력된 ID와 PW 가져오기
        id_text = self.id_line_edit.text()
        pw_text = self.pw_line_edit.text()

        # 메모장에 저장
        save_credentials(self.app_name, id_text, pw_text)

        # 저장 성공 메시지 박스 표시
        QMessageBox.information(self, "Saved", f"{self.app_name} ID and PW have been saved successfully.")

    def calculate_difference(self):

        print(self.input_startday.text())

        today = QDate.currentDate()

        # 두 날짜 가져오기
        date1 = self.input_startday.date()
        date2 = self.input_endday.date()

        self.diff1 = date1.daysTo(today)  # 인스턴스 변수에 저장
        self.diff2 = date2.daysTo(today)  # 인스턴스 변수에 저장

        # QLabel에 결과 출력
        self.txt_xdays_before.setText(f"{self.diff1}일 전 부터 {self.diff2}일 전 까지")

    def run_ss_down(self):
        
        print(self.brand_comboBox.currentText())

        if self.brand_comboBox.currentText() == "하엔":
            pass
        
        else:

            # 날짜 구하기
            today = date.today()
            # 하루를 나타내는 timedelta 객체 생성
            # 어제 날짜를 구함

            today_date = today.strftime("%d")
            today_Ym = today.strftime("%Y. %m.")

            dayx = datetime.timedelta(days=self.diff1)
            dayy = datetime.timedelta(days=self.diff2-1)
            day1 = datetime.timedelta(days=1)

            # 오늘 날짜 구하기
            today_yday = today-day1
            startday = today-dayx
            endday = today-dayy
            tday_Ym = startday.strftime("%Y. %m.")
            tday_d = startday.strftime("%d")

            # while startday != endday:
            #         # 날짜 구하기
            #         today = date.today()
            #         # 하루를 나타내는 timedelta 객체 생성
            #         # 어제 날짜를 구함

            #         today_date = today.strftime("%d")
            #         today_Ym = today.strftime("%Y. %m.")

            #         dayx = datetime.timedelta(days)

            #         # 오늘 날짜 구하기
            #         startday += datetime.timedelta(days=1)
            #         print(startday)

            # EdgeOptions 객체 생성
            edge_options = webdriver.EdgeOptions()
            edge_options.use_chromium = True
            edge_options.add_argument("disable-gpu")
            edge_options.add_argument("no-sandbox")

            # 사용자의 프로필 경로 설정
            profile_path = 'C:\\Users\\A\\AppData\\Local\\Microsoft\\Edge\\User Data1'
            edge_options.add_argument(f"user-data-dir={profile_path}")
            edge_options.add_argument("--profile-directory=Default")

            # Edge 드라이버 서비스 시작
            edge_service = Service(EdgeChromiumDriverManager().install())
            edge_driver = webdriver.Edge(service=edge_service, options=edge_options)


            def count_files(download_folder):
                            """ 폴더 내 파일의 개수를 반환합니다. """
                            return len([name for name in os.listdir(download_folder) if os.path.isfile(os.path.join(download_folder, name))])

            def get_latest_file(download_folder):
                            """ 폴더 내에서 가장 최신의 파일을 반환합니다. """
                            files = [os.path.join(download_folder, f) for f in os.listdir(download_folder) if os.path.isfile(os.path.join(download_folder, f))]
                            latest_file = max(files, key=os.path.getctime)
                            return latest_file

            def check_download():
                # 다운로드 전의 파일 개수 확인
                initial_file_count = count_files(download_folder)

                # 다운로드 시작 ...

                # 새 파일이 다운로드될 때까지 기다림
                global check
                check = 0
                i = 0

                while i < 20:
                    current_file_count = count_files(download_folder)
                    if current_file_count > initial_file_count:
                        print("A new file has been downloaded.")
                        latest_file = get_latest_file(download_folder)
                        print(f"Downloaded file: {latest_file}")
                        # 여기서 필요한 작업을 수행하세요, 예를 들면 파일 열기 등
                        check = 1
                        break
                    else:
                        print("Still no new file")
                    time.sleep(1)  # 폴더 상태를 1초마다 체크
                    i += 1

                    get_latest_file(download_folder)
                return check

            edge_driver.get("https://bizadvisor.naver.com/shopping/product")

            # 로그인
            WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#wrap > div > div > div.login_box > ul > li:nth-child(1) > a"))).click()
            WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#root > div > div.Layout_wrap__3uDBh > div > div > div.Login_simple_box__2bfAS > button"))).click()

            # 상품별 이동
            WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#include_nav > div > div > div:nth-child(1) > ul > li:nth-child(4) > a"))).click()
            WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#include_nav > div > div > div:nth-child(1) > ul > li.on > div > ul > li:nth-child(1) > a"))).click()


            print(edge_driver.find_element(By.CSS_SELECTOR, "#include_header > div > div.header_tit > div > div:nth-child(2) > div > div > div > div > a > h2 > span > p").text)

            if edge_driver.find_element(By.CSS_SELECTOR, "#include_header > div > div.header_tit > div > div:nth-child(2) > div > div > div > div > a > h2 > span > p").text == self.brand_comboBox.currentText():
                pass
            
            else:
                edge_driver.find_element(By.XPATH, f'//*[contains(text(), "{self.brand_comboBox.currentText()}")]').click()


            while startday != endday:

                # 날짜 클릭(달력오픈)
                WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.date_select > a.btn.select_data'))).click()

                tday_Ym = startday.strftime("%Y. %m.")
                tday_d = str(int(startday.strftime("%d")))
                print(startday)
                trick = (startday-day1).strftime("%Y. %m.")

                while True:
                    DPmonth = edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_area > div.pick_calendar_layout > div.DayPicker.DayPicker > div > div.DayPicker-Month.rdp-caption_start")

                    if trick == DPmonth.text[:9]:

                        if tday_Ym == DPmonth.text[:9]:

                            days = edge_driver.find_elements(By.XPATH, f"//*[@id='wrap']/div[1]/section/div/div[2]/div[1]/div/ul/li[2]/div/div/div[2]/div/div[1]/div[2]/div[2]/div/div[1]/table/tbody//td[not(contains(@class, 'DayPicker-Day--outside'))]")
                            for day in days:
                                if day.text == tday_d:
                                    print("target: ", day.text)
                                    day.click()
                                    time.sleep(0.1)
                                    day.click()
                                    break

                            # 적용
                            edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_info_area > div.btn_area > a:nth-child(1)").click()
                            break

                        WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_area > div.pick_calendar_layout > div.DayPicker-NavBar > span.DayPicker-NavButton.DayPicker-NavButton--next'))).click()

                        DPmonth = edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_area > div.pick_calendar_layout > div.DayPicker.DayPicker > div > div.DayPicker-Month.rdp-caption_start")

                        if tday_Ym == DPmonth.text[:9]:

                            days = edge_driver.find_elements(By.XPATH, f"//*[@id='wrap']/div[1]/section/div/div[2]/div[1]/div/ul/li[2]/div/div/div[2]/div/div[1]/div[2]/div[2]/div/div[1]/table/tbody//td[not(contains(@class, 'DayPicker-Day--outside'))]")
                            for day in days:
                                if day.text == tday_d:
                                    print("target: ", day.text)
                                    day.click()
                                    time.sleep(0.1)
                                    day.click()
                                    break

                            # 적용
                            edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_info_area > div.btn_area > a:nth-child(1)").click()
                            break

                        

                    else:
                        # 월 이동 버튼
                        WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_area > div.pick_calendar_layout > div.DayPicker-NavBar > span.DayPicker-NavButton.DayPicker-NavButton--prev'))).click()


                    if tday_Ym == DPmonth.text[:9]:

                            days = edge_driver.find_elements(By.XPATH, f"//*[@id='wrap']/div[1]/section/div/div[2]/div[1]/div/ul/li[2]/div/div/div[2]/div/div[1]/div[2]/div[2]/div/div[1]/table/tbody//td[not(contains(@class, 'DayPicker-Day--outside'))]")
                            for day in days:
                                if day.text == tday_d:
                                    print("target: ", day.text)
                                    day.click()
                                    time.sleep(0.1)
                                    day.click()
                                    break

                            # 적용
                            edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_info_area > div.btn_area > a:nth-child(1)").click()
                            break

                # 다운로드 버튼
                WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(1) > span > a'))).click()

                check_download()

                startday += datetime.timedelta(days=1)

            edge_driver.close()

    def run_ss_write(self):

        if self.brand_comboBox.currentText() == "하엔":
            pass
        
        else:
            target_days = self.diff1
            target_num = self.diff1-self.diff2+1
            input_num = 2


            today = date.today()
            today_date = today.strftime("%d")
            today_month = str(int(today.strftime("%m")))

            dayx = datetime.timedelta(days=target_days)
            day1 = datetime.timedelta(days=1)

            today_yday = today-day1
            today_tday = today-dayx
            print(today_tday)

            weekday_num = today.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)
            weekday_numy = today_yday.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)
            weekday_numt = today_tday.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)

            download_folder = "C:\\Users\\A\\Downloads"

            def count_files(folder):
                """ 폴더 내 파일의 개수를 반환합니다. """
                return len([name for name in os.listdir(folder) if os.path.isfile(os.path.join(folder, name))])

            def get_latest_file(folder):
                """ 폴더 내에서 가장 최신의 파일을 반환합니다. """
                files = [os.path.join(folder, f) for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
                latest_file = max(files, key=os.path.getctime)
                return latest_file

            def check_download():
                # 다운로드 전의 파일 개수 확인
                    initial_file_count = count_files(download_folder)

                    # 다운로드 시작 ...

                    # 새 파일이 다운로드될 때까지 기다림
                    global check
                    check = 0
                    i = 0

                    while i < 20:
                        current_file_count = count_files(download_folder)
                        if current_file_count > initial_file_count:
                            print("A new file has been downloaded.")
                            latest_file = get_latest_file(download_folder)
                            print(f"Downloaded file: {latest_file}")
                            # 여기서 필요한 작업을 수행하세요, 예를 들면 파일 열기 등
                            check = 1
                            break
                        else:
                            print("Still no new file")
                        time.sleep(1)  # 폴더 상태를 1초마다 체크
                        i += 1
                    return check

            def get_nth_latest_file(folder, n):
                            """폴더 내에서 n번째로 최신 파일을 반환합니다. n이 1이면 가장 최신, 2면 두 번째로 최신 파일을 반환합니다."""
                            # 폴더 내의 파일들의 전체 경로와 함께 리스트를 생성합니다.
                            files = [os.path.join(folder, f) for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
                            
                            # 파일이 없다면 None을 반환
                            if not files:
                                return None
                            
                            # 파일들을 생성 시간 기준으로 정렬합니다.
                            files.sort(key=os.path.getctime, reverse=True)
                            
                            # 요청한 순위의 파일을 반환합니다. n이 파일 수보다 많거나 0 이하인 경우 None을 반환합니다.
                            if 1 <= n <= len(files):
                                nth_latest_file = files[n-1]  # n번째 파일 선택
                                print(nth_latest_file)
                                return nth_latest_file
                            else:
                                return None

            def check_data_in_second_row(file_path):
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active
                second_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
                if second_row and any(cell is not None for cell in second_row[0]):
                    return True
                return False

            row_index = 1
            while target_num > 0:

                i = get_nth_latest_file(download_folder, target_num)

                wb = openpyxl.load_workbook(i)
                sheet = wb.active  # 활성 시트 선택

                # 새 워크북 생성
                new_wb = openpyxl.load_workbook('new_file.xlsx')
                new_sheet = new_wb.active

                
                if check_data_in_second_row(i):
                    pass
                else:
                    new_sheet.cell(row=row_index, column=1, value=today_tday.strftime("%Y-%m-%d"))
                    row_index += 1  # 새 시트에서의 행 인덱스 증가
                    today_tday += timedelta(days=1)  # 날짜 하루 증가
                    target_num -= 1
                    new_wb.save('new_file.xlsx')
                    continue

                # 원본 시트의 행을 반복하며 첫 번째 행을 제외하고 데이터가 있는 행만 복사
                for row in sheet.iter_rows(min_row=2):  # 첫 번째 행을 제외하고 시작
                    # 각 셀에 데이터가 있는지 확인
                    data_exists = any(cell.value not in (None, '', ' ') for cell in row)  # 빈 문자열과 공백도 무시

                    # 날짜 입력
                    new_sheet.cell(row=row_index, column=1, value=today_tday.strftime("%Y-%m-%d"))

                    for col_index, cell in enumerate(row, start=2):
                        new_sheet.cell(row=row_index, column=col_index, value=cell.value)


                    row_index += 1  # 새 시트에서의 행 인덱스 증가
                today_tday += timedelta(days=1)  # 날짜 하루 증가

                # 새로운 파일에 저장
                new_wb.save('new_file.xlsx')

                print(target_num)
                target_num -= 1

    def run_coup_down(self):
        download_folder = "C:\\Users\\A\\Downloads"

        if self.brand_comboBox.currentText() == "노마셀":
            id, pw = self.knowmycell_credential()
            print(id,"\n",pw,"\n")

        elif self.brand_comboBox.currentText() == "러브슬라임":
            id, pw = self.loveslime_credential()
            print(id,"\n",pw,"\n")

        elif self.brand_comboBox.currentText() == "하엔":
            id, pw = self.haen_credential()
            print(id,"\n",pw,"\n")


        def count_files(folder):
            """ 폴더 내 파일의 개수를 반환합니다. """
            return len([name for name in os.listdir(folder) if os.path.isfile(os.path.join(folder, name))])

        def get_latest_file(folder):
            """ 폴더 내에서 가장 최신의 파일을 반환합니다. """
            files = [os.path.join(folder, f) for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
            latest_file = max(files, key=os.path.getctime)
            return latest_file

        def check_download():
            # 다운로드 전의 파일 개수 확인
                initial_file_count = count_files(download_folder)

                # 다운로드 시작 ...

                # 새 파일이 다운로드될 때까지 기다림
                global check
                check = 0
                i = 0

                while i < 20:
                    current_file_count = count_files(download_folder)
                    if current_file_count > initial_file_count:
                        print("A new file has been downloaded.")
                        latest_file = get_latest_file(download_folder)
                        print(f"Downloaded file: {latest_file}")
                        # 여기서 필요한 작업을 수행하세요, 예를 들면 파일 열기 등
                        check = 1
                        break
                    else:
                        print("Still no new file")
                    time.sleep(1)  # 폴더 상태를 1초마다 체크
                    i += 1
                return check
                        

        # 변수 지정
        today = date.today()
        today_date = today.strftime("%d")
        today_month = str(int(today.strftime("%m")))

        dayx = datetime.timedelta(days=self.diff1)
        dayy = datetime.timedelta(days=self.diff2-1)
        day1 = datetime.timedelta(days=1)

        today_yday = today-day1
        startday = today-dayx
        endday = today-dayy
        print(startday)

        weekday_num = today.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)
        weekday_numy = today_yday.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)
        weekday_numt = startday.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)

        # 크롬 On
        chromedriver_path = chromedriver_autoinstaller.install()
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_argument("--start-maximized") #최대 크기로 시작
        chrome_options.add_experimental_option('detach', True)
        user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36"
        headers = {'user-agent' : user_agent}

        driver = webdriver.Chrome(
                        service=Service(chromedriver_path),
                        options=chrome_options
                    )


        # 접속
        driver.get("https://wing.coupang.com/seller/notification/metrics/dashboard")  # 로그인 시작

                
        # 로그인
        input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#username")))
        input_field.click()
        time.sleep(0.7)
        input_field.send_keys(Keys.CONTROL + "a")
        input_field.send_keys(Keys.BACKSPACE)
        driver.find_element(By.CSS_SELECTOR, "#username").send_keys(id)
        input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#password")))
        input_field.click()
        input_field.send_keys(Keys.CONTROL + "a")
        input_field.send_keys(Keys.BACKSPACE)
        driver.find_element(By.CSS_SELECTOR, "#password").send_keys(pw)
        driver.find_element(By.CSS_SELECTOR,'#kc-login').click()


        # target_days 날짜만큼 파일 다운로드
        while startday != endday:


            startdayInputOnly = startday.strftime("%Y-%m-%d")
            
            # 날짜 입력
            WebDriverWait(driver,15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#search-filter-panel > div:nth-child(1) > span.sc-common-date-group > span:nth-child(1) > i")))
            input_field = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#dateEnd")))
            input_field.send_keys(Keys.CONTROL + "a")
            input_field.send_keys(Keys.BACKSPACE)
            input_field.send_keys(startdayInputOnly)

            time.sleep(0.1)
            input_field = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#dateStart")))
            input_field.send_keys(Keys.CONTROL + "a")
            input_field.send_keys(Keys.BACKSPACE)
            input_field.send_keys(startdayInputOnly)

            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#date_range"))) # 날짜변경확인


            time.sleep(1)
            
            try:
                try:
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#download-product-info"))).click() # 다운로드
                    driver.find_element(By.CSS_SELECTOR,"#download-product-info").click() # 다운로드
                except:
                    time.sleep(1)
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#download-product-info"))).click() # 다운로드
                    driver.find_element(By.CSS_SELECTOR,"#download-product-info").click() # 다운로드

            except:
                # element = driver.find_element(By.CSS_SELECTOR, "#container-wing-v2 > div:nth-child(5) > div.sweet-alert.showSweetAlert.visible > button.confirm")
                # ActionChains(driver).move_to_element_with_offset(element,15,20).click().perform()
                # ActionChains(driver).move_to_element_with_offset(element,15,25).click().perform()
                confirm_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, "confirm")))
                driver.execute_script("arguments[0].click();", confirm_button)

                try:
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#download-product-info"))).click() # 다운로드
                    driver.find_element(By.CSS_SELECTOR,"#download-product-info").click() # 다운로드
                except:
                    time.sleep(1)
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#download-product-info"))).click() # 다운로드
                    driver.find_element(By.CSS_SELECTOR,"#download-product-info").click() # 다운로드
            
            check_download()

            startday += datetime.timedelta(days=1)

        driver.close()

    def run_coup_write(self):


        target_days = self.diff1
        target_num = self.diff1-self.diff2+1
        input_num = 2

        today = date.today()
        today_date = today.strftime("%d")
        today_month = str(int(today.strftime("%m")))

        dayx = datetime.timedelta(days=target_days)
        day1 = datetime.timedelta(days=1)

        today_yday = today-day1
        today_tday = today-dayx
        print(today_tday)

        weekday_num = today.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)
        weekday_numy = today_yday.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)
        weekday_numt = today_tday.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)

        download_folder = "C:\\Users\\A\\Downloads"

        def count_files(folder):
            """ 폴더 내 파일의 개수를 반환합니다. """
            return len([name for name in os.listdir(folder) if os.path.isfile(os.path.join(folder, name))])

        def get_latest_file(folder):
            """ 폴더 내에서 가장 최신의 파일을 반환합니다. """
            files = [os.path.join(folder, f) for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
            latest_file = max(files, key=os.path.getctime)
            return latest_file

        def check_download():
            # 다운로드 전의 파일 개수 확인
                initial_file_count = count_files(download_folder)

                # 다운로드 시작 ...

                # 새 파일이 다운로드될 때까지 기다림
                global check
                check = 0
                i = 0

                while i < 20:
                    current_file_count = count_files(download_folder)
                    if current_file_count > initial_file_count:
                        print("A new file has been downloaded.")
                        latest_file = get_latest_file(download_folder)
                        print(f"Downloaded file: {latest_file}")
                        # 여기서 필요한 작업을 수행하세요, 예를 들면 파일 열기 등
                        check = 1
                        break
                    else:
                        print("Still no new file")
                    time.sleep(1)  # 폴더 상태를 1초마다 체크
                    i += 1
                return check

        def get_nth_latest_file(folder, n):
                        """폴더 내에서 n번째로 최신 파일을 반환합니다. n이 1이면 가장 최신, 2면 두 번째로 최신 파일을 반환합니다."""
                        # 폴더 내의 파일들의 전체 경로와 함께 리스트를 생성합니다.
                        files = [os.path.join(folder, f) for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
                        
                        # 파일이 없다면 None을 반환
                        if not files:
                            return None
                        
                        # 파일들을 생성 시간 기준으로 정렬합니다.
                        files.sort(key=os.path.getctime, reverse=True)
                        
                        # 요청한 순위의 파일을 반환합니다. n이 파일 수보다 많거나 0 이하인 경우 None을 반환합니다.
                        if 1 <= n <= len(files):
                            nth_latest_file = files[n-1]  # n번째 파일 선택
                            print(nth_latest_file)
                            return nth_latest_file
                        else:
                            return None

        
        while target_num > 0:
            dayx = datetime.timedelta(days=target_days)
            today_tday = today-dayx

            xlsx_file = get_nth_latest_file(download_folder, target_num)

            df_uploaded_new = pd.read_excel(xlsx_file)
            # '러브슬라임'이라는 단어가 포함된 모든 행을 '옵션명' 열을 기준으로 필터링합니다.
            filtered_rows_with_loveslime = df_uploaded_new[df_uploaded_new['옵션명'].astype(str).str.contains(self.brand_comboBox.currentText())]

            # 필터링된 행들의 데이터를 리스트로 변환합니다.
            rows_list_with_loveslime = filtered_rows_with_loveslime.values.tolist()

            # 두 번째 값만 정수형으로 변환한 후 문자열로 변환하여 업데이트하는 과정
            updated_data_list = []
            for row in rows_list_with_loveslime:
                new_row = row.copy()  # 원본 데이터의 복사본 생성
                if len(row) > 1:  # 두 번째 값이 존재하는지 확인
                    new_row[1] = str(int(row[1]))  # 두 번째 값을 정수형으로 변환 후 문자열로 변환
                updated_data_list.append(new_row)


            print(updated_data_list,"\n")
            print(len(updated_data_list))


            # 결과 출력

            wb = load_workbook('new_file.xlsx')
            ws = wb.active
            col_index = 2  # B열부터 시작

            if len(updated_data_list) > 1:

                i = 0
                while i < len(updated_data_list):
                    col_index = 2  # B열부터 시작
                    for item in updated_data_list[i]:  # 각 리스트의 항목을 셀에 할당
                        cell = ws.cell(row=input_num, column=col_index)
                        cell.value = item
                        

                        ws[f'A{input_num}'] = today_tday

                        col_index += 1
                    input_num += 1
                    i += 1



            # 10개의 행에 숫자 입력
            elif len(updated_data_list) == 1:
                ws[f'A{input_num}'] = today_tday

                for data_list in updated_data_list:  # updated_data_list 내의 각 리스트에 대해
                    for item in data_list:  # 각 리스트의 항목을 셀에 할당
                        cell = ws.cell(row=input_num, column=col_index)
                        cell.value = item
                        col_index += 1
                input_num += 1
            

            # 파일 저장
            wb.save('new_file.xlsx')

            target_days -= 1
            target_num -= 1

    def run_cafe(self):

        if self.brand_comboBox.currentText() == "노마셀":
            id, pw = self.knowmycell_credential()
            print(id,"\n",pw,"\n")

        elif self.brand_comboBox.currentText() == "러브슬라임":
            id, pw = self.loveslime_credential()
            print(id,"\n",pw,"\n")

        elif self.brand_comboBox.currentText() == "하엔":
            id, pw = self.haen_credential()
            print(id,"\n",pw,"\n")

         # 크롬 On
        ### chromedriver_autoinstaller.install() 사용 추가
        chromedriver_path = chromedriver_autoinstaller.install()
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_argument("--start-maximized") #최대 크기로 시작
        # chrome_options.add_argument('--incognito')
        # chrome_options.add_argument('--window-size=1920,1080')  
        # chrome_options.add_argument('--headless')
        chrome_options.add_experimental_option('detach', True)

        user_data = self.chrome_path_folder.text()
        user_data = 'C:\\Users\\A\\AppData\\Local\\Google\\Chrome\\User Data1'
        chrome_options.add_argument(f"user-data-dir={user_data}")
        chrome_options.add_argument("--profile-directory=Profile 1")
        
        user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36"
        headers = {'user-agent' : user_agent}

        driver = webdriver.Chrome(
            service=Service(chromedriver_path),
            options=chrome_options
        )

        
        
        driver.get("https://eclogin.cafe24.com/Shop/")

        ##################################### 로그인
        ##################################### 로그인
        ##################################### 로그인
        ##################################### 로그인

        # ID
        input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#mall_id")))
        input_field.click()
        time.sleep(1)
        input_field.send_keys(Keys.CONTROL + "a")
        input_field.send_keys(Keys.BACKSPACE)
        driver.find_element(By.CSS_SELECTOR, "#mall_id").send_keys(id)

        # PW
        input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#userpasswd")))
        input_field.click()
        input_field.send_keys(Keys.CONTROL + "a")
        input_field.send_keys(Keys.BACKSPACE)
        driver.find_element(By.CSS_SELECTOR, "#userpasswd").send_keys(pw)

        # 로그인클릭
        driver.find_element(By.CSS_SELECTOR,'#frm_user > div > div.mButton > button').click()

        #비밀번호변경안내
        try: WebDriverWait(driver, 5).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#iptBtnEm")))).click() 
        except: pass

        try:
            time.sleep(3)
            popup = driver.find_element(By.XPATH, '//*[contains(text(), "오늘 하루 보지 않기")]')
            popup.click()

        except: pass

        #화면로딩대기
        WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, '//*[contains(text(), "오늘의 할 일")]')))

        #통계 화면 이동
        WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.ID, 'QA_Lnb_Menu2060'))).click()
        WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.ID, 'QA_Lnb_Menu2062'))).click()
        WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.ID, 'QA_Lnb_Menu2063'))).click()

        input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#pr_start_date")))
        input_field.click()
        time.sleep(1)
        input_field.send_keys(Keys.CONTROL + "a")
        input_field.send_keys(Keys.BACKSPACE)
        input_field.send_keys(self.startdate.text())

        input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#pr_end_date")))
        input_field.click()
        time.sleep(1)
        input_field.send_keys(Keys.CONTROL + "a")
        input_field.send_keys(Keys.BACKSPACE)
        input_field.send_keys(self.)

        # 자세히보기클릭
        element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#QA_day3 > div.mBoard.gScroll > table")))
        driver.execute_script("arguments[0].scrollIntoView(true);", element) # 스크롤다운
        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#sReportGabView"))).click() 
        
        element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#QA_day3 > div.mBoard.gScroll > table")))
        driver.execute_script("arguments[0].scrollIntoView(true);", element) # 스크롤다운
        rows = driver.find_elements(By.CSS_SELECTOR, 'tbody.right tr')

    def folderopen(self):
        fname = QFileDialog.getExistingDirectory(self,'폴더선택','')
        self.path_folder.setText(fname)

    def loadText(self):
            try:
                with open('saved_text.txt', 'r') as f:
                    saved_text = f.read()
                    texts = saved_text.split("\n")

                    self.path_folder.setText(texts[0])

            except FileNotFoundError:
                pass

    def saveText(self):
        text = self.path_folder.text()
        with open('saved_text.txt', 'w') as file:
            file.write(text)
        QMessageBox.information(self,'알림','저장되었습니다.')

    def my_exception_hook(exctype, value, traceback):
        # Print the error and traceback
        print(exctype, value, traceback)
        # Call the normal Exception hook after
        sys._excepthook(exctype, value, traceback)
        # sys.exit(1)

    # Back up the reference to the exceptionhook
    sys._excepthook = sys.excepthook

    # Set the exception hook to our wrapping function
    sys.excepthook = my_exception_hook

if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = data_synchronization()
    win.show()
    sys.exit(app.exec_())
