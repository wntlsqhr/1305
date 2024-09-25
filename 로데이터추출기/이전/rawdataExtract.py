# 1.2.1 메타(노마셀), 구글(하엔, 노마셀) 다시 사용
# 코드수정 점검
# 1. 카페24 판매 데이터 날짜 못찾으면 넘어가도록 코드 수정
# 2. 쿠팡C 다운로드 클릭 안되는 문제 수정(try:구문 넣고 다운로드 클릭 한줄 추가)
# 3. 쿠팡R 바로 다운로드 되면 check download 감지 안되는 문제 -> 감지 기능 삭제
# 4. 쿠팡C 데이터 없으면 0이 입력되도록 코드수정
# 5. 네이버 스스 이전 달 날짜 선택 안되는 코드 변경
# 6. 메타 캠페인, 일 체크 동작 삭제(CSS선택자 변경 이슈)
# 7. 카페24 R, 데이터 없으면 0 입력 되도록 코드 수정
# 8. 메타 n일전 데이터 불러오기(데이터 없으면 더미데이터 입력)
#240613
# 9. 로그인창으로 가는 선택자가 바뀌어, 다른 로직의 이동방법 적용
# 10. 카페24(파컨) 홍보용 팝업창 대비, 다른 페이지 이동 후 통계 팝업 창 띄우도록 변경
# 11. 구글 다운로드 클릭, CSS선택자 방식에서 XPATH - contains, text() 방식으로 변경(CSS선택자 매번바뀜)
#240614
# 네이버검색광고 추가(러블로, 노마셀)
#240618
# 로데이터, 쿠팡 다운로드 체크 오류 -> 다운로드 확인 방식 변경 (check_download() -> 수식 풀어서 확인)
# 스스 엣지드라이버 꺼짐 오류 -> 엣지드라이버 업데이트
#240620
# 네이버검색광고 하엔 추가
# 구글ads n일 전 데이터 추출 기능 추가
# 파컨 화면 로딩대기 로직 수정
# 네이버검색광고 로그인 확인 창 제거 로직 수정(visibility_of_all_elements_located -> element_to_be_clickable)
#240709
# 쿠팡 보고서 -> 어제 클릭 -> 새로고침 과정에서 새로고침 선택 안되는 문제(CSS요소 변동) 수정 // 텍스트로 새로고침 인식하게 함
## 구글 다운로드 실패 시 재시도 적용
#240710
### 보고서 생성 실패 시 한 번 더 시도
#240725
### chromedriver_autoinstaller.install() 사용 추가
#chromedriver_path = chromedriver_autoinstaller.install()
#driver = webdriver.Chrome(
#    service=Service(chromedriver_path),
#    options=chrome_options
#)
#240730
### 보고서 생성 실패하면 페이지 다시 로딩 후 생성







from PyQt5.QtGui import QFont, QIcon, QStandardItemModel, QStandardItem, QTextBlock, QTextCursor
from PyQt5.QtCore import Qt, QThread, QObject, pyqtSignal, QCoreApplication
from selenium.common.exceptions import SessionNotCreatedException
from openpyxl.utils.exceptions import InvalidFileException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoAlertPresentException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from gspread.utils import rowcol_to_a1
from gspread_formatting import *
from gspread.exceptions import APIError
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from datetime import datetime, date, timedelta
from PyQt5.QtWidgets import *
import pandas as pd
import chromedriver_autoinstaller
import datetime
import threading
import openpyxl
import gspread
import json
import time
import glob
import csv
import sys
import os
import re


class Rawdata_extractor(QWidget):

    def __init__(self):
        super().__init__()
        self.UI초기화()

    def UI초기화(self):

        self.setWindowTitle("Raw data 자동 추출기")
        self.setWindowIcon(QIcon('1305.ico'))
        self.setStyleSheet("background-color: rgb(192, 197, 215)")
        self.setFixedSize(800, 600)

        self.maintitle = QLabel("Raw data 자동 추출기", self)
        self.maintitle.move(50, 50)
        self.maintitle.setFont(QFont('Helvetia', 30, QFont.Bold))
        self.maintitle.setStyleSheet("color: rgb(11, 27, 83);"
                            "background-color: #FA8072;"
                             "border-style: solid;"
                             "border-width: 2px;"
                             "border-color: #FA8072;"
                             "border-radius: 8px")
        
        #플랫폼
        self.lable_cafe = QLabel("카페24",self)
        self.lable_cafe.move(120,150)
        self.lable_cafe.setFont(QFont('Helvetia', 12, QFont.Bold))

        self.lable_coup = QLabel("쿠팡",self)
        self.lable_coup.move(240,150)
        self.lable_coup.setFont(QFont('Helvetia', 12, QFont.Bold))

        self.lable_meta = QLabel("메타",self)
        self.lable_meta.move(340,150)
        self.lable_meta.setFont(QFont('Helvetia', 12, QFont.Bold))

        self.lable_goog = QLabel("구글",self)
        self.lable_goog.move(440,150)
        self.lable_goog.setFont(QFont('Helvetia', 12, QFont.Bold))

        self.lable_ss = QLabel("스스",self)
        self.lable_ss.move(540,150)
        self.lable_ss.setFont(QFont('Helvetia', 12, QFont.Bold))

        self.lable_pc = QLabel("파컨",self)
        self.lable_pc.move(640,150)
        self.lable_pc.setFont(QFont('Helvetia', 12, QFont.Bold))

        self.lable_pc = QLabel("네이버검색광고",self)
        self.lable_pc.move(740,150)
        self.lable_pc.setFont(QFont('Helvetia', 12, QFont.Bold))      


        #밑줄
        self.line = QLabel("─────────────────────────────────────────",self)
        self.line.move(35,180)
        self.line.setFont(QFont('Helvetia', 12, QFont.Bold))

        # 옆줄
        # self.log = QLabel("test", self)
        # self.log.move(620,180)
        # self.log.setFont(QFont('Helvetia', 12))

        #브랜드
        self.lable_haen = QLabel("하엔",self)
        self.lable_haen.move(60,230)
        self.lable_haen.setFont(QFont('Helvetia', 12, QFont.Bold))

        self.lable_lovl = QLabel("러블로",self)
        self.lable_lovl.move(60,280)
        self.lable_lovl.setFont(QFont('Helvetia', 12, QFont.Bold))

        self.lable_know = QLabel("노마셀",self)
        self.lable_know.move(60,330)
        self.lable_know.setFont(QFont('Helvetia', 12, QFont.Bold))


        #체크박스 카페        
        font = QFont()
        font.setPointSize(16)  # 폰트 크기 설정
        self.chk_cafe_haen = QCheckBox("하엔", self)
        self.chk_cafe_haen.move(150,230)
        self.chk_cafe_haen.setFont(font)
        self.chk_cafe_lovl = QCheckBox("러블로", self)
        self.chk_cafe_lovl.move(150,280)
        self.chk_cafe_lovl.setFont(font)
        self.chk_cafe_know = QCheckBox("노마셀", self)
        self.chk_cafe_know.move(150,330)
        self.chk_cafe_know.setFont(font)
        self.chk_cafe_ZQ = QCheckBox("제니크", self)
        self.chk_cafe_ZQ.move(150,380)
        self.chk_cafe_ZQ.setFont(font)

        #체크박스 쿠팡
        self.chk_coup_haen = QCheckBox("하엔", self)
        self.chk_coup_haen.move(250,230)
        self.chk_coup_haen.setFont(font)
        self.chk_coup_lovl = QCheckBox("러블로", self)
        self.chk_coup_lovl.move(250,280)
        self.chk_coup_lovl.setFont(font)
        self.chk_coup_know = QCheckBox("노마셀", self)
        self.chk_coup_know.move(250,330)
        self.chk_coup_know.setFont(font)

        #체크박스 메타
        self.chk_meta_haen = QCheckBox("하엔", self)
        self.chk_meta_haen.move(350,230)
        self.chk_meta_haen.setFont(font)
        self.chk_meta_lovl = QCheckBox("러블로", self)
        self.chk_meta_lovl.move(350,280)
        self.chk_meta_lovl.setFont(font)
        self.chk_meta_know = QCheckBox("노마셀", self)
        self.chk_meta_know.move(350,330)
        self.chk_meta_know.setFont(font)

        #체크박스 구글
        self.chk_goog_haen = QCheckBox("하엔", self)
        self.chk_goog_haen.move(450,230)
        self.chk_goog_haen.setFont(font)
        self.chk_goog_lovl = QCheckBox("러블로", self)
        self.chk_goog_lovl.move(450,280)
        self.chk_goog_lovl.setFont(font)
        self.chk_goog_know = QCheckBox("노마셀", self)
        self.chk_goog_know.move(450,330)
        self.chk_goog_know.setFont(font)

        #체크박스 스스
        self.chk_ss_haen = QCheckBox("하엔", self)
        self.chk_ss_haen.move(550,230)
        self.chk_ss_haen.setFont(font)
        self.chk_ss_lovl = QCheckBox("러블로", self)
        self.chk_ss_lovl.move(550,280)
        self.chk_ss_lovl.setFont(font)
        self.chk_ss_know = QCheckBox("노마셀", self)
        self.chk_ss_know.move(550,330)
        self.chk_ss_know.setFont(font)


        #체크박스 파컨
        self.chk_pc_haen = QCheckBox("하엔", self)
        self.chk_pc_haen.move(650,230)
        self.chk_pc_haen.setFont(font)
        self.chk_pc_lovl = QCheckBox("러블로", self)
        self.chk_pc_lovl.move(650,280)
        self.chk_pc_lovl.setFont(font)
        self.chk_pc_know = QCheckBox("노마셀", self)
        self.chk_pc_know.move(650,330)
        self.chk_pc_know.setFont(font)


        #체크박스 네이버검색광고
        self.chk_nad_haen = QCheckBox("하엔", self)
        self.chk_nad_haen.move(750,230)
        self.chk_nad_haen.setFont(font)
        self.chk_nad_lovl = QCheckBox("러블로", self)
        self.chk_nad_lovl.move(750,280)
        self.chk_nad_lovl.setFont(font)
        self.chk_nad_know = QCheckBox("노마셀", self)
        self.chk_nad_know.move(750,330)
        self.chk_nad_know.setFont(font)

        #불러오기 체크박스설정
        self.loadCheckboxState()

        #버튼 추출하기
        self.extr_button = QPushButton('추출하기',self)
        self.extr_button.setGeometry(130,502,410,40)
        self.extr_button.clicked.connect(self.extract)
        self.extr_button.setStyleSheet(
            """
            QPushButton {
                background-color: white;
                border-radius: 1.5px;
                border-width: 1px;
                border-color: black;
                border-style: solid;
            }
            QPushButton:hover {
                background-color: rgb(120,120,120);
            }
            QPushButton:pressed {
                background-color: rgb(50, 50, 50);
            }
            """
        )

        #버튼 다운로드폴더
        self.slt_folder = QPushButton('다운로드폴더',self)
        self.slt_folder.setGeometry(330,401,100,29)
        self.slt_folder.clicked.connect(self.folderopen)
        self.slt_folder.setStyleSheet(
            """
            QPushButton {
                background-color: white;
                border-radius: 1.5px;
                border-width: 1px;
                border-color: black;
                border-style: solid;
            }
            QPushButton:hover {
                background-color: rgb(120,120,120);
            }
            QPushButton:pressed {
                background-color: rgb(50, 50, 50);
            }
            """
        )

        #폴더 경로
        self.path_folder = QLineEdit(self)
        self.path_folder.setGeometry(80,401,240,27)
        self.path_folder.setStyleSheet(
                        "background-color: white;"
                        "border-radius: 1.5px;"
                        "border-width: 1px;"
                        "border-color: black;"
                        "border-style: solid;")  # 테두리 스타일 추가
        self.path_folder.setReadOnly(True)

        #크롬 폴더 경로
        self.chrome_path_folder = QLineEdit(self)
        self.chrome_path_folder.setGeometry(80,450,240,27)
        self.chrome_path_folder.setStyleSheet(
                        "background-color: white;"
                        "border-radius: 1.5px;"
                        "border-width: 1px;"
                        "border-color: black;"
                        "border-style: solid;")  # 테두리 스타일 추가
        self.chrome_path_folder.setReadOnly(True)
        self.loadText()

        # 크롬 폴더 버튼
        self.chrome_slt_folder = QPushButton('크롬 폴더',self)
        self.chrome_slt_folder.setGeometry(330,450,100,29)
        self.chrome_slt_folder.clicked.connect(self.chromefolderopen)
        self.chrome_slt_folder.setStyleSheet(
            """
            QPushButton {
                background-color: white;
                border-radius: 1.5px;
                border-width: 1px;
                border-color: black;
                border-style: solid;
            }
            QPushButton:hover {
                background-color: rgb(120,120,120);
            }
            QPushButton:pressed {
                background-color: rgb(50, 50, 50);
            }
            """
        )

        #버튼 설정저장
        self.saveButton = QPushButton('설정저장', self)
        self.saveButton.setGeometry(440,401,100,29)
        self.saveButton.clicked.connect(self.saveText)
        self.saveButton.setStyleSheet(
            """
            QPushButton {
                background-color: white;
                border-radius: 1.5px;
                border-width: 1px;
                border-color: black;
                border-style: solid;
            }
            QPushButton:hover {
                background-color: rgb(120,120,120);
            }
            QPushButton:pressed {
                background-color: rgb(50, 50, 50);
            }
            """
        )

        # 날짜 선택
        self.combo = QComboBox(self)
        self.combo.setGeometry(75, 503, 50, 39)
        self.combo.addItems(["1", "2", "3", "4", "5", "6", "7"])
        self.combo.setFont(QFont('Helvetia', 12, QFont.Bold))

        # 날짜 레이블
        self.daybefore = QLabel("일 전까지", self)
        self.daybefore.move(75, 545)
        self.daybefore.setFont(QFont('Helvetia', 12, QFont.Bold))

    def extract(self):
        try:
            
            # 크롬 On
            ### chromedriver_autoinstaller.install() 사용 추가
            chromedriver_path = chromedriver_autoinstaller.install()
            chrome_options = webdriver.ChromeOptions()
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_argument("--start-maximized") #최대 크기로 시작
            # chrome_options.add_argument('--incognito')
            # chrome_options.add_argument('--window-size=1920,1080')  
            # chrome_options.add_argument('--headless')
            chrome_options.add_experimental_option('detach', True)

            user_data = self.chrome_path_folder.text()
            user_data = 'C:\\Users\\A\\AppData\\Local\\Google\\Chrome\\User Data1'
            chrome_options.add_argument(f"user-data-dir={user_data}")
            chrome_options.add_argument("--profile-directory=Profile 1")

            
            user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36"
            headers = {'user-agent' : user_agent}


            driver = webdriver.Chrome(
                service=Service(chromedriver_path),
                options=chrome_options
            )

            download_folder = self.path_folder.text()


            def count_files(folder):
                """ 폴더 내 파일의 개수를 반환합니다. """
                return len([name for name in os.listdir(folder) if os.path.isfile(os.path.join(folder, name))])

            def get_latest_file(folder):
                """ 폴더 내에서 가장 최신의 파일을 반환합니다. """
                files = [os.path.join(folder, f) for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
                latest_file = max(files, key=os.path.getctime)
                return latest_file
            
            def get_previous_latest_file(folder):
                """폴더 내에서 가장 최신 파일을 제외한 이전 파일을 반환합니다."""
                # 폴더 내의 파일들의 전체 경로와 함께 리스트를 생성합니다.
                files = [os.path.join(folder, f) for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
                
                # 파일이 없다면 None을 반환
                if not files:
                    return None
                
                # 파일들을 생성 시간 기준으로 정렬합니다.
                files.sort(key=os.path.getctime)
                
                # 가장 최신 파일을 제외한 가장 최신 파일을 찾습니다.
                # 파일이 하나만 있는 경우에는 그 파일이 최신 파일이므로, None을 반환합니다.
                if len(files) > 1:
                    previous_latest_file = files[-2]  # 뒤에서 두 번째 항목 선택
                    print(previous_latest_file)
                    return previous_latest_file
                else:
                    return None
                
            def get_nth_latest_file(folder, n):
                """폴더 내에서 n번째로 최신 파일을 반환합니다. n이 1이면 가장 최신, 2면 두 번째로 최신 파일을 반환합니다."""
                # 폴더 내의 파일들의 전체 경로와 함께 리스트를 생성합니다.
                files = [os.path.join(folder, f) for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
                
                # 파일이 없다면 None을 반환
                if not files:
                    return None
                
                # 파일들을 생성 시간 기준으로 정렬합니다.
                files.sort(key=os.path.getctime, reverse=True)
                
                # 요청한 순위의 파일을 반환합니다. n이 파일 수보다 많거나 0 이하인 경우 None을 반환합니다.
                if 1 <= n <= len(files):
                    nth_latest_file = files[n-1]  # n번째 파일 선택
                    print(nth_latest_file)
                    return nth_latest_file
                else:
                    return None

            def check_download():
                # 다운로드 전의 파일 개수 확인
                    initial_file_count = count_files(download_folder)

                    # 다운로드 시작 ...

                    # 새 파일이 다운로드될 때까지 기다림
                    global check
                    check = 0
                    i = 0

                    while i < 20:
                        current_file_count = count_files(download_folder)
                        if current_file_count > initial_file_count:
                            print("A new file has been downloaded.")
                            latest_file = get_latest_file(download_folder)
                            print(f"Downloaded file: {latest_file}")
                            # 여기서 필요한 작업을 수행하세요, 예를 들면 파일 열기 등
                            check = 1
                            break
                        else:
                            print("Still no new file")
                        time.sleep(0.3)  # 폴더 상태를 0.3초마다 체크
                        i += 1
                    return check

            def check_data_in_second_row(file_path):
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active
                second_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
                if second_row and any(cell is not None for cell in second_row[0]):
                    return True
                return False
            
            def convert_data(data):
                result = []
                for item in data:
                    if isinstance(item, str) and '%' in item:
                        result.append(float(item.strip('%')) / 100)
                    elif isinstance(item, str) and ',' in item:
                        result.append(int(item.replace(',', '')))
                    elif item.isdigit():
                        result.append(int(item))
                    else:
                        result.append(item)
                return result
            
            def retry_operation(operation, max_retries=5, delay=30):
                retries = 0
                while retries < max_retries:
                    try:
                        return operation()
                    except APIError as e:
                        retries += 1
                        if retries >= max_retries:
                            raise
                        print(f"오류 발생: {e}. {delay}초 후에 재시도합니다...")
                        time.sleep(delay)
            
            
            target_days_input = int(self.combo.currentText())
            # 날짜 구하기
            today = date.today()
            # 하루를 나타내는 timedelta 객체 생성
            one_day = datetime.timedelta(days=1)
            # 어제 날짜를 구함
            yesterday = today - one_day
            formatted_date = yesterday.strftime("%Y-%m-%d")
            
            ######################                         어제 추출                            ####################
            ######################                         어제 추출                            ####################
            ######################                         어제 추출                            ####################
            
            if target_days_input == 1:

                # 날짜 구하기
                today = datetime.date.today()
                # 하루를 나타내는 timedelta 객체 생성
                # 어제 날짜를 구함

                today_date = today.strftime("%d")
                today_month = str(int(today.strftime("%m")))

                weekday_korean = {
                    0: '월',
                    1: '화',
                    2: '수',
                    3: '목',
                    4: '금',
                    5: '토',
                    6: '일'
                }
                target_days = 1
                dayx = datetime.timedelta(days=target_days)
                day1 = datetime.timedelta(days=1)

                # 오늘 날짜 구하기
                today_yday = today-day1
                today_tday = today-dayx

                #카페24
                def cafe24(url_cafe24, url_cafe24_req, cafe24_id, cafe24_pw, sheet_urlR, sheet_nameR, sheet_nameD, brand):
                    
                    driver.get(url_cafe24)

                    ##################################### 로그인
                    ##################################### 로그인
                    ##################################### 로그인
                    ##################################### 로그인

                    # ID
                    input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#mall_id")))
                    input_field.click()
                    time.sleep(1)
                    input_field.send_keys(Keys.CONTROL + "a")
                    input_field.send_keys(Keys.BACKSPACE)
                    driver.find_element(By.CSS_SELECTOR, "#mall_id").send_keys(cafe24_id)

                    # PW
                    input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#userpasswd")))
                    input_field.click()
                    input_field.send_keys(Keys.CONTROL + "a")
                    input_field.send_keys(Keys.BACKSPACE)
                    driver.find_element(By.CSS_SELECTOR, "#userpasswd").send_keys(cafe24_pw)

                    # 로그인클릭
                    driver.find_element(By.CSS_SELECTOR,'#frm_user > div > div.mButton > button').click()

                    #비밀번호변경안내
                    try: WebDriverWait(driver, 5).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#iptBtnEm")))).click() 
                    except: pass

                    # 카페24 팝업, 새소식 알림 발생 시 변수 jump 코드 적용(화면로딩대기 코드 변경 -> 오늘의 할 일 contains(text()) 코드)
                    try:
                        time.sleep(3)
                        popup = driver.find_element(By.XPATH, '//*[contains(text(), "오늘 하루 보지 않기")]')
                        popup.click()

                    except: pass

                    #화면로딩대기
                    WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, '//*[contains(text(), "오늘의 할 일")]')))
                    

                    # 데이터 접근
                    driver.get(url_cafe24_req)

                    # 자세히보기클릭
                    element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#QA_day3 > div.mBoard.gScroll > table")))
                    driver.execute_script("arguments[0].scrollIntoView(true);", element) # 스크롤다운
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#sReportGabView"))).click() 
                    
                    element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#QA_day3 > div.mBoard.gScroll > table")))
                    driver.execute_script("arguments[0].scrollIntoView(true);", element) # 스크롤다운


                    #################################################### 데이터 처리
                    #################################################### 데이터 처리
                    #################################################### 데이터 처리
                    #################################################### 데이터 처리


                    # 날짜 구하기
                    today = datetime.date.today()
                    # 하루를 나타내는 timedelta 객체 생성
                    # 어제 날짜를 구함

                    today_date = today.strftime("%d")
                    today_month = str(int(today.strftime("%m")))

                    weekday_korean = {
                        0: '월',
                        1: '화',
                        2: '수',
                        3: '목',
                        4: '금',
                        5: '토',
                        6: '일'
                    }
                    target_days = 1
                    dayx = datetime.timedelta(days=target_days)
                    day1 = datetime.timedelta(days=1)

                    # 오늘 날짜 구하기
                    today_yday = today-day1
                    today_tday = today-dayx

                    weekday_num = today.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)
                    weekday_numy = today_yday.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)
                    weekday_numt = today_tday.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)
                    # 요일을 한국어로 변환
                    weekday_kr = weekday_korean[weekday_num]
                    weekday_kry = weekday_korean[weekday_numy]
                    weekday_krt = weekday_korean[weekday_numt]

                    weekday = f"{today}({weekday_kr})"
                    weekday_y = f"{today_yday}({weekday_kry})"
                    weekday_t = f"{today_tday}({weekday_krt})"

                    rows = driver.find_elements(By.CSS_SELECTOR, 'tbody.right tr')

                    cover = []
                    cover0 = []
                    for element in rows:
                        new_data_list = []
                        rawdata = element.text
                        # 문자열을 공백을 기준으로 분리하여 리스트로 변환
                        data_list = rawdata.split()
                        data_list = [x.replace(',', '') for x in data_list]

                        for items in data_list:

                        # 숫자인 경우 숫자로 변환
                            try:
                                numeric_value = int(items)
                                new_data_list.append(numeric_value)
                            except:
                                # 숫자가 아닌 경우 원래 값 유지
                                new_data_list.append(items)
                        cover.append(new_data_list[1:])
                        cover0.append(new_data_list[0])

                    # 데이터가 비어있을 경우
                    today_ydayTemp = yesterday.strftime(f"%Y-%m-%d({weekday_kry})")
                    print(today_ydayTemp)
                    print("check")
                    print(cover0)

                    # 서비스 계정 키 파일 경로
                    credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                    # gspread 클라이언트 초기화
                    client = gspread.service_account(filename=credential_file)

                    # Google 시트 열기
                    spreadsheet = client.open_by_url(sheet_urlR)

                    # 첫 번째 시트 선택
                    sheet = spreadsheet.worksheet(sheet_nameR)

                    last_row = len(sheet.col_values(3))
                    next_row = last_row + 1
                    print(last_row)
                    print(next_row)
                        
                    if today_ydayTemp in cover0:
                        keynum = cover0.index(today_ydayTemp)

                        data_to_paste = cover[keynum]
                        # print(cover0)
                        # print(weekday_t)

                        # t=cover0.index(weekday_t)
                        # y=cover0.index(weekday_y)

                        # data_to_paste = cover[y:t+1]
                        # print(data_to_paste)

                        data1 = data_to_paste[:9]
                        data2 = data_to_paste[9]
                        data3 = data_to_paste[10:]
                    # 카페24 R, 데이터 없으면 0 입력 되도록 코드 수정
                    else:
                        data1 = [0, 0, 0, 0, 0, 0, 0, 0, 0]
                        data2 = 0
                        data3 = [0, 0, 0]

                    print(data1)
                    print(data2)
                    print(data3)

                    range1 = f'C{next_row}:K{next_row}'
                    range2 = f'M{next_row}'
                    range3 = f'O{next_row}:Q{next_row}'
                    
                    sheet.update([data1], range1)
                    sheet.update([[data2]], range2)
                    sheet.update([data3], range3)



                    time.sleep(1)

                    ###################################### 조회수
                    ###################################### 조회수
                    ###################################### 조회수
                    ###################################### 조회수

                    driver.find_element(By.CSS_SELECTOR, "#mCSB_2_container > ul:nth-child(1) > li:nth-child(9)").click() #통계 클릭
                    WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#mCSB_3_container > div.depthList > ul > li:nth-child(8)"))).click() #접속통계클릭

                    #새 창 대기
                    current_window_handle = driver.current_window_handle

                    new_window_handle = None
                    while not new_window_handle:
                        for handle in driver.window_handles:
                            if handle != current_window_handle:
                                new_window_handle = handle
                                break


                    #팝업으로 제어 변경
                    driver.switch_to.window(driver.window_handles[1])

                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#rpt_pth")))).click() #방문경로분석
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#rpt_pth_dod")))).click() #방문도메인(상세)
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#body_center > table:nth-child(5) > tbody > tr:nth-child(2) > td:nth-child(3) > a:nth-child(2) > img")))).click()

                    # 데이터 처리
                    visitors = driver.find_elements(By.ID, "summary_pfm_total")
                    for num in visitors:
                        the_num = driver.find_element(By.CSS_SELECTOR, "#summary_pfm_total > td:nth-child(2)").text
                        print(the_num)

                    # 서비스 계정 키 파일 경로
                    credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                    # gspread 클라이언트 초기화
                    client = gspread.service_account(filename=credential_file)

                    # Google 시트 열기
                    spreadsheet = client.open_by_url(sheet_urlR)

                    # 첫 번째 시트 선택
                    sheet = spreadsheet.worksheet(sheet_nameD)
                    today = datetime.date.today().strftime("%Y-%m-%d")
                    column_values = sheet.col_values(1)
                    for idx, cell_value in enumerate(column_values, start=1):  # start=1로 설정하여 행 번호를 1부터 시작
                        if cell_value == today:
                            print(cell_value)
                            print(gspread.utils.rowcol_to_a1(idx, 1))
                            cell_addr = gspread.utils.rowcol_to_a1(idx, 1)
                            # return f"{gspread.utils.rowcol_to_a1(idx, 1)}"  # 셀 주소 반환
                        
                    (start_row, start_col) = gspread.utils.a1_to_rowcol(cell_addr)

                    # Google 시트에 데이터 쓰기
                    numeric_value = int(the_num.replace(',', ''))
                    range_to_write = f'C{start_row-1}'
                    sheet.update([[numeric_value]], range_to_write)

                    driver.close()
                    driver.switch_to.window(driver.window_handles[0]) 


                #카페24 하엔
                if self.chk_cafe_haen.isChecked() == True:

                    label = self.chk_cafe_haen
                    url_cafe24 = "https://eclogin.cafe24.com/Shop/"
                    url_cafe24_req_haen = "https://woo8425.cafe24.com/disp/admin/shop1/report/DailyList"
                    
                    cafe24_id_haen = self.login_info("CAFE_HAEN_ID")
                    cafe24_pw_haen = self.login_info("CAFE_HAEN_PW")

                    sheet_haenR_url = 'https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=1894651086'
                    sheet_haenR = '하엔R'
                    sheet_haenD = "하엔D"
                    brand = "하엔"
                
                    cafe24(url_cafe24, url_cafe24_req_haen, cafe24_id_haen, cafe24_pw_haen, sheet_haenR_url, sheet_haenR, sheet_haenD, brand)

                #카페24 러블로
                if self.chk_cafe_lovl.isChecked() == True:

                    url_cafe24 = "https://eclogin.cafe24.com/Shop/" 
                    url_cafe24_req_lovelo = "https://wooo8425.cafe24.com/disp/admin/shop1/report/DailyList"

                    cafe24_id_lovelo = self.login_info("CAFE_LOVE_ID")
                    cafe24_pw_lovelo = self.login_info("CAFE_LOVE_PW")

                    sheet_loveR_url = 'https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=872830966'
                    sheet_loveR = '러블로R'
                    sheet_loveD = "러블로D"
                    brand = "러블로"

                    cafe24(url_cafe24, url_cafe24_req_lovelo, cafe24_id_lovelo, cafe24_pw_lovelo, sheet_loveR_url, sheet_loveR, sheet_loveD, brand)

                #카페24 노마셀
                if self.chk_cafe_know.isChecked() == True:

                    url_cafe24 = "https://eclogin.cafe24.com/Shop/" 
                    url_cafe24_req_knowmycell = "https://fkark12.cafe24.com/disp/admin/shop1/report/DailyList"

                    cafe24_id_knowmycell = self.login_info("CAFE_KNOW_ID")
                    cafe24_pw_knowmycell = self.login_info("CAFE_KNOW_PW")

                    sheet_knowR_url = 'https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=567505346'
                    sheet_knowR = '노마셀R'
                    sheet_knowD = "노마셀D"
                    brand = "노마셀"

                    cafe24(url_cafe24, url_cafe24_req_knowmycell, cafe24_id_knowmycell, cafe24_pw_knowmycell, sheet_knowR_url, sheet_knowR, sheet_knowD, brand)


                #카페24 제니크
                if self.chk_cafe_ZQ.isChecked() == True:

                    url_cafe24 = "https://eclogin.cafe24.com/Shop/" 
                    url_cafe24_req_ZQ = "https://fkark08.cafe24.com/disp/admin/shop1/report/DailyList"

                    cafe24_id_ZQ = self.login_info("CAFE_ZQ_ID")
                    cafe24_pw_ZQ = self.login_info("CAFE_ZQ_PW")

                    sheet_ZQR_url = 'https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=567505346'
                    sheet_ZQR = '제니크R'
                    sheet_ZQD = "제니크D"
                    brand = "제니크"

                    cafe24(url_cafe24, url_cafe24_req_ZQ, cafe24_id_ZQ, cafe24_pw_ZQ, sheet_ZQR_url, sheet_ZQR, sheet_ZQD, brand)

                def power(url, url2,  id, pw, sheetUrl, sheetName, key, key2, brand):

                    # 서비스 계정 키 파일 경로
                    credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                    # gspread 클라이언트 초기화
                    client = gspread.service_account(filename=credential_file)

                    # Google 시트 열기
                    spreadsheet = client.open_by_url(sheetUrl)

                    # 첫 번째 시트 선택
                    sheet = spreadsheet.worksheet(sheetName)

                    driver.get(url)

                    ##################################### 로그인
                    ##################################### 로그인
                    ##################################### 로그인
                    ##################################### 로그인

                    # ID
                    input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#mall_id")))
                    input_field.click()
                    time.sleep(1)
                    input_field.send_keys(Keys.CONTROL + "a")
                    input_field.send_keys(Keys.BACKSPACE)
                    driver.find_element(By.CSS_SELECTOR, "#mall_id").send_keys(id)

                    # PW
                    input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#userpasswd")))
                    input_field.click()
                    input_field.send_keys(Keys.CONTROL + "a")
                    input_field.send_keys(Keys.BACKSPACE)
                    driver.find_element(By.CSS_SELECTOR, "#userpasswd").send_keys(pw)

                    # 로그인클릭
                    driver.find_element(By.CSS_SELECTOR,'#frm_user > div > div.mButton > button').click()

                    #비밀번호변경안내
                    try: WebDriverWait(driver, 5).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#iptBtnEm")))).click() 
                    except: pass

                    #화면로딩대기
                    # 파컨 화면 로딩대기 로직 수정
                    WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, '//*[contains(text(), "오늘의 할 일")]')))


                    # 10. 카페24(파컨) 홍보용 팝업창 대비, 다른 페이지 이동 후 통계 팝업 창 띄우도록 변경
                    driver.get(url2)
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#mCSB_2_container > ul:nth-child(1) > li:nth-child(9)"))).click() #통계 클릭
                    # driver.find_element(By.CSS_SELECTOR, "#mCSB_2_container > ul:nth-child(1) > li:nth-child(9)").click() #통계 클릭
                    WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#mCSB_3_container > div.depthList > ul > li:nth-child(8)"))).click() #접속통계클릭

                    #새 창 대기
                    current_window_handle = driver.current_window_handle

                    new_window_handle = None
                    while not new_window_handle:
                        for handle in driver.window_handles:
                            if handle != current_window_handle:
                                new_window_handle = handle
                                break


                    #팝업으로 제어 변경
                    driver.switch_to.window(driver.window_handles[1]) 

                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#rpt_pth")))).click() #방문경로분석
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#rpt_pth_dod")))).click() #방문도메인(상세)

                    keywords = [key, key2]
                    # 검색어 입력(NV, NPO, GS)
                    for item in keywords:
                        print(item, "검색")
                        search = driver.find_element(By.CSS_SELECTOR, "#body_center > table:nth-child(13) > tbody > tr > td:nth-child(1) > input")
                        search.send_keys(Keys.CONTROL + "a")
                        search.send_keys(Keys.BACKSPACE)
                        search.click()
                        search.send_keys(item)
                        driver.find_element(By.CSS_SELECTOR, "#body_center > table:nth-child(13) > tbody > tr > td:nth-child(1) > a > img").click()

                        # 어제 클릭
                        WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#body_center > table:nth-child(5) > tbody > tr:nth-child(2) > td:nth-child(3) > a:nth-child(2) > img")))).click()

                        for i in range(1,30):
                            try:
                                line = driver.find_element(By.CSS_SELECTOR, f"#detail_pfm_total > tr:nth-child({i})").text
                                lineSplit = line.strip().split(" ")
                                print(lineSplit)

                                last_row = len(sheet.get_all_values())
                                print(last_row)
                                next_row = last_row + 1  # 다음 행 번호

                                def convert_data(data):
                                    result = []
                                    for item in data:
                                        if isinstance(item, str) and '%' in item:
                                            result.append(float(item.strip('%')) / 100)
                                        elif isinstance(item, str) and ',' in item:
                                            result.append(int(item.replace(',', '')))
                                        elif item.isdigit():
                                            result.append(int(item))
                                        else:
                                            result.append(item)
                                    return result

                                # 입력할 데이터
                                converted_data = convert_data(lineSplit)

                                # 구글 시트에서 퍼센트 형식으로 변경하기 위해 셀 범위를 지정
                                # 예시에서는 C1, E1 셀을 퍼센트 형식으로 설정
                                sheet.format(f"D{next_row}", {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}})
                                sheet.format(f"F{next_row}", {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}})
                                    
                                sheet.update([[str(today_yday)]], f"A{next_row}")
                                range_to_write = f'B{next_row}:I{next_row}'
                                sheet.update([converted_data], range_to_write)
                                print("OK")
                            except:
                                if i == 1:
                                    dummyData = [item, '0', '0', '0', '0', '0', '0', '0']

                                    last_row = len(sheet.get_all_values())
                                    print(last_row)
                                    next_row = last_row + 1  # 다음 행 번호

                                    def convert_data(data):
                                        result = []
                                        for item in data:
                                            if isinstance(item, str) and '%' in item:
                                                result.append(float(item.strip('%')) / 100)
                                            elif isinstance(item, str) and ',' in item:
                                                result.append(int(item.replace(',', '')))
                                            elif item.isdigit():
                                                result.append(int(item))
                                            else:
                                                result.append(item)
                                        return result

                                    # 입력할 데이터
                                    converted_data = convert_data(dummyData)

                                    # 구글 시트에서 퍼센트 형식으로 변경하기 위해 셀 범위를 지정
                                    # 예시에서는 C1, E1 셀을 퍼센트 형식으로 설정
                                    sheet.format(f"D{next_row}", {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}})
                                    sheet.format(f"F{next_row}", {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}})
                                        
                                    sheet.update([[str(today_yday)]], f"A{next_row}")
                                    range_to_write = f'B{next_row}:I{next_row}'
                                    sheet.update([converted_data], range_to_write)
                                    print("OK")
                                    break
                    driver.close()
                    time.sleep(0.1)
                    driver.switch_to.window(driver.window_handles[0])



                url_cafe24 = "https://eclogin.cafe24.com/Shop/" 
                
                if self.chk_pc_haen.isChecked() == True:
                    cafe24_id_haen = self.login_info("CAFE_HAEN_ID")
                    cafe24_pw_haen = self.login_info("CAFE_HAEN_PW")
                    url2 = "https://woo8425.cafe24.com/disp/admin/shop1/report/DailyList"

                    sheetUrl_haen = "https://docs.google.com/spreadsheets/d/1V8b3FRe_8witwHXQceekgm-BAvTQLwwkcyuaW-mIi30/edit#gid=128334801"
                    sheetName_haenPCR = "하엔 파워콘텐츠 R"
                    Keyword = "NPO"
                    Keyword2 = "GS"
                    brand = "하엔"

                    power(url_cafe24, url2, cafe24_id_haen, cafe24_pw_haen, sheetUrl_haen, sheetName_haenPCR, Keyword, Keyword2, brand)


                if self.chk_pc_lovl.isChecked() == True:
                    cafe24_id_love = self.login_info("CAFE_LOVE_ID")
                    cafe24_pw_love = self.login_info("CAFE_LOVE_PW")
                    url2 = "https://wooo8425.cafe24.com/disp/admin/shop1/report/DailyList"

                    sheetUrl_love = "https://docs.google.com/spreadsheets/d/1NVnVJsCj0Ap_o2xabua9ANUw_1IUREVMJKteY_O1yks/edit#gid=311448069"
                    sheetName_lovePCR = "러블로 파워콘텐츠 R"
                    Keyword = "NV"
                    Keyword2 = "GS"
                    brand = "러블로"

                    power(url_cafe24, url2, cafe24_id_love, cafe24_pw_love, sheetUrl_love, sheetName_lovePCR, Keyword, Keyword2, brand)


                if self.chk_pc_know.isChecked() == True:
                    cafe24_id_know = self.login_info("CAFE_KNOW_ID")
                    cafe24_pw_know = self.login_info("CAFE_KNOW_PW")
                    url2 = "https://fkark12.cafe24.com/disp/admin/shop1/report/DailyList"

                    sheetUrl_know = "https://docs.google.com/spreadsheets/d/12FWmZMuznsxOY_IDbBWeSis-EW1Ds1f9TB6X7K6TFBc/edit#gid=1722856727"
                    sheetName_knowPCR = "노마셀 파워콘텐츠 R"
                    Keyword = "NV"
                    Keyword2 = "GS"
                    brand = "노마셀"

                    power(url_cafe24, url2, cafe24_id_know, cafe24_pw_know, sheetUrl_know, sheetName_knowPCR, Keyword, Keyword2, brand)


                #########쿠팡로데이터##########
                def coupang_rawdata(sheet_url, sheet_nameR, options, sheet_nameC):

                    xlsx_file = get_previous_latest_file(download_folder)

                    df_uploaded_new = pd.read_excel(xlsx_file)
                    # '러브슬라임'이라는 단어가 포함된 모든 행을 '옵션명' 열을 기준으로 필터링합니다.
                    filtered_rows_with_loveslime = df_uploaded_new[df_uploaded_new['입찰유형'].astype(str).str.contains("cpc")]

                    # 필터링된 행들의 데이터를 리스트로 변환합니다.
                    rows_list_with_loveslime = filtered_rows_with_loveslime.values.tolist()
                    updated_data_list = []
                    for row in rows_list_with_loveslime:
                        new_row = row.copy()  # 원본 데이터의 복사본 생성
                        if len(row) > 1:  # 두 번째 값이 존재하는지 확인
                            new_row[1] = str(row[1])  # 두 번째 값을 정수형으로 변환 후 문자열로 변환
                        updated_data_list.append(new_row)

                    print(updated_data_list)
                
                
                    # 서비스 계정 키 파일 경로
                    credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                    # gspread 클라이언트 초기화
                    client = gspread.service_account(filename=credential_file)

                    # Google 시트 열기
                    spreadsheet = client.open_by_url(sheet_url)

                    # 첫 번째 시트 선택
                    sheet = spreadsheet.worksheet(sheet_nameR)

                    last_row = len(sheet.get_all_values())
                    print(last_row)
                    next_row = last_row + 1  # 다음 행 번호

                    # 날짜 구하기
                    today = datetime.date.today()
                    # 하루를 나타내는 timedelta 객체 생성
                    one_day = datetime.timedelta(days=1)
                    # 어제 날짜를 구함
                    yesterday = today - one_day

                    formatted_date = yesterday.strftime("%Y-%m-%d")
                    # Google 시트에 데이터 쓰기

                    if len(updated_data_list) > 1:
                        i = 0
                        while i < len(updated_data_list):
                            print((updated_data_list[i])[1:])
                            range_to_write = f'B{next_row+i}:AI{next_row+i}'
                            sheet.update([(updated_data_list[i])[1:-1]], range_to_write)
                            sheet.update([[formatted_date]], f'A{next_row+i}')
                            i += 1
                    else:
                        range_to_write = f'B{next_row}:AI{next_row}'
                        sheet.update([(updated_data_list[0])[1:-1]], range_to_write)
                        sheet.update([[formatted_date]], f'A{next_row}')


                    ###                                                         쿠팡 C                                                  ###
                    ###                                                         쿠팡 C                                                  ###
                    ###                                                         쿠팡 C                                                  ###

                    xlsx_file = get_latest_file(download_folder)


                    df_uploaded_new = pd.read_excel(xlsx_file)
                    # '러브슬라임'이라는 단어가 포함된 모든 행을 '옵션명' 열을 기준으로 필터링합니다.

                    filtered_rows_with_loveslime = df_uploaded_new[df_uploaded_new['옵션명'].astype(str).str.contains(options)]
                    print(filtered_rows_with_loveslime)

                    # 4. 쿠팡C 데이터 없으면 0이 입력되도록 코드수정
                    if filtered_rows_with_loveslime.empty:
                        filtered_rows_with_loveslime = pd.DataFrame([["-", "-", "-", "-", "-", 0, 0, 0, 0, 0, 0, 0, 0]],
                                                    columns=df_uploaded_new.columns)
                

                    # 필터링된 행들의 데이터를 리스트로 변환합니다.
                    rows_list_with_loveslime = filtered_rows_with_loveslime.values.tolist()

                    # 숫자인지 확인하는 함수
                    def is_number(s):
                        try:
                            float(s)
                            return True
                        except ValueError:
                            return False

                    # 두 번째 값만 숫자로 변환한 후 문자열로 변환하여 업데이트하는 과정
                    updated_data_list = []
                    for row in rows_list_with_loveslime:
                        new_row = row.copy()  # 원본 데이터의 복사본 생성
                        if len(row) > 1 and is_number(row[1]):  # 두 번째 값이 존재하고 숫자인지 확인
                            new_row[1] = str(float(row[1]))  # 두 번째 값을 float으로 변환 후 문자열로 변환
                        else:
                            new_row[1] = "-"  # 숫자가 아닌 경우 "-"로 설정
                        updated_data_list.append(new_row)


                    # # 두 번째 값만 정수형으로 변환한 후 문자열로 변환하여 업데이트하는 과정
                    # updated_data_list = []
                    # for row in rows_list_with_loveslime:
                    #     new_row = row.copy()  # 원본 데이터의 복사본 생성
                    #     if len(row) > 1:  # 두 번째 값이 존재하는지 확인
                    #         new_row[1] = str(int(row[1]))  # 두 번째 값을 정수형으로 변환 후 문자열로 변환
                    #     updated_data_list.append(new_row)


                

                    # 결과 출력
                    print(updated_data_list)


                    # 서비스 계정 키 파일 경로
                    credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                    # gspread 클라이언트 초기화
                    client = gspread.service_account(filename=credential_file)

                    # Google 시트 열기
                    spreadsheet = client.open_by_url("https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=374561563")

                    # 첫 번째 시트 선택
                    sheet = spreadsheet.worksheet(sheet_nameC)

                    last_row = len(sheet.get_all_values())
                    print(last_row)
                    next_row = last_row + 1  # 다음 행 번호

                    # 날짜 구하기
                    today = datetime.date.today()
                    # 하루를 나타내는 timedelta 객체 생성
                    one_day = datetime.timedelta(days=1)
                    # 어제 날짜를 구함
                    yesterday = today - one_day

                    formatted_date = yesterday.strftime("%Y-%m-%d")
                        
                    # Google 시트에 데이터 쓰기
                    if len(updated_data_list) > 1:
                        i = 0
                        while i < len(updated_data_list):
                            range_to_write = f'B{next_row+i}:N{next_row+i}'
                            sheet.update([updated_data_list[i]], range_to_write)
                            sheet.update([[formatted_date]], f'A{next_row+i}')
                            i += 1
                    else:
                        range_to_write = f'B{next_row}:N{next_row}'
                        sheet.update([updated_data_list[0]], range_to_write)
                        sheet.update([[formatted_date]], f'A{next_row}')

                # 쿠팡
                def coupang(url_coupang_daily, coupang_id, coupang_pw, coupC_url):
                    try:
                        driver.get(url_coupang_daily)  # 로그인 시작
                        if driver.find_element(By.CSS_SELECTOR, "body > pre"):
                            driver.get(url_coupang_daily)  # 요소가 존재하면 페이지를 다시 로드
                    except NoSuchElementException:
                        # 요소가 없을 때 처리할 로직
                        pass

                    try:
                        driver.get(url_coupang_daily)  # 로그인 시작
                        if driver.find_element(By.CSS_SELECTOR, "body > h1"):
                            driver.get(url_coupang_daily)  # 요소가 존재하면 페이지를 다시 로드
                    except NoSuchElementException:
                        pass
                    
                    # 9. 쿠팡 로그인창으로 가는 선택자가 바뀌어, 다른 로직의 이동방법 적용
                    try:
                        loginElements = driver.find_elements(By.XPATH, '//*[contains(text(), "로그인하기")]')

                        if len(loginElements) > 1:  # 요소가 두 개 이상 있는지 확인
                            
                            loginElements[0].click()
                        # if driver.find_element(By.CSS_SELECTOR, "#main-container > div > div.sc-30ec2de1-0.cZrQsU > ul > li:nth-child(1) > a > span"):
                        #     driver.find_element(By.CSS_SELECTOR, "#main-container > div > div.sc-30ec2de1-0.cZrQsU > ul > li:nth-child(1) > a > span").click()

                    except NoSuchElementException:
                        # 요소가 없을 때 처리할 로직
                        pass
                            

                    input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#username")))
                    try:
                        input_field.click()

                    except:
                        driver.get("https://xauth.coupang.com/auth/realms/seller/protocol/openid-connect/auth?client_id=wing&scope=openid%20phone%20web-origins%20profile%20roles%20address%20microprofile-jwt%20email&response_type=code&redirect_uri=https%3A%2F%2Fadvertising.coupang.com%2Fuser%2Fwing%2Fauthorization-callback&state=SDlDgC-VXCFPEL7w0O0y29Th1gpa3u_zYmU0BSu8a3A&code_challenge=C20glxOs1oHF9XNibHiActQzd7jqe7q-MPH820V9CQI&code_challenge_method=S256")
                        
                        input_field.click()


                    print(coupang_id)
                    print(coupang_pw)
                    time.sleep(0.7)
                    input_field.send_keys(Keys.CONTROL + "a")
                    input_field.send_keys(Keys.BACKSPACE)
                    driver.find_element(By.CSS_SELECTOR, "#username").send_keys(coupang_id)
                    input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#password")))
                    input_field.click()
                    input_field.send_keys(Keys.CONTROL + "a")
                    input_field.send_keys(Keys.BACKSPACE)
                    driver.find_element(By.CSS_SELECTOR, "#password").send_keys(coupang_pw)
                    driver.find_element(By.CSS_SELECTOR,'#kc-login').click()

                    # 로그인 오류 발생하면 재시도
                    ### 비밀번호 오류 예외문
                    cnt = 0
                    while cnt < 4:
                        try:
                            loginErrorMessage = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, '//*[contains(text(), "아이디 또는 비밀번호가 다릅니다.")]')))
                            if loginErrorMessage:
                                input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#username")))
                                input_field.click()
                                time.sleep(0.7)
                                input_field.send_keys(Keys.CONTROL + "a")
                                input_field.send_keys(Keys.BACKSPACE)
                                driver.find_element(By.CSS_SELECTOR, "#username").send_keys(coupang_id)
                                input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#password")))
                                input_field.click()
                                input_field.send_keys(Keys.CONTROL + "a")
                                input_field.send_keys(Keys.BACKSPACE)
                                driver.find_element(By.CSS_SELECTOR, "#password").send_keys(coupang_pw)
                                driver.find_element(By.CSS_SELECTOR,'#kc-login').click()
                            
                            cnt += 1
                            print("쿠팡 로그인에러... 재시도...", cnt)
                            

                        except: break
                            
                    try:
                        WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#endDateId"))) #클릭 종료일
                    except:
                        driver.refresh()
                        WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#endDateId"))) #클릭 종료일

                    driver.find_element(By.CSS_SELECTOR, "#endDateId").click()
                    time.sleep(0.3)
                    driver.find_element(By.XPATH, '//*[@id="ad-reporting-app"]/div[2]/div/div[2]/div[1]/div[2]/div[3]/div[2]/div[2]/div[1]/div/div/div[2]/div/div/div[2]/div/span[2]/button').click() #어제 클릭
                    time.sleep(0.3)
# 쿠팡 보고서 -> 어제 클릭 -> 새로고침 과정에서 새로고침 선택 안되는 문제(CSS요소 변동) 수정 // 텍스트로 새로고침 인식하게 함
                    driver.find_element(By.XPATH, '//*[contains(text(), "목록 새로 고침")]').click()
                    time.sleep(0.3)

                    element = driver.find_element(By.CSS_SELECTOR, '#ad-reporting-app > div.self-service-ad-reporting-ui > div > div.sc-11l2gxs-0.fcpsUc > div.sc-ipia07-0.iCqAxH > div.panel-options > div.sc-19odvm9-0.kgfJLF > div.select-date-group')#기간 구분
                    element.click() 
                    ActionChains(driver).move_to_element_with_offset(element,5,75).click().perform() #클릭 일별
                    time.sleep(0.3)

                    driver.find_element(By.CSS_SELECTOR,'#ad-reporting-app > div.self-service-ad-reporting-ui > div > div.sc-11l2gxs-0.fcpsUc > div.sc-ipia07-0.iCqAxH > div.panel-options > div.sc-1jpf51e-0.hSjByk > div > div.campaign-picker-container > div > button > span.text').click() #캠페인 선택
                    time.sleep(0.3)
                    checkbox = driver.find_element(By.CSS_SELECTOR, "body > div.sc-1jpf51e-1.jljGiJ.popper > div > div.select-all-campaigns > label > span.ant-checkbox > input[type='checkbox']")

                    # 체크박스가 체크되어 있지 않다면 클릭하여 체크합니다.
                    if not checkbox.is_selected():
                        checkbox.click()  
                    driver.find_element(By.CSS_SELECTOR, "body > div.sc-1jpf51e-1.jljGiJ.popper > div > div.button-container > button.ant-btn.ant-btn-primary > span").click()
                    time.sleep(0.3)

### 보고서 생성 실패 시 한 번 더 시도
                    try:
                        # 보고서 생성
                        driver.find_element(By.CSS_SELECTOR, "#generateReport > span").click() 
                        time.sleep(5)

                        if driver.find_element(By.CSS_SELECTOR, "#rc-tabs-0-panel-requestedReport > div > div.react-grid-Container > div > div > div:nth-child(2) > div > div > div:nth-child(2) > div:nth-child(1) > div:nth-child(5) > div > div > span > div").text == "생성 실패":
                            driver.find_element(By.CSS_SELECTOR, "#generateReport > span").click() 
                            time.sleep(5)

                        # 보고서 다운로드
                        element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#rc-tabs-0-panel-requestedReport > div > div.react-grid-Container > div > div > div:nth-child(2) > div > div > div:nth-child(2) > div:nth-child(1) > div:nth-child(6) > div > div > span > div > div:nth-child(2) > button > span"))) 

                        # 다운로드 확인
                        cnt = 1
                        while cnt < 10:
                            current_file_count1 = count_files(download_folder)
                            element.click()
                            time.sleep(3)
                            current_file_count2 = count_files(download_folder)
                            if current_file_count1 != current_file_count2:
                                break

                            cnt += 1

                    except:

                        driver.get(url_coupang_daily)

                        WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#endDateId"))) #클릭 종료일 
                        driver.find_element(By.CSS_SELECTOR, "#endDateId").click()
                        time.sleep(0.3)
                        driver.find_element(By.XPATH, '//*[@id="ad-reporting-app"]/div[2]/div/div[2]/div[1]/div[2]/div[3]/div[2]/div[2]/div[1]/div/div/div[2]/div/div/div[2]/div/span[2]/button').click() #어제 클릭
                        time.sleep(0.3)
    # 쿠팡 보고서 -> 어제 클릭 -> 새로고침 과정에서 새로고침 선택 안되는 문제(CSS요소 변동) 수정 // 텍스트로 새로고침 인식하게 함
                        driver.find_element(By.XPATH, '//*[contains(text(), "목록 새로 고침")]').click()
                        time.sleep(0.3)

                        element = driver.find_element(By.CSS_SELECTOR, '#ad-reporting-app > div.self-service-ad-reporting-ui > div > div.sc-11l2gxs-0.fcpsUc > div.sc-ipia07-0.iCqAxH > div.panel-options > div.sc-19odvm9-0.kgfJLF > div.select-date-group')#기간 구분
                        element.click() 
                        ActionChains(driver).move_to_element_with_offset(element,5,75).click().perform() #클릭 일별
                        time.sleep(0.3)

                        driver.find_element(By.CSS_SELECTOR,'#ad-reporting-app > div.self-service-ad-reporting-ui > div > div.sc-11l2gxs-0.fcpsUc > div.sc-ipia07-0.iCqAxH > div.panel-options > div.sc-1jpf51e-0.hSjByk > div > div.campaign-picker-container > div > button > span.text').click() #캠페인 선택
                        time.sleep(0.3)
                        checkbox = driver.find_element(By.CSS_SELECTOR, "body > div.sc-1jpf51e-1.jljGiJ.popper > div > div.select-all-campaigns > label > span.ant-checkbox > input[type='checkbox']")

                        # 체크박스가 체크되어 있지 않다면 클릭하여 체크합니다.
                        if not checkbox.is_selected():
                            checkbox.click()  
                        driver.find_element(By.CSS_SELECTOR, "body > div.sc-1jpf51e-1.jljGiJ.popper > div > div.button-container > button.ant-btn.ant-btn-primary > span").click()
                        time.sleep(0.3)


                        # 보고서 생성
                        driver.find_element(By.CSS_SELECTOR, "#generateReport > span").click() 
                        time.sleep(5)

                        if driver.find_element(By.CSS_SELECTOR, "#rc-tabs-0-panel-requestedReport > div > div.react-grid-Container > div > div > div:nth-child(2) > div > div > div:nth-child(2) > div:nth-child(1) > div:nth-child(5) > div > div > span > div").text == "생성 실패":
                            driver.find_element(By.CSS_SELECTOR, "#generateReport > span").click() 
                            time.sleep(5)

                        # 보고서 다운로드
                        element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#rc-tabs-0-panel-requestedReport > div > div.react-grid-Container > div > div > div:nth-child(2) > div > div > div:nth-child(2) > div:nth-child(1) > div:nth-child(6) > div > div > span > div > div:nth-child(2) > button > span"))) 

                        # 다운로드 확인
                        cnt = 1
                        while cnt < 10:
                            current_file_count1 = count_files(download_folder)
                            element.click()

                            time.sleep(3)
                            current_file_count2 = count_files(download_folder)
                            if current_file_count1 != current_file_count2:
                                break

                            cnt += 1

                    # check_download()
                    time.sleep(4)
                    try:
                        WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "body > div.MuiDialog-root.sc-852clq-0.efPzRF > div.MuiDialog-container.MuiDialog-scrollPaper > div > div:nth-child(3) > button"))).click()
                    except: pass


                    ###                                                         쿠팡 C                                                  ###
                    ###                                                         쿠팡 C                                                  ###
                    ###                                                         쿠팡 C                                                  ###

                    driver.get(coupC_url)


                    try:
                        time.sleep(1)
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#search-filter-panel > div:nth-child(1) > span.sc-common-date-group > span:nth-child(1)"))) # 시작 날짜
                        time.sleep(1)
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#search-filter-panel > div:nth-child(1) > span.sc-common-date-group > span:nth-child(1)"))).click() # 시작 날짜
                    except:
                        driver.get(coupC_url)
                        time.sleep(1)
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#search-filter-panel > div:nth-child(1) > span.sc-common-date-group > span:nth-child(1)"))) # 시작 날짜
                        time.sleep(1)
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#search-filter-panel > div:nth-child(1) > span.sc-common-date-group > span:nth-child(1)"))).click() # 시작 날짜

                    input_field = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#dateStart")))
                    time.sleep(0.1)
                    input_field.send_keys(Keys.CONTROL + "a")
                    input_field.send_keys(Keys.BACKSPACE)
                    input_field.send_keys(formatted_date)
                    time.sleep(0.1)

                    click_field = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#search-filter-panel > div:nth-child(1) > span.sc-common-date-group > span:nth-child(2)")))
                    click_field.click() # 끝 날짜

                    input_field = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#dateEnd")))
                    time.sleep(0.1)
                    input_field.send_keys(Keys.CONTROL + "a")
                    input_field.send_keys(Keys.BACKSPACE)
                    input_field.send_keys(formatted_date)
                    
                    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#date_range"))) # 날짜변경확인
                    
                    time.sleep(1)
                    

                    # 로데이터, 쿠팡 다운로드 체크 오류 -> 다운로드 확인 방식 변경 (check_download() -> 수식 풀어서 확인)
                    # 다운로드 확인
                    cnt = 1
                    while cnt < 10:
                        current_file_count1 = count_files(download_folder)
                        time.sleep(3)
                        try:
                            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#download-product-info"))).click() # 다운로드
                            driver.find_element(By.CSS_SELECTOR, "#download-product-info").click()
                        except:
                            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#download-product-info"))).click() # 다운로드
                            driver.find_element(By.CSS_SELECTOR, "#download-product-info").click()
                        current_file_count2 = count_files(download_folder)
                        if current_file_count1 != current_file_count2:
                            break

                        cnt += 1

                    time.sleep(1)
                    # 쿠팡 로그아웃
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#top-header-menu > div.top-header-control.etc-buttons > ul > li.my-user-menu > div'))).click()
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#wing-top-header-user-action-layer > div > div.my-user-menu-bottom > span'))).click()

                    
                # 쿠팡 하엔
                coupC_url = "https://wing.coupang.com/seller/notification/metrics/dashboard"
                coup_report_url = 'https://advertising.coupang.com/marketing-reporting/billboard/reports/pa'

                
                if self.chk_coup_haen.isChecked() == True:
                    
                    coupang_id_haen = self.login_info("COUP_HAEN_ID")
                    coupang_pw_haen = self.login_info("COUP_HAEN_PW")
                    sheet_url_haen_all = 'https://docs.google.com/spreadsheets/d/1V8b3FRe_8witwHXQceekgm-BAvTQLwwkcyuaW-mIi30/edit#gid=1338112098'
                    sheet_name_haenR = '하엔 쿠팡 R'
                    sheet_name_haenC = '하엔C'
                    options = "하엔"

                    coupang(coup_report_url, coupang_id_haen, coupang_pw_haen, coupC_url)
                    coupang_rawdata(sheet_url_haen_all, sheet_name_haenR, options, sheet_name_haenC)

                # 쿠팡 러블로
                if self.chk_coup_lovl.isChecked() == True:

                    coupang_id_lovelo = self.login_info("COUP_LOVE_ID")
                    coupang_pw_lovelo = self.login_info("COUP_LOVE_PW")
                    sheet_url_love_all = 'https://docs.google.com/spreadsheets/d/1NVnVJsCj0Ap_o2xabua9ANUw_1IUREVMJKteY_O1yks/edit#gid=392530415'
                    sheet_name_loveR = '러블로 쿠팡 R'
                    sheet_name_loveC = '러블로C'
                    options = "러브슬라임"

                    coupang(coup_report_url, coupang_id_lovelo, coupang_pw_lovelo, coupC_url)
                    coupang_rawdata(sheet_url_love_all, sheet_name_loveR, options, sheet_name_loveC)

                # 쿠팡 노마셀
                if self.chk_coup_know.isChecked() == True:
                    coupang_id_knowmycell = self.login_info("COUP_KNOW_ID")
                    coupang_pw_knowmycell = self.login_info("COUP_KNOW_PW")
                    sheet_url_know_all = 'https://docs.google.com/spreadsheets/d/12FWmZMuznsxOY_IDbBWeSis-EW1Ds1f9TB6X7K6TFBc/edit#gid=1042061913'
                    sheet_name_knowR = '노마셀 쿠팡 R'
                    sheet_name_knowC = '노마셀C'
                    options = "노마셀"

                    coupang(coup_report_url, coupang_id_knowmycell, coupang_pw_knowmycell, coupC_url)
                    coupang_rawdata(sheet_url_know_all, sheet_name_knowR, options, sheet_name_knowC)


                #########메타로데이터##########
                def meta_rawdata(sheet_url, sheet_name, know_TF):

                    xlsx_file = get_latest_file(download_folder)
                    wb = load_workbook(xlsx_file)
                    ws = wb.active

                    # 서비스 계정 키 파일 경로
                    credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                    # gspread 클라이언트 초기화
                    client = gspread.service_account(filename=credential_file)

                    # Google 시트 열기
                    spreadsheet = client.open_by_url(sheet_url)

                    # 첫 번째 시트 선택
                    sheet = spreadsheet.worksheet(sheet_name)

                    try:

                        if know_TF == 0:
                            data_to_paste = []

                            targe = ws['4']
                            
                            for cell in targe:
                                data_to_paste.append(cell.value)
                            data_to_paste = data_to_paste[2:]
                            print(data_to_paste)

                            today = datetime.date.today().strftime("%Y-%m-%d")
                            column_values = sheet.col_values(1)
                            for idx, cell_value in enumerate(column_values, start=1):  # start=1로 설정하여 행 번호를 1부터 시작
                                if cell_value == today:
                                    print(cell_value)
                                    print(gspread.utils.rowcol_to_a1(idx, 1))
                                    cell_addr = gspread.utils.rowcol_to_a1(idx, 1)
                                    # return f"{gspread.utils.rowcol_to_a1(idx, 1)}"  # 셀 주소 반환
                                
                            (start_row, start_col) = gspread.utils.a1_to_rowcol(cell_addr)

                            # Google 시트에 데이터 쓰기
                            range_to_write = f'B{start_row-1}:P{start_row-1}'
                            sheet.update([data_to_paste], range_to_write)
                        
                        else:
                            last_row = len(sheet.get_all_values())
                            print(last_row)
                            next_row = last_row + 1  # 다음 행 번호

                            if metaDataEmpty:
                                # 데이터 추가
                                range_to_write = f'A{next_row}:Q{next_row}'
                                sheet.update([metaDataEmpty], range_to_write) #한줄

                            else:

                                targe1 = ws['2']
                                targe2 = ws['3']
                                data_to_paste_know1 = []
                                data_to_paste_know2 = []
                                for cell in targe1:
                                    data_to_paste_know1.append(cell.value)
                                for cell in targe2:
                                    data_to_paste_know2.append(cell.value)
                                print(data_to_paste_know1)
                                print(data_to_paste_know2)

                                # 데이터 추가
                                range_to_write = f'A{next_row}:Q{next_row}'
                                sheet.update([data_to_paste_know1], range_to_write) #한줄
                                
                                # K 값 지정
                                K_value = sheet.acell(f'K{next_row}').value
                                K_value_previous = sheet.acell(f'K{next_row-1}').value

                                if sheet.acell(f'A{next_row}').value is not None and K_value is None: #웹사이트 구매 빈칸일 때
                                    sheet.update([[K_value_previous]], f'K{next_row}')

                                range_to_write = f'A{next_row+1}:Q{next_row+1}'
                                sheet.update([data_to_paste_know2], range_to_write) #두줄

                                time.sleep(2)
                                K_value = sheet.acell(f'K{next_row+1}').value
                                if sheet.acell(f'A{next_row+1}').value is not None and K_value is None: #웹사이트 구매 빈칸일 때
                                    sheet.update([[K_value_previous]], f'K{next_row+1}')

                    except:
                        time.sleep(60)

                        if know_TF == 0:
                            data_to_paste = []

                            targe = ws['4']
                            
                            for cell in targe:
                                data_to_paste.append(cell.value)
                            data_to_paste = data_to_paste[2:]
                            print(data_to_paste)

                            today = datetime.date.today().strftime("%Y-%m-%d")
                            column_values = sheet.col_values(1)
                            for idx, cell_value in enumerate(column_values, start=1):  # start=1로 설정하여 행 번호를 1부터 시작
                                if cell_value == today:
                                    print(cell_value)
                                    print(gspread.utils.rowcol_to_a1(idx, 1))
                                    cell_addr = gspread.utils.rowcol_to_a1(idx, 1)
                                    # return f"{gspread.utils.rowcol_to_a1(idx, 1)}"  # 셀 주소 반환
                                
                            (start_row, start_col) = gspread.utils.a1_to_rowcol(cell_addr)

                            # Google 시트에 데이터 쓰기
                            range_to_write = f'B{start_row-1}:P{start_row-1}'
                            sheet.update([data_to_paste], range_to_write)
                        
                        else:
                            last_row = len(sheet.get_all_values())
                            print(last_row)
                            next_row = last_row + 1  # 다음 행 번호

                            if metaDataEmpty:
                                # 데이터 추가
                                range_to_write = f'A{next_row}:Q{next_row}'
                                sheet.update([metaDataEmpty], range_to_write) #한줄

                            else:

                                targe1 = ws['2']
                                targe2 = ws['3']
                                data_to_paste_know1 = []
                                data_to_paste_know2 = []
                                for cell in targe1:
                                    data_to_paste_know1.append(cell.value)
                                for cell in targe2:
                                    data_to_paste_know2.append(cell.value)
                                print(data_to_paste_know1)
                                print(data_to_paste_know2)

                                # 데이터 추가
                                range_to_write = f'A{next_row}:Q{next_row}'
                                sheet.update([data_to_paste_know1], range_to_write) #한줄
                                
                                # K 값 지정
                                K_value = sheet.acell(f'K{next_row}').value
                                K_value_previous = sheet.acell(f'K{next_row-1}').value

                                if sheet.acell(f'A{next_row}').value is not None and K_value is None: #웹사이트 구매 빈칸일 때
                                    sheet.update([[K_value_previous]], f'K{next_row}')

                                range_to_write = f'A{next_row+1}:Q{next_row+1}'
                                sheet.update([data_to_paste_know2], range_to_write) #두줄

                                time.sleep(2)
                                K_value = sheet.acell(f'K{next_row+1}').value
                                if sheet.acell(f'A{next_row+1}').value is not None and K_value is None: #웹사이트 구매 빈칸일 때
                                    sheet.update([[K_value_previous]], f'K{next_row+1}')

                # 메타
                def meta(url_meta, know_TF):
                        
                    driver.get(url_meta)

                    # meta_id = 'healer10@kakao.com'
                    # meta_pw = 'fhdifxmfl1305!!'


### 로그인 정보가 있는 브라우저에서, 로그인 버튼만 찾아서 클릭하는 작업
                    # '비밀번호를' 텍스트를 포함하는 요소 찾기
                    time.sleep(2)

                    try:
                        WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "비밀번호를")]')))
                        pw = driver.find_element(By.XPATH, '//*[contains(text(), "비밀번호를")]')

                        if pw:
                            print("pw 만족")
                            # 이전 형제 요소 찾기
                            parent_element = pw.find_element(By.XPATH, '..')
                            previous_sibling = parent_element.find_element(By.XPATH, 'preceding-sibling::*[1]')
                            print("Previous sibling found:", previous_sibling.text)
                            print
                            previous_sibling.click()
                        
                        else:
                            print("요소를 찾을 수 없습니다.")

                        time.sleep(1)
                        driver.get(url_meta)

                    except:
                        pass
                    #알림 제거
                    try:
                        body = driver.find_element(By.CSS_SELECTOR, 'body')
                        ActionChains(driver).move_to_element(body).click().perform()
                    except: pass
# 6. 메타 캠페인, 일 체크 동작 삭제(CSS선택자 변경 이슈)
                    # if know_TF == 0:  
                    #     try:
                    #         WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#js_1v")))
                    #         checkbox = driver.find_element(By.CSS_SELECTOR, "#js_1v")

                    #         if checkbox.is_selected():
                    #             checkbox.click()  # 체크박스가 체크되어 있다면 클릭하여 체크 해제합니다.#캠페인이름
                    #         time.sleep(0.3)

                    #     except:
                    #         WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#left_rail_nux_target_node > div._5jln > div > div > div.x1ye3gou.xn6708d.xz9dl7a.xjkvuk6 > div:nth-child(1) > div > div > div.x19kh74d > div > div > div > div:nth-child(1)")))
                    #         checkbox = driver.find_element(By.CSS_SELECTOR, "#left_rail_nux_target_node > div._5jln > div > div > div.x1ye3gou.xn6708d.xz9dl7a.xjkvuk6 > div:nth-child(1) > div > div > div.x19kh74d > div > div > div > div:nth-child(1)")
                    #         if know_TF == 0:
                    #             if checkbox.is_selected():
                    #                 checkbox.click()  # 체크박스가 체크되어 있다면 클릭하여 체크 해제합니다.#캠페인이름
                    #             time.sleep(0.3)

                    #     try:
                    #         checkbox = driver.find_element(By.CSS_SELECTOR, "#js_3w")
                    #         if checkbox.is_selected():
                    #             checkbox.click()  # 체크박스가 체크되어 있다면 클릭하여 체크 해제합니다.# 일

                    #     except:
                    #         checkbox = driver.find_element(By.CSS_SELECTOR, "#left_rail_nux_target_node > div._5jln > div > div > div.x1ye3gou.xn6708d.xz9dl7a.xjkvuk6 > div:nth-child(1) > div > div > div.x19kh74d > div > div > div > div:nth-child(13)")
                    #         if checkbox.is_selected():
                    #             checkbox.click()  # 체크박스가 체크되어 있다면 클릭하여 체크 해제합니다.# 일


                    # else:
                        # # 캠페인 이름
                        # WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div/div/div/div/span/div/div[1]/div[1]/div[2]/span/div/div/div/div/div[3]/div[2]/div/div/div/div/div/div[1]/div[5]/div[2]/div[1]/div/div[2]/div[1]/div/div/div[2]/div/div/div/div[1]/div[1]/label/div/div/div[1]/div/div/div[1]/input')))
                        
                        # checkbox = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div/div/div/span/div/div[1]/div[1]/div[2]/span/div/div/div/div/div[3]/div[2]/div/div/div/div/div/div[1]/div[5]/div[2]/div[1]/div/div[2]/div[1]/div/div/div[2]/div/div/div/div[1]/div[1]/label/div/div/div[1]/div/div/div[1]/input')

                        # if not checkbox.is_selected():
                        #     checkbox.click()  
                        # time.sleep(0.3)

                        # # 일
                        # checkbox = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div/div/div/span/div/div[1]/div[1]/div[2]/span/div/div/div/div/div[3]/div[2]/div/div/div/div/div/div[1]/div[5]/div[2]/div[1]/div/div[2]/div[1]/div/div/div[2]/div/div/div/div[13]/div[1]/label/div/div/div[1]/div/div/div[1]/input')
                        # if not checkbox.is_selected():
                        #     checkbox.click()  

                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="PNG_EXPORT"]/div/div[3]/div[1]/div[1]/div/div/div/div[1]/div[2]/div[2]/div/div/div/span')))
                    driver.find_element(By.XPATH, '//*[@id="PNG_EXPORT"]/div/div[3]/div[1]/div[1]/div/div/div/div[1]/div[2]/div[2]/div/div/div/span').click()

                    elements = driver.find_element(By.XPATH, '//*[contains(text(), "어제")]')
                    elements = driver.find_element(By.XPATH, '//*[contains(text(), "어제")]').click()
                    # elements.find_element(By.XPATH, '//*[contains(text(), "2024년 5월")]')


                    try:
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "가장 가까운 기간 보기")]')))
                        global metaDataEmpty
                        metaDataEmpty = [str(today_yday), '-', 0, 0, 0, 0, 0, 0, 0, 0, '-', 0, 0, 0, 0, str(today_yday), str(today_yday)]

                    except:
                        time.sleep(0.5)
                        driver.find_element(By.CSS_SELECTOR, "#PNG_EXPORT > div > div:nth-child(1) > div > div._4bl7 > div > div.x3nfvp2.x193iq5w.xxymvpz").click() #새로고침
                        time.sleep(1.5)
                        element = driver.find_element(By.CSS_SELECTOR, "#export_button > div > div > span > div > div.xeuugli.x2lwn1j.x6s0dn4.x78zum5.x1q0g3np.x1iyjqo2.xozqiw3.x19lwn94.x1hc1fzr.x13dflua.x6o7n8i.xxziih7.x12w9bfk.xl56j7k.xh8yej3 > div > div")
                        driver.find_element(By.CSS_SELECTOR, "#export_button > div > div > span > div > div.xeuugli.x2lwn1j.x6s0dn4.x78zum5.x1q0g3np.x1iyjqo2.xozqiw3.x19lwn94.x1hc1fzr.x13dflua.x6o7n8i.xxziih7.x12w9bfk.xl56j7k.xh8yej3 > div > div").click() #내보내기
                        time.sleep(1.5)
                        ActionChains(driver).move_to_element_with_offset(element,-579,497).click().perform() #다운로드

                    
                        while True:
                            check_download()
                            if check == 0:
                                    driver.find_element(By.CSS_SELECTOR, "#export_button > div > div > span > div > div.xeuugli.x2lwn1j.x6s0dn4.x78zum5.x1q0g3np.x1iyjqo2.xozqiw3.x19lwn94.x1hc1fzr.x13dflua.x6o7n8i.xxziih7.x12w9bfk.xl56j7k.xh8yej3 > div > div").click() #내보내기
                                    time.sleep(1)
                                    ActionChains(driver).move_to_element_with_offset(element,-579,497).click().perform() #다운로드
                            else: break

                        
                    time.sleep(3)
                
                #메타 하엔
                if self.chk_meta_haen.isChecked() == True:
                    url_meta_haen = 'https://adsmanager.facebook.com/adsmanager/reporting/view?act=774078054299392&business_id=341660836507461&selected_report_id=120202962853720679' #하엔
                    sheet_url_haen_all = 'https://docs.google.com/spreadsheets/d/1V8b3FRe_8witwHXQceekgm-BAvTQLwwkcyuaW-mIi30/edit#gid=168246212'
                    sheet_name_haen = '하엔 페이스북 R'
                    know_TF = 0

                    meta(url_meta_haen, know_TF)
                    meta_rawdata(sheet_url_haen_all, sheet_name_haen, know_TF)

                #메타 러블로
                if self.chk_meta_lovl.isChecked() == True:
                    url_meta_lovelo = 'https://adsmanager.facebook.com/adsmanager/reporting/view?act=1913234209031352&business_id=267018165996779&selected_report_id=120200964749160675' #러블로
                    sheet_url_love_all = 'https://docs.google.com/spreadsheets/d/1NVnVJsCj0Ap_o2xabua9ANUw_1IUREVMJKteY_O1yks/edit#gid=1607702031'
                    sheet_name_love = '러블로 페이스북 R'
                    know_TF = 0

                    meta(url_meta_lovelo, know_TF)
                    meta_rawdata(sheet_url_love_all, sheet_name_love, know_TF)

                #메타 노마셀
                if self.chk_meta_know.isChecked() == True:
                    url_meta_knowmycell = 'https://adsmanager.facebook.com/adsmanager/reporting/view?act=238068255778220&business_id=635001998695042&selected_report_id=120200841324100083' #노마셀
                    sheet_url_know_all = 'https://docs.google.com/spreadsheets/d/12FWmZMuznsxOY_IDbBWeSis-EW1Ds1f9TB6X7K6TFBc/edit#gid=137297262'
                    sheet_name_know = '노마셀 페이스북 R'
                    know_TF = 1

                    meta(url_meta_knowmycell, know_TF)
                    meta_rawdata(sheet_url_know_all, sheet_name_know, know_TF)


                #########구글로데이터##########
                def google_rawdata(sheet_url, sheet_name, brand):
                    
                    csv_file = get_latest_file(download_folder)

                    # 서비스 계정 키 파일 경로
                    credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                    # gspread 클라이언트 초기화
                    client = gspread.service_account(filename=credential_file)

                    # Google 시트 열기
                    spreadsheet = client.open_by_url(sheet_url)

                    # 첫 번째 시트 선택
                    sheet = spreadsheet.worksheet(sheet_name)
        
                    selected_rows = []

                    with open(csv_file, newline='', encoding='utf-16') as csvfile:
                        reader = csv.reader(csvfile)
                        for i, row in enumerate(reader):
                            if 3 <= i <= 10:  # 범위 내 행 적용

                                # 데이터를 올바르게 파싱하기 위해 먼저 전체 문자열을 하나로 합친다
                                full_data = "".join(row)

                                # 탭(\t)으로 데이터를 분리한다
                                parsed_data = full_data.split('\t')

                                cleaned_data = [item.replace('"', '') for item in parsed_data]
                                selected_rows.append(cleaned_data)
                            print(selected_rows)

                    for items in selected_rows:
                        new_selected_rows = []

                        result = []
                        for item in items:
                            if isinstance(item, str) and '%' in item:
                                result.append(float(item.strip('%')) / 100)
                            elif isinstance(item, str) and ',' in item:
                                result.append(int(item.replace(',', '')))
                            elif isinstance(item, str) and item.replace('.', '', 1).isdigit() and item.count('.') == 1:
                                result.append(float(item))
                            elif isinstance(item, str) and item.isdigit():
                                result.append(int(item))
                            else:
                                result.append(item)
                        new_selected_rows.append(result)
                        

                        last_row = len(sheet.col_values(1))
                        print(last_row)
                        # remove_set = {''}
                        # last_row_removed = len([i for i in last_row if i not in remove_set])

                        # a_column = [row[0] for row in last_row]
                        next_row = int(last_row) + 1  # 다음 행 번호
                        print(next_row)
                        if brand == "하엔":
                            print("하엔 입력 시작")
                            print(new_selected_rows)

                            range_to_write = f'A{next_row}:M{next_row}'

                            sheet.update(new_selected_rows, range_to_write)

                            # 셀 포맷 설정
                            sheet.format(f"H{next_row}", {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}})
                            sheet.format(f'I{next_row}:L{next_row}', {"numberFormat": {"type":'NUMBER'}})

                        if brand == "노마셀":
                            print("노마셀 입력 시작")
                            print(new_selected_rows)

                            range_to_write = f'A{next_row}:K{next_row}'

                            sheet.update(new_selected_rows, range_to_write)

                            # 셀 포맷 설정
                            sheet.format(f"G{next_row}", {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}})
                            sheet.format(f'H{next_row}:J{next_row}', {"numberFormat": {"type":'NUMBER'}})

                            

                # 구글애즈
                def google(url_google):
                    driver.get(url_google)
                    calOpen = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CLASS_NAME, 'button-text')))

                    try:
                        time.sleep(1)
                        # 달력 열기
                        calOpen.click()

                    except:
                        driver.find_element(By.XPATH, "//*[contains(text(), '보기 필터 지우기')]").click()
                        time.sleep(1)
                        # 달력 열기
                        calOpen.click()

                        
                    # 어제 선택
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CLASS_NAME, 'visible-month')))
                    time.sleep(1.5)
                    lists = driver.find_element(By.CLASS_NAME, "preset-container").find_elements(By.CLASS_NAME, "item")

                    for item in lists:
                        print(item.text)
                        if item.text == "어제":
                            item.click()
                            break

                    time.sleep(0.5)
                    # 11. 구글 다운로드 클릭, CSS선택자 방식에서 XPATH - contains, text() 방식으로 변경(CSS선택자 매번바뀜)
                    schedule = driver.find_element(By.XPATH, "//*[contains(text(), '일정')]")
                    #다운
                    # 부모의 부모의 부모의 부모의 부모의 이전 요소 찾기 및 클릭
                    previous_sibling = schedule.find_element(By.XPATH, "ancestor::*[4]/preceding-sibling::*[1]")
                    previous_sibling.click()
                    
                    
                    # Excel .csv 선택
                    lists = driver.find_element(By.CLASS_NAME, "download-dropdown").find_elements(By.CLASS_NAME, "item")

                    for item in lists:
                        print(item.text)
                        if item.text == "Excel .csv":
                            item.click()
                            break
## 구글 다운로드 실패 시 재시도 적용
                    try:
                        check_download()
                    except:
                        schedule = driver.find_element(By.XPATH, "//*[contains(text(), '일정')]")
                        #다운
                        # 부모의 부모의 부모의 부모의 부모의 이전 요소 찾기 및 클릭
                        previous_sibling = schedule.find_element(By.XPATH, "ancestor::*[4]/preceding-sibling::*[1]")
                        previous_sibling.click()
                        
                        
                        # Excel .csv 선택
                        lists = driver.find_element(By.CLASS_NAME, "download-dropdown").find_elements(By.CLASS_NAME, "item")

                        for item in lists:
                            print(item.text)
                            if item.text == "Excel .csv":
                                item.click()
                                break
                        check_download()


                    time.sleep(2)

                # 구글 하엔
                if self.chk_goog_haen.isChecked() == True:
                    url_ads_haen = 'https://ads.google.com/aw/reporteditor/view?ocid=1181720304&workspaceId=0&reportId=927965366&euid=1114690018&__u=8943315282&uscid=1181720304&__c=5821258096&authuser=0'
                    sheet_url_goog = "https://docs.google.com/spreadsheets/d/1V8b3FRe_8witwHXQceekgm-BAvTQLwwkcyuaW-mIi30/edit#gid=1966867512"
                    sheet_name_goog = "하엔 구글 R"
                    brand = "하엔"
                    google(url_ads_haen)
                    google_rawdata(sheet_url_goog, sheet_name_goog, brand)


                if self.chk_goog_know.isChecked() == True:
                    url_ads_know = 'https://ads.google.com/aw/reporteditor/view?ocid=1379143590&workspaceId=-1615213561&reportId=928192574&euid=1114690018&__u=8943315282&uscid=1379143590&__c=4267857910&authuser=0'
                    sheet_url_goog = "https://docs.google.com/spreadsheets/d/12FWmZMuznsxOY_IDbBWeSis-EW1Ds1f9TB6X7K6TFBc/edit#gid=1001228164"
                    sheet_name_goog = "노마셀 구글 R"
                    brand = "노마셀"
                    google(url_ads_know)
                    google_rawdata(sheet_url_goog, sheet_name_goog, brand)



                def ssDown(brand):

                    # 날짜 구하기
                    today = date.today()

                    today_date = today.strftime("%d")
                    today_Ym = today.strftime("%Y. %m.")

                    number = target_days_input
                    dayx = datetime.timedelta(days=number)
                    dayy = datetime.timedelta(days=1)
                    day1 = datetime.timedelta(days=1)

                    today_yday = today-day1
                    startday = today-dayx
                    endday = today-dayy
                    tday_Ym = startday.strftime("%Y. %m.")
                    tday_d = startday.strftime("%d")

                    # EdgeOptions 객체 생성
                    edge_options = webdriver.EdgeOptions()
                    edge_options.use_chromium = True
                    edge_options.add_argument("disable-gpu")
                    edge_options.add_argument("no-sandbox")


                    # 사용자의 프로필 경로 설정
                    profile_path = 'C:\\Users\\A\\AppData\\Local\\Microsoft\\Edge\\User Data1'
                    edge_options.add_argument(f"user-data-dir={profile_path}")
                    edge_options.add_argument("--profile-directory=Default")


                    # Edge 드라이버 서비스 시작
                    edge_service = Service(executable_path='C:\\path\\to\\edgedriver_win64\\msedgedriver.exe')
                    edge_driver = webdriver.Edge(service=edge_service, options=edge_options)
                    # 스스 엣지드라이버 꺼짐 오류 -> 엣지드라이버 업데이트


                    edge_driver.get("https://bizadvisor.naver.com/shopping/product")
                    # 로그인
                    
                    WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#wrap > div > div > div.login_box > ul > li:nth-child(1) > a"))).click()
                    try:
                        WebDriverWait(edge_driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#root > div > div.Layout_wrap__3uDBh > div > div > div.Login_simple_box__2bfAS > button"))).click()

                    except:
                        driver.find_element(By.CSS_SELECTOR, '[class^="Login_btn_more"]').click()

                        current_window_handle = driver.current_window_handle

                        new_window_handle = None
                        while not new_window_handle:
                            for handle in driver.window_handles:
                                if handle != current_window_handle:
                                    new_window_handle = handle
                                    break

                        #팝업으로 제어 변경
                        driver.switch_to.window(driver.window_handles[1])


                        WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#log\.login")))
                        
                        txtInput = driver.find_element(By.CSS_SELECTOR, "#id")
                        txtInput.send_keys("wntlsqhr")
                        time.sleep(0.1)
                        txtInput = driver.find_element(By.CSS_SELECTOR, "#pw")
                        txtInput.send_keys("dnflskfk00@")
                        time.sleep(0.1)
                        driver.find_element(By.CSS_SELECTOR, "#log\.login")

                        #원래 페이지로 제어 변경
                        driver.switch_to.window(driver.window_handles[0])
            





                    # 상품별 이동
                    WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#include_nav > div > div > div:nth-child(1) > ul > li:nth-child(4) > a"))).click()
                    WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#include_nav > div > div > div:nth-child(1) > ul > li.on > div > ul > li:nth-child(1) > a"))).click()

                    brandtext = edge_driver.find_element(By.CSS_SELECTOR, "#include_header > div > div.header_tit > div > div:nth-child(2) > div > div > div > span:nth-child(1)").text[:3]


                    # 브랜드 변경
                    if not brandtext == brand:
                        edge_driver.find_element(By.CSS_SELECTOR, "#include_header > div > div.header_tit > div > div:nth-child(2) > div > div > div > div > a > div > span").click()

                    while startday != today:

                        # 날짜 클릭(달력오픈)
                        WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.date_select > a.btn.select_data'))).click()


                        # 날짜 변수 지정
                        tday_Ym = startday.strftime("%Y. %m.")
                        tday_d = str(int(startday.strftime("%d")))
                        trick = (startday-day1).strftime("%Y. %m.")
                        print(startday)


                        # %Y. %m 표시(웹상)
                        DPmonthStart = edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_area > div.pick_calendar_layout > div.DayPicker.DayPicker > div > div.DayPicker-Month.rdp-caption_start")
                        DPmonthBtw = edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_area > div.pick_calendar_layout > div.DayPicker.DayPicker > div > div.DayPicker-Month.rdp-caption_between")


                        # 첫번째 단락 년,월 대조
                        if tday_Ym == DPmonthStart.text[:9]:

                                days = edge_driver.find_elements(By.XPATH, f"//*[@id='wrap']/div[1]/section/div/div[2]/div[1]/div/ul/li[2]/div/div/div[2]/div/div[1]/div[2]/div[2]/div/div[1]/table/tbody//td[not(contains(@class, 'DayPicker-Day--outside'))]")
                                for day in days:
                                    if day.text == tday_d:
                                        print("target: ", day.text)
                                        day.click()
                                        time.sleep(0.1)
                                        day.click()
                                        break

                                # 적용
                                edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_info_area > div.btn_area > a:nth-child(1)").click()
                        
                        
                        # 두번째 단락 년,월 대조
                        elif tday_Ym == DPmonthBtw.text[:9]:
                        
                                days = edge_driver.find_elements(By.XPATH, f"//*[@id='wrap']/div[1]/section/div/div[2]/div[1]/div/ul/li[2]/div/div/div[2]/div/div[1]/div[2]/div[2]/div/div[2]/table/tbody//td[not(contains(@class, 'DayPicker-Day--outside'))]")
                                for day in days:
                                    if day.text == tday_d:
                                        print("target: ", day.text)
                                        day.click()
                                        time.sleep(0.1)
                                        day.click()
                                        break
                                    
                                # 적용
                                edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_info_area > div.btn_area > a:nth-child(1)").click()


                        # 다운로드 확인
                        cnt = 1
                        while cnt < 10:
                            current_file_count1 = count_files(download_folder)
                            # 다운로드 버튼
                            WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(1) > span > a'))).click()
                            time.sleep(3)
                            current_file_count2 = count_files(download_folder)
                            if current_file_count1 != current_file_count2:
                                break

                            cnt += 1

                        startday += datetime.timedelta(days=1)

                def ssWrite(brandSheet, sheet_url):
                    defaultData = ["화장품/미용", "바디케어", "입욕제", "-", "러블로 러브슬라임 슬라임탕 젤 입욕제 젤탕", "9019908272",	"일반배송",	"0", "0", "0", "0.00%"]

                    # 날짜 구하기
                    today = date.today()

                    today_date = today.strftime("%d")
                    today_Ym = today.strftime("%Y. %m.")

                    number = target_days_input
                    dayx = datetime.timedelta(days=number)
                    dayy = datetime.timedelta(days=1)
                    day1 = datetime.timedelta(days=1)

                    today_yday = today-day1
                    startday = today-dayx
                    endday = today-dayy
                    tday_Ym = startday.strftime("%Y. %m.")
                    tday_d = startday.strftime("%d")

                    while number > 0:

                        # 서비스 계정 키 파일 경로
                        credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                        # gspread 클라이언트 초기화
                        client = gspread.service_account(filename=credential_file)

                        # Google 시트 열기
                        spreadsheet = client.open_by_url(sheet_url)

                        # 첫 번째 시트 선택
                        sheet = spreadsheet.worksheet(brandSheet)

                        last_row = len(sheet.get_all_values())
                        print(last_row)
                        next_row = last_row + 1  # 다음 행 번호

                        i = get_nth_latest_file(download_folder, number)

                        wb = openpyxl.load_workbook(i)
                        sheet = wb.active  # 활성 시트 선택

                        if check_data_in_second_row(i):
                            pass

                        else:
                            sheet = spreadsheet.worksheet(brandSheet)

                            # 날짜 입력
                            sheet.update([[str(startday)]], f"A{next_row}")
                            range_to_write = f'B{next_row}:L{next_row}'
                            sheet.update([defaultData], range_to_write)

                            number -= 1
                            startday += timedelta(days=1)  # 날짜 하루 증가
                            continue


                        # 원본 시트의 행을 반복하며 첫 번째 행을 제외하고 데이터가 있는 행만 복사
                        for row in sheet.iter_rows(min_row=2):  # 첫 번째 행을 제외하고 시작
                            # 각 셀에 데이터가 있는지 확인
                            data_exists = any(cell.value not in (None, '', ' ') for cell in row)  # 빈 문자열과 공백도 무시

                            # 서비스 계정 키 파일 경로
                            credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                            # gspread 클라이언트 초기화
                            client = gspread.service_account(filename=credential_file)

                            # Google 시트 열기
                            spreadsheet = client.open_by_url(sheet_url)

                            # 첫 번째 시트 선택
                            sheet = spreadsheet.worksheet(brandSheet)

                            # 날짜 입력
                            sheet.update([[str(startday)]], f"A{next_row}")
                            
                            values = []
                            for col_index, cell in enumerate(row, start=2):
                                values.append(cell.value)

                            range_to_write = f'B{next_row}:L{next_row}'
                            sheet.update([values], range_to_write)

                            
                            next_row += 1

                        startday += timedelta(days=1)  # 날짜 하루 증가
                        number -= 1

                # if self.chk_ss_haen.isChecked() == True:

                #     label = self.chk_cafe_haen
                #     url_cafe24 = "https://eclogin.cafe24.com/Shop/"
                #     url_cafe24_req_haen = "https://woo8425.cafe24.com/disp/admin/shop1/report/DailyList"
                    
                #     cafe24_id_haen = self.login_info("CAFE_HAEN_ID")
                #     cafe24_pw_haen = self.login_info("CAFE_HAEN_PW")

                #     sheet_haenR_url = 'https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=1894651086'
                #     sheet_haenR = '하엔R'
                #     sheet_haenD = "하엔D"

                sheet_url = "https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=374561563"
                if self.chk_ss_lovl.isChecked() == True:

                    brand = "러브슬"
                    brandSheet = "러블로N"

                    ssDown(brand)
                    ssWrite(brandSheet, sheet_url)


                    label = self.chk_cafe_haen
                    url_cafe24 = "https://eclogin.cafe24.com/Shop/"
                    url_cafe24_req_haen = "https://woo8425.cafe24.com/disp/admin/shop1/report/DailyList"
                    
                    cafe24_id_haen = self.login_info("CAFE_HAEN_ID")
                    cafe24_pw_haen = self.login_info("CAFE_HAEN_PW")

                    sheet_haenR_url = 'https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=1894651086'
                    sheet_haenR = '하엔R'
                    sheet_haenD = "하엔D"

                if self.chk_ss_know.isChecked() == True:

                    brand = "노마셀"
                    brandSheet = "노마셀N"

                    ssDown(brand)
                    ssWrite(brandSheet, sheet_url)


                    label = self.chk_cafe_haen
                    url_cafe24 = "https://eclogin.cafe24.com/Shop/"
                    url_cafe24_req_haen = "https://woo8425.cafe24.com/disp/admin/shop1/report/DailyList"
                    
                    cafe24_id_haen = self.login_info("CAFE_HAEN_ID")
                    cafe24_pw_haen = self.login_info("CAFE_HAEN_PW")

                    sheet_haenR_url = 'https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=1894651086'
                    sheet_haenR = '하엔R'
                    sheet_haenD = "하엔D"


                def naverad(url):

                    # EdgeOptions 객체 생성
                    edge_options = webdriver.EdgeOptions()
                    edge_options.use_chromium = True
                    edge_options.add_argument("disable-gpu")
                    edge_options.add_argument("no-sandbox")


                    # 사용자의 프로필 경로 설정
                    profile_path = 'C:\\Users\\A\\AppData\\Local\\Microsoft\\Edge\\User Data1'
                    edge_options.add_argument(f"user-data-dir={profile_path}")
                    edge_options.add_argument("--profile-directory=Default")


                    # Edge 드라이버 서비스 시작
                    edge_service = Service(executable_path='C:\\path\\to\\edgedriver_win64\\msedgedriver.exe')
                    edge_driver = webdriver.Edge(service=edge_service, options=edge_options)

                    edge_driver.get(url)

                    # 네이버검색광고 로그인 확인 창 제거 로직 수정(visibility_of_all_elements_located -> element_to_be_clickable)
                    try:
                        WebDriverWait(edge_driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "환영합니다")]')))
                        closeButton = edge_driver.find_element(By.CLASS_NAME, "btn_name")
                        closeButton.click()
                        print("로그인확인 창 제거")
                    except: 
                        print("로그인확인 창 없음")
                        pass

                    # 네이버검색광고 로그인 확인 창 제거 로직 수정
                    try:
                        WebDriverWait(edge_driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "환영합니다")]')))
                        checkbox = edge_driver.find_element(By.ID, "chk_cls")
                        checkbox.click()

                    except: pass

                    # 캘린더 열기
                    WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-nclick="datePicker"]'))).click()

                    # 날짜 선택
                    WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "지난 7일")]'))).click()

                    time.sleep(0.5)
                    element = edge_driver.find_element(By.XPATH, "//*[contains(text(), '다운로드') and not(contains(text(), '대용량 다운로드 보고서'))]")
                    time.sleep(0.5)

                    # 다운로드 확인
                    cnt = 1
                    while cnt < 10:
                        current_file_count1 = count_files(download_folder)
                        element.click()
                        time.sleep(3)
                        current_file_count2 = count_files(download_folder)
                        if current_file_count1 != current_file_count2:
                            break

                        cnt += 1
                        
                    
                    time.sleep(1)
                    edge_driver.close()

                def naveradInput(url, name):

                    target_days = 1
                    dayx = datetime.timedelta(days=target_days)
                    day1 = datetime.timedelta(days=1)

                    # 오늘 날짜 구하기
                    today_yday = today-day1
                    today_tday = today-dayx
                    # CSV 파일 읽기 (첫 번째 행은 건너뛰고 두 번째 행을 열 이름으로 사용)
                    df = pd.read_csv(get_latest_file(download_folder), skiprows=1)

                    # 날짜 열 이름 추출 (A열, 즉 첫 번째 열)
                    date_column = df.columns[0]

                    # 'date' 열을 datetime 형식으로 변환
                    df[date_column] = pd.to_datetime(df[date_column], format='%Y.%m.%d.')

                    df[date_column] = df[date_column].dt.strftime('%Y-%m-%d')
                    # 필터링된 데이터프레임 출력
                    dataList = df.values.tolist()
                    print(df.values.tolist())

                    csv_file = get_latest_file(download_folder)

                    # 서비스 계정 키 파일 경로
                    credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                    # gspread 클라이언트 초기화
                    client = gspread.service_account(filename=credential_file)

                    # Google 시트 열기
                    spreadsheet = client.open_by_url(url)

                    # 첫 번째 시트 선택
                    sheet = spreadsheet.worksheet(name)

                    while today_tday != today:
                        print(today_tday, "검색 시작")
                        for i in dataList:
                            if str(today_tday) in i:
                                result = []
                                print(today_tday, "찾음!")

                                for item in i:
                                    if isinstance(item, str) and '%' in item:
                                        result.append(float(item.strip('%')) / 100)
                                    elif isinstance(item, str) and ',' in item and '.' in item.replace(',', ''):
                                        result.append(float(item.replace(',', '')))
                                    elif isinstance(item, str) and ',' in item:
                                        result.append(int(item.replace(',', '')))
                                    elif isinstance(item, str) and item.replace('.', '', 1).isdigit() and item.count('.') == 1:
                                        result.append(float(item))
                                    elif isinstance(item, str) and item.isdigit():
                                        result.append(int(item))
                                    else:
                                        result.append(item)


                                last_row = len(sheet.col_values(1))
                                next_row = int(last_row) + 1
                                range_to_write = f'A{next_row}:S{next_row}'
                                sheet.update([result], range_to_write)
                        today_tday += timedelta(days=1)

                if self.chk_nad_haen.isChecked() == True:

                        sheet_url = 'https://docs.google.com/spreadsheets/d/1V8b3FRe_8witwHXQceekgm-BAvTQLwwkcyuaW-mIi30/edit?gid=2136174248#gid=2136174248'
                        sheet_name = '하엔 네이버 R'
                        target_url = "https://manage.searchad.naver.com/customers/2621471/reports/rtt-a001-000000000650376"
                        
                        naverad(target_url)
                        naveradInput(sheet_url, sheet_name)

                if self.chk_nad_lovl.isChecked() == True:

                    sheet_url = 'https://docs.google.com/spreadsheets/d/1NVnVJsCj0Ap_o2xabua9ANUw_1IUREVMJKteY_O1yks/edit?gid=910059812#gid=910059812'
                    sheet_name = '러블로 네이버 R'
                    target_url = "https://manage.searchad.naver.com/customers/2914810/reports/rtt-a001-000000000651901"
                    
                    naverad(target_url)
                    naveradInput(sheet_url, sheet_name)

                if self.chk_nad_know.isChecked() == True:

                    sheet_url = 'https://docs.google.com/spreadsheets/d/12FWmZMuznsxOY_IDbBWeSis-EW1Ds1f9TB6X7K6TFBc/edit?gid=1997928779#gid=1997928779'
                    sheet_name = '노마셀 네이버 R'
                    target_url = "https://manage.searchad.naver.com/customers/2957190/reports/rtt-a001-000000000651985"
                    
                    naverad(target_url)
                    naveradInput(sheet_url, sheet_name)


            elif target_days_input >= 2:
                
                # 날짜 구하기
                today = date.today()
                # 하루를 나타내는 timedelta 객체 생성
                # 어제 날짜를 구함

                today_date = today.strftime("%d")
                today_month = str(int(today.strftime("%m")))

                weekday_korean = {
                    0: '월',
                    1: '화',
                    2: '수',
                    3: '목',
                    4: '금',
                    5: '토',
                    6: '일'
                }
                
                dayx = datetime.timedelta(days=target_days_input)
                day1 = datetime.timedelta(days=1)

                # 오늘 날짜 구하기
                today_yday = today-day1
                today_tday = today-dayx
                today_Tday년월 = (today-dayx).strftime("%Y년 %m월")
                today_Yday년월 = (today-day1).strftime("%Y년 %m월")
                Tday_month월 = str(int(today_tday.strftime("%m"))) + "월"
                Yday_month월 = str(int(today_yday.strftime("%m"))) + "월"
                today_Tday일 = str(int((today-dayx).strftime("%d")))
                today_Yday일 = str(int((today-day1).strftime("%d")))


                weekday_num = today.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)
                weekday_numy = today_yday.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)
                weekday_numt = today_tday.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)
                # 요일을 한국어로 변환
                weekday_kr = weekday_korean[weekday_num]
                weekday_kry = weekday_korean[weekday_numy]
                weekday_krt = weekday_korean[weekday_numt]

                weekday = f"{today}({weekday_kr})"
                weekday_y = f"{today_yday}({weekday_kry})"
                weekday_t = f"{today_tday}({weekday_krt})"

                #카페24
                def cafe24(label, url_cafe24, url_cafe24_req, cafe24_id, cafe24_pw, sheet_urlR, sheet_nameR, sheet_nameD):
                    
                    driver.get(url_cafe24)

                    ##################################### 로그인
                    ##################################### 로그인
                    ##################################### 로그인
                    ##################################### 로그인

                    # ID
                    input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#mall_id")))
                    input_field.click()
                    time.sleep(1)
                    input_field.send_keys(Keys.CONTROL + "a")
                    input_field.send_keys(Keys.BACKSPACE)
                    driver.find_element(By.CSS_SELECTOR, "#mall_id").send_keys(cafe24_id)

                    # PW
                    input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#userpasswd")))
                    input_field.click()
                    input_field.send_keys(Keys.CONTROL + "a")
                    input_field.send_keys(Keys.BACKSPACE)
                    driver.find_element(By.CSS_SELECTOR, "#userpasswd").send_keys(cafe24_pw)

                    # 로그인클릭
                    driver.find_element(By.CSS_SELECTOR,'#frm_user > div > div.mButton > button').click()

                    #비밀번호변경안내
                    try: WebDriverWait(driver, 5).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#iptBtnEm")))).click() 
                    except: pass

                    try:
                        time.sleep(3)
                        popup = driver.find_element(By.XPATH, '//*[contains(text(), "오늘 하루 보지 않기")]')
                        popup.click()

                    except: pass

                    #화면로딩대기
                    WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, '//*[contains(text(), "오늘의 할 일")]')))

                    # 데이터 접근
                    driver.get(url_cafe24_req)

                    # 자세히보기클릭
                    element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#QA_day3 > div.mBoard.gScroll > table")))
                    driver.execute_script("arguments[0].scrollIntoView(true);", element) # 스크롤다운
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#sReportGabView"))).click() 
                    
                    element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#QA_day3 > div.mBoard.gScroll > table")))
                    driver.execute_script("arguments[0].scrollIntoView(true);", element) # 스크롤다운
                    rows = driver.find_elements(By.CSS_SELECTOR, 'tbody.right tr')

                    cover = []
                    cover0 = []
                    for element in rows:
                        new_data_list = []
                        rawdata = element.text
                        # 문자열을 공백을 기준으로 분리하여 리스트로 변환
                        data_list = rawdata.split()
                        data_list = [x.replace(',', '') for x in data_list]

                        for items in data_list:

                        # 숫자인 경우 숫자로 변환
                            try:
                                numeric_value = int(items)
                                new_data_list.append(numeric_value)
                            except:
                                # 숫자가 아닌 경우 원래 값 유지
                                new_data_list.append(items)
                        cover.append(new_data_list[1:])
                        cover0.append(new_data_list[0])


                    today_tdayTemp = today_tday
                    today_tdayTempDay = today_tday.strftime(f"%Y-%m-%d({weekday_krt})")
                    cover.reverse()
                    cover0.reverse()
                    print(cover)

                    print(cover0)
                    print(today_tdayTemp)
                    for i in range(target_days_input):
                        print(today_tdayTemp)

                         # 서비스 계정 키 파일 경로
                        credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                        # gspread 클라이언트 초기화
                        client = gspread.service_account(filename=credential_file)

                        # Google 시트 열기
                        spreadsheet = client.open_by_url(sheet_urlR)

                        # 첫 번째 시트 선택
                        sheet = spreadsheet.worksheet(sheet_nameR)

                        column_values = sheet.col_values(1)
                        for idx, cell_value in enumerate(column_values, start=1):  # start=1로 설정하여 행 번호를 1부터 시작
                            if cell_value == str(today_tdayTemp):
                                print(cell_value)
                                print(gspread.utils.rowcol_to_a1(idx, 1))
                                cell_addr = gspread.utils.rowcol_to_a1(idx, 1)
                                # return f"{gspread.utils.rowcol_to_a1(idx, 1)}"  # 셀 주소 반환
                            
                        (start_row, start_col) = gspread.utils.a1_to_rowcol(cell_addr)
                        print(start_row)
                        print(today_tdayTempDay)

                        # last_row = len(sheet.col_values(3))
                        # next_row = last_row + 1
                        # print(last_row)
                        # print(next_row)

                        if today_tdayTempDay in cover0:
                            print("성립")
                            keynum = cover0.index(today_tdayTempDay)
                            data_to_paste = cover[keynum]    

                            data1 = data_to_paste[:9]
                            data2 = data_to_paste[9]
                            data3 = data_to_paste[10:]
                        # 카페24 R, 데이터 없으면 0 입력 되도록 코드 수정 -2
                        else:
                            data1 = [0, 0, 0, 0, 0, 0, 0, 0, 0]
                            data2 = 0
                            data3 = [0, 0, 0]


                        print(data1)
                        print(data2)
                        print(data3)

                        range1 = f'C{start_row}:K{start_row}'
                        range2 = f'M{start_row}'
                        range3 = f'O{start_row}:Q{start_row}'
                        
                        sheet.update([data1], range1)
                        sheet.update([[data2]], range2)
                        sheet.update([data3], range3)


                        weekday_korean = {
                            0: '월',
                            1: '화',
                            2: '수',
                            3: '목',
                            4: '금',
                            5: '토',
                            6: '일'
                        }
                        
                        dayx = datetime.timedelta(days=target_days_input)
                        day1 = datetime.timedelta(days=1)

                        # 오늘 날짜 구하기
                        today_yday = today-day1
                        today_tdayTemp = today_tdayTemp + timedelta(days=1)
                        today_Tday년월 = (today-dayx).strftime("%Y년 %m월")
                        today_Yday년월 = (today-day1).strftime("%Y년 %m월")
                        today_Tday일 = str(int((today-dayx).strftime("%d")))
                        today_Yday일 = str(int((today-day1).strftime("%d")))


                        weekday_num = today.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)
                        weekday_numy = today_yday.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)
                        weekday_numtTemp = today_tdayTemp.weekday()  # 요일 번호 (월요일이 0, 일요일이 6)
                        # 요일을 한국어로 변환
                        weekday_kr = weekday_korean[weekday_num]
                        weekday_kry = weekday_korean[weekday_numy]
                        weekday_krtTemp = weekday_korean[weekday_numt]

                        weekday = f"{today}({weekday_kr})"
                        weekday_y = f"{today_yday}({weekday_kry})"
                        weekday_t = f"{today_tday}({weekday_krt})"

                        print(today_tdayTemp)
                        weekday_krtTemp = weekday_korean[weekday_numtTemp]
                        print(weekday_krtTemp)
                        today_tdayTempDay = f"{today_tdayTemp}({weekday_krtTemp})"
                        print(today_tdayTempDay)

                    ###################################### 조회수
                    ###################################### 조회수
                    ###################################### 조회수
                    ###################################### 조회수

                    driver.find_element(By.CSS_SELECTOR, "#mCSB_2_container > ul:nth-child(1) > li:nth-child(9)").click() #통계 클릭
                    WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#mCSB_3_container > div.depthList > ul > li:nth-child(8)"))).click() #접속통계클릭

                    #새 창 대기
                    current_window_handle = driver.current_window_handle

                    new_window_handle = None
                    while not new_window_handle:
                        for handle in driver.window_handles:
                            if handle != current_window_handle:
                                new_window_handle = handle
                                break


                    #팝업으로 제어 변경
                    driver.switch_to.window(driver.window_handles[1]) 

                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#rpt_pth")))).click() #방문경로분석
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#rpt_pth_dod")))).click() #방문도메인(상세)
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#body_center > table:nth-child(5) > tbody > tr:nth-child(2) > td:nth-child(3) > a:nth-child(2) > img")))).click()

                    for ii in range(1, target_days_input+1):

                        # 달력클릭
                        WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#body_center > table:nth-child(5) > tbody > tr:nth-child(2) > td:nth-child(2) > a:nth-child(3)")))).click()

                        days_vis = datetime.timedelta(days=ii)
                        before_day_vis = today-days_vis
                        

                        #시작
                        #년도 선택
                        before_year = (before_day_vis).strftime("%Y")
                        select_element = driver.find_element(By.CSS_SELECTOR, '#yearSB_1')
                        select = Select(select_element)
                        select.select_by_value(before_year)

                        #달 선택
                        before_month = str(int((before_day_vis).strftime("%m")))
                        select_element = driver.find_element(By.CSS_SELECTOR, '#monthSB_1')
                        select = Select(select_element)
                        select.select_by_value(before_month)

                        #일 선택
                        before_day1 = str(int((before_day_vis).strftime("%d")))
                        for i in range(1, 43):
                            try:
                                element = driver.find_element(By.ID, f'li_{i}')
                                if element.text == before_day1:
                                    element.click()
                                    print("before_day1 clicked")
                                    break

                            except:
                                print(f'li_{i} not found')


                        #끝
                        #년도 선택
                        select_element = driver.find_element(By.CSS_SELECTOR, '#yearSB_2')
                        select = Select(select_element)
                        select.select_by_value(before_year)

                        #달 선택
                        select_element = driver.find_element(By.CSS_SELECTOR, '#monthSB_2')
                        select = Select(select_element)
                        select.select_by_value(before_month)

                        #일 선택
                        for i in range(1, 43):
                            try:
                                element = driver.find_element(By.ID, f'le_{i}')
                                if element.text == before_day1:
                                    element.click()
                                    print("before_day1 clicked")
                                    break

                            except:
                                print(f'le_{i} not found')

                        # 조회
                        driver.find_element(By.CSS_SELECTOR, "#body_center > table:nth-child(5) > tbody > tr:nth-child(2) > td:nth-child(2) > a:nth-child(4) > img").click()

                        visitors = driver.find_elements(By.ID, "summary_pfm_total")
                        for num in visitors:
                            the_num = driver.find_element(By.CSS_SELECTOR, "#summary_pfm_total > td:nth-child(2)").text
                            print(the_num)

                        # 서비스 계정 키 파일 경로
                        credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                        # gspread 클라이언트 초기화
                        client = gspread.service_account(filename=credential_file)

                        # Google 시트 열기
                        spreadsheet = client.open_by_url(sheet_urlR)

                        # 첫 번째 시트 선택
                        sheet = spreadsheet.worksheet(sheet_nameD)
                        todayy = today.strftime("%Y-%m-%d")
                        column_values = sheet.col_values(1)
                        for idx, cell_value in enumerate(column_values, start=1):  # start=1로 설정하여 행 번호를 1부터 시작
                            if cell_value == todayy:
                                print(cell_value)
                                print(gspread.utils.rowcol_to_a1(idx, 1))
                                cell_addr = gspread.utils.rowcol_to_a1(idx, 1)
                                # return f"{gspread.utils.rowcol_to_a1(idx, 1)}"  # 셀 주소 반환
                            
                        (start_row, start_col) = gspread.utils.a1_to_rowcol(cell_addr)

                        # Google 시트에 데이터 쓰기
                        numeric_value = int(the_num.replace(',', ''))
                        range_to_write = f'C{start_row-ii}'
                        sheet.update([[numeric_value]], range_to_write)

                    driver.close()
                    driver.switch_to.window(driver.window_handles[0]) #팝업으로 제어 변경
                    

                #카페24 하엔
                if self.chk_cafe_haen.isChecked() == True:

                    label = self.chk_cafe_haen
                    url_cafe24 = "https://eclogin.cafe24.com/Shop/"
                    url_cafe24_req_haen = "https://woo8425.cafe24.com/disp/admin/shop1/report/DailyList"
                    
                    cafe24_id_haen = self.login_info("CAFE_HAEN_ID")
                    cafe24_pw_haen = self.login_info("CAFE_HAEN_PW")

                    sheet_haenR_url = 'https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=1894651086'
                    sheet_haenR = '하엔R'
                    sheet_haenD = "하엔D"
                
                    cafe24(label, url_cafe24, url_cafe24_req_haen, cafe24_id_haen, cafe24_pw_haen, sheet_haenR_url, sheet_haenR, sheet_haenD)

                #카페24 러블로
                if self.chk_cafe_lovl.isChecked() == True:

                    label = self.chk_cafe_lovl
                    url_cafe24 = "https://eclogin.cafe24.com/Shop/" 
                    url_cafe24_req_lovelo = "https://wooo8425.cafe24.com/disp/admin/shop1/report/DailyList"

                    cafe24_id_lovelo = self.login_info("CAFE_LOVE_ID")
                    cafe24_pw_lovelo = self.login_info("CAFE_LOVE_PW")

                    sheet_loveR_url = 'https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=872830966'
                    sheet_loveR = '러블로R'
                    sheet_loveD = "러블로D"

                    cafe24(label, url_cafe24, url_cafe24_req_lovelo, cafe24_id_lovelo, cafe24_pw_lovelo, sheet_loveR_url, sheet_loveR, sheet_loveD)

                #카페24 노마셀
                if self.chk_cafe_know.isChecked() == True:

                    label = self.chk_cafe_know
                    url_cafe24 = "https://eclogin.cafe24.com/Shop/" 
                    url_cafe24_req_knowmycell = "https://fkark12.cafe24.com/disp/admin/shop1/report/DailyList"

                    cafe24_id_knowmycell = self.login_info("CAFE_KNOW_ID")
                    cafe24_pw_knowmycell = self.login_info("CAFE_KNOW_PW")

                    sheet_knowR_url = 'https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=567505346'
                    sheet_knowR = '노마셀R'
                    sheet_knowD = "노마셀D"

                    cafe24(label, url_cafe24, url_cafe24_req_knowmycell, cafe24_id_knowmycell, cafe24_pw_knowmycell, sheet_knowR_url, sheet_knowR, sheet_knowD)

                                
                #카페24 제니크
                if self.chk_cafe_ZQ.isChecked() == True:

                    label = self.chk_cafe_ZQ
                    url_cafe24 = "https://eclogin.cafe24.com/Shop/" 
                    url_cafe24_req_ZQ = "https://fkark08.cafe24.com/disp/admin/shop1/report/DailyList"

                    cafe24_id_ZQ = self.login_info("CAFE_ZQ_ID")
                    cafe24_pw_ZQ = self.login_info("CAFE_ZQ_PW")

                    sheet_ZQR_url = 'https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=567505346'
                    sheet_ZQR = '제니크R'
                    sheet_ZQD = "제니크D"

                    cafe24(label, url_cafe24, url_cafe24_req_ZQ, cafe24_id_ZQ, cafe24_pw_ZQ, sheet_ZQR_url, sheet_ZQR, sheet_ZQD)

                    ##################################### 파컨
                    ##################################### 파컨
                    ##################################### 파컨
                    ##################################### 파컨

                def power(url, url2, id, pw, sheetUrl, sheetName, key, key2, brand):

                    # 서비스 계정 키 파일 경로
                    credential_file = 'triple-nectar-412808-da4dac0cc16e.json'
                    # gspread 클라이언트 초기화
                    client = gspread.service_account(filename=credential_file)
                    # Google 시트 열기
                    spreadsheet = client.open_by_url(sheetUrl)
                    # 첫 번째 시트 선택
                    sheet = spreadsheet.worksheet(sheetName)

                    driver.get(url)

### 로그인
                    # ID
                    input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#mall_id")))
                    input_field.click()
                    time.sleep(1)
                    input_field.send_keys(Keys.CONTROL + "a")
                    input_field.send_keys(Keys.BACKSPACE)
                    driver.find_element(By.CSS_SELECTOR, "#mall_id").send_keys(id)

                    # PW
                    input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#userpasswd")))
                    input_field.click()
                    input_field.send_keys(Keys.CONTROL + "a")
                    input_field.send_keys(Keys.BACKSPACE)
                    driver.find_element(By.CSS_SELECTOR, "#userpasswd").send_keys(pw)

                    # 로그인클릭
                    driver.find_element(By.CSS_SELECTOR,'#frm_user > div > div.mButton > button').click()

                    #비밀번호변경안내
                    try: WebDriverWait(driver, 5).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#iptBtnEm")))).click() 
                    except: pass

                    #화면로딩대기
                    WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, '//*[contains(text(), "오늘의 할 일")]')))

### 데이터 화면 접근
                    driver.get(url2)
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#mCSB_2_container > ul:nth-child(1) > li:nth-child(9)"))).click() #통계 클릭
                    WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#mCSB_3_container > div.depthList > ul > li:nth-child(8)"))).click() #접속통계클릭

                    #새 창 대기
                    current_window_handle = driver.current_window_handle

                    new_window_handle = None
                    while not new_window_handle:
                        for handle in driver.window_handles:
                            if handle != current_window_handle:
                                new_window_handle = handle
                                break


                    #팝업으로 제어 변경
                    driver.switch_to.window(driver.window_handles[1]) 

                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#rpt_pth")))).click() #방문경로분석
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#rpt_pth_dod")))).click() #방문도메인(상세)

### 데이터 검색
                    # 어제 클릭
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#body_center > table:nth-child(5) > tbody > tr:nth-child(2) > td:nth-child(3) > a:nth-child(2) > img")))).click()

                    updates = []
                    formats = []
                    dayUpdates = []

                    for ii in range(target_days_input, 0, -1):

                        # 달력클릭
                        WebDriverWait(driver, 15).until(EC.element_to_be_clickable(((By.CSS_SELECTOR,"#body_center > table:nth-child(5) > tbody > tr:nth-child(2) > td:nth-child(2) > a:nth-child(3)")))).click()

                        days_vis = datetime.timedelta(days=ii)
                        before_day_vis = today-days_vis
                        
                        #시작
                        #년도 선택
                        before_year = (before_day_vis).strftime("%Y")
                        select_element = driver.find_element(By.CSS_SELECTOR, '#yearSB_1')
                        select = Select(select_element)
                        select.select_by_value(before_year)

                        #달 선택
                        before_month = str(int((before_day_vis).strftime("%m")))
                        select_element = driver.find_element(By.CSS_SELECTOR, '#monthSB_1')
                        select = Select(select_element)
                        select.select_by_value(before_month)

                        #일 선택
                        before_day1 = str(int((before_day_vis).strftime("%d")))
                        for i in range(1, 43):
                            try:
                                element = driver.find_element(By.ID, f'li_{i}')
                                if element.text == before_day1:
                                    element.click()
                                    print("before_day1 clicked")
                                    break

                            except:
                                print(f'li_{i} not found')


                        #끝
                        #년도 선택
                        select_element = driver.find_element(By.CSS_SELECTOR, '#yearSB_2')
                        select = Select(select_element)
                        select.select_by_value(before_year)

                        #달 선택
                        select_element = driver.find_element(By.CSS_SELECTOR, '#monthSB_2')
                        select = Select(select_element)
                        select.select_by_value(before_month)

                        #일 선택
                        for i in range(1, 43):
                            try:
                                element = driver.find_element(By.ID, f'le_{i}')
                                if element.text == before_day1:
                                    element.click()
                                    print("before_day1 clicked")
                                    break

                            except:
                                print(f'le_{i} not found')

                        keywords = [key, key2]
                        # 검색어 입력(NV, NPO, GS)
                        for item in keywords:
                            print(item, "검색")
                            
                            search = driver.find_element(By.CSS_SELECTOR, "#body_center > table:nth-child(13) > tbody > tr > td:nth-child(1) > input")
                            search.send_keys(Keys.CONTROL + "a")
                            search.send_keys(Keys.BACKSPACE)
                            search.click()
                            search.send_keys(item)
                            driver.find_element(By.CSS_SELECTOR, "#body_center > table:nth-child(13) > tbody > tr > td:nth-child(1) > a > img").click()
                            
                            # 조회
                            driver.find_element(By.CSS_SELECTOR, "#body_center > table:nth-child(5) > tbody > tr:nth-child(2) > td:nth-child(2) > a:nth-child(4) > img").click()

                            last_row = len(sheet.get_all_values())
                            print(last_row)
                            next_row = last_row + 1  # 다음 행 번호

                            # 데이터 불러오기
                            for i in range(1,30):
                                
                                try:
                                    
                                    line = driver.find_element(By.CSS_SELECTOR, f"#detail_pfm_total > tr:nth-child({i})").text
                                    lineSplit = line.strip().split(" ")
                                    print(lineSplit)

                                    

                                    def convert_data(data):
                                        result = []
                                        for item in data:
                                            if isinstance(item, str) and '%' in item:
                                                result.append(float(item.strip('%')) / 100)
                                            elif isinstance(item, str) and ',' in item:
                                                result.append(int(item.replace(',', '')))
                                            elif item.isdigit():
                                                result.append(int(item))
                                            else:
                                                result.append(item)
                                        return result

                                    # 입력할 데이터
                                    converted_data = convert_data(lineSplit)

                                    # 구글 시트에서 퍼센트 형식으로 변경하기 위해 셀 범위를 지정
                                    # 예시에서는 C1, E1 셀을 퍼센트 형식으로 설정

                                    range_to_write_day = f"A{next_row}"
                                        
                                    # batch로 입력할 날짜 모으기
                                    dayUpdates.append({'range': range_to_write_day, 'values': [[str(before_day_vis)]]})

                                    # data 입력 범위
                                    range_to_write = f'B{next_row}:I{next_row}'

                                    # batch로 입력할 data 모으기
                                    updates.append({'range': range_to_write, 'values': [converted_data]})

                                    print("OK")

                                    # 정렬할 format 세팅
                                    formats.append({
                                    'range': f"D{next_row}",
                                    'format': {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}}
                                    })
                                    formats.append({
                                    'range': f"F{next_row}",
                                    'format': {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}}
                                    })

                                    next_row += 1

                                except:
                                    if i == 1:
                                        dummyData = [item, '0', '0', '0', '0', '0', '0', '0']

                                        last_row = len(sheet.get_all_values())
                                        print(last_row)
                                        next_row = last_row + 1  # 다음 행 번호

                                        def convert_data(data):
                                            result = []
                                            for item in data:
                                                if isinstance(item, str) and '%' in item:
                                                    result.append(float(item.strip('%')) / 100)
                                                elif isinstance(item, str) and ',' in item:
                                                    result.append(int(item.replace(',', '')))
                                                elif item.isdigit():
                                                    result.append(int(item))
                                                else:
                                                    result.append(item)
                                            return result

                                        # 입력할 데이터
                                        converted_data = convert_data(dummyData)

                                        # 구글 시트에서 퍼센트 형식으로 변경하기 위해 셀 범위를 지정
                                        # 예시에서는 C1, E1 셀을 퍼센트 형식으로 설정
                                        sheet.format(f"D{next_row}", {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}})
                                        sheet.format(f"F{next_row}", {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}})
                                            
                                        sheet.update([[str(before_day_vis)]], f"A{next_row}")
                                        range_to_write = f'B{next_row}:I{next_row}'
                                        sheet.update([converted_data], range_to_write)
                                        print("OK")
                                        break

                                if updates:
                                    print(updates)
                                    print(dayUpdates)
                                    sheet.batch_update(updates)
                                    sheet.batch_update(dayUpdates)
                                    for fmt in formats:
                                        sheet.format(fmt['range'], fmt['format'])
                                    updates.clear()
                                    dayUpdates.clear()
                                    formats.clear()
                                    time.sleep(0.5)  # 각 배치 요청 사이에 지연 시간을 추가
                                    print("Batch update and format applied.")

                    driver.close()
                    time.sleep(0.1)
                    
                    driver.switch_to.window(driver.window_handles[0])



                url_cafe24 = "https://eclogin.cafe24.com/Shop/" 
                
                if self.chk_pc_haen.isChecked() == True:
                    cafe24_id_haen = self.login_info("CAFE_HAEN_ID")
                    cafe24_pw_haen = self.login_info("CAFE_HAEN_PW")
                    url2 = "https://woo8425.cafe24.com/disp/admin/shop1/report/DailyList"

                    sheetUrl_haen = "https://docs.google.com/spreadsheets/d/1V8b3FRe_8witwHXQceekgm-BAvTQLwwkcyuaW-mIi30/edit#gid=128334801"
                    sheetName_haenPCR = "하엔 파워콘텐츠 R"
                    Keyword = "NPO"
                    Keyword2 = "GS"
                    brand = "하엔"

                    power(url_cafe24, url2, cafe24_id_haen, cafe24_pw_haen, sheetUrl_haen, sheetName_haenPCR, Keyword, Keyword2, brand)


                if self.chk_pc_lovl.isChecked() == True:
                    cafe24_id_love = self.login_info("CAFE_LOVE_ID")
                    cafe24_pw_love = self.login_info("CAFE_LOVE_PW")
                    url2 = "https://wooo8425.cafe24.com/disp/admin/shop1/report/DailyList"

                    sheetUrl_love = "https://docs.google.com/spreadsheets/d/1NVnVJsCj0Ap_o2xabua9ANUw_1IUREVMJKteY_O1yks/edit#gid=311448069"
                    sheetName_lovePCR = "러블로 파워콘텐츠 R"
                    Keyword = "NV"
                    Keyword2 = "GS"
                    brand = "러블로"

                    power(url_cafe24, url2, cafe24_id_love, cafe24_pw_love, sheetUrl_love, sheetName_lovePCR, Keyword, Keyword2, brand)


                if self.chk_pc_know.isChecked() == True:
                    cafe24_id_know = self.login_info("CAFE_KNOW_ID")
                    cafe24_pw_know = self.login_info("CAFE_KNOW_PW")
                    url2 = "https://fkark12.cafe24.com/disp/admin/shop1/report/DailyList"

                    sheetUrl_know = "https://docs.google.com/spreadsheets/d/12FWmZMuznsxOY_IDbBWeSis-EW1Ds1f9TB6X7K6TFBc/edit#gid=1722856727"
                    sheetName_knowPCR = "노마셀 파워콘텐츠 R"
                    Keyword = "NV"
                    Keyword2 = "GS"
                    brand = "노마셀"

                    power(url_cafe24, url2, cafe24_id_know, cafe24_pw_know, sheetUrl_know, sheetName_knowPCR, Keyword, Keyword2,brand)


                # 쿠팡
                def coupang(url_coupang_daily, coupang_id, coupang_pw, coupC_url):
                    try:
                        driver.get(url_coupang_daily)  # 로그인 시작
                        if driver.find_element(By.CSS_SELECTOR, "body > pre"):
                            driver.get(url_coupang_daily)  # 요소가 존재하면 페이지를 다시 로드
                    except NoSuchElementException:
                        # 요소가 없을 때 처리할 로직
                        pass

                    try:
                        driver.get(url_coupang_daily)  # 로그인 시작
                        if driver.find_element(By.CSS_SELECTOR, "body > h1"):
                            driver.get(url_coupang_daily)  # 요소가 존재하면 페이지를 다시 로드
                    except NoSuchElementException:
                        pass
                    
                    try:
                        loginElements = driver.find_elements(By.XPATH, '//*[contains(text(), "로그인하기")]')

                        if len(loginElements) > 1:  # 요소가 두 개 이상 있는지 확인
                            
                            loginElements[0].click()

                        # if driver.find_element(By.CSS_SELECTOR, "#main-container > div > div.sc-30ec2de1-0.tedRR > ul > li:nth-child(1) > a > span"):
                        #     driver.find_element(By.CSS_SELECTOR, "#main-container > div > div.sc-30ec2de1-0.tedRR > ul > li:nth-child(1) > a > span").click()

                    except NoSuchElementException:
                        # 요소가 없을 때 처리할 로직
                        pass

                    print(coupang_id)
                    print(coupang_pw)
                    input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#username")))
                    input_field.click()
                    time.sleep(0.7)
                    input_field.send_keys(Keys.CONTROL + "a")
                    input_field.send_keys(Keys.BACKSPACE)
                    driver.find_element(By.CSS_SELECTOR, "#username").send_keys(coupang_id)
                    input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#password")))
                    input_field.click()
                    input_field.send_keys(Keys.CONTROL + "a")
                    input_field.send_keys(Keys.BACKSPACE)
                    driver.find_element(By.CSS_SELECTOR, "#password").send_keys(coupang_pw)
                    driver.find_element(By.CSS_SELECTOR,'#kc-login').click()

                    # 로그인 오류 발생하면 재시도
                    ### 비밀번호 오류 예외문
                    try:
                        loginErrorMessage = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, '//*[contains(text(), "아이디 또는 비밀번호가 다릅니다.")]')))
                        if loginErrorMessage:
                            input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#username")))
                            input_field.click()
                            time.sleep(0.7)
                            input_field.send_keys(Keys.CONTROL + "a")
                            input_field.send_keys(Keys.BACKSPACE)
                            driver.find_element(By.CSS_SELECTOR, "#username").send_keys(coupang_id)
                            input_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#password")))
                            input_field.click()
                            input_field.send_keys(Keys.CONTROL + "a")
                            input_field.send_keys(Keys.BACKSPACE)
                            driver.find_element(By.CSS_SELECTOR, "#password").send_keys(coupang_pw)
                            driver.find_element(By.CSS_SELECTOR,'#kc-login').click()
                        

                    except: pass

                    try:
                        WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#startDateId"))) #클릭 시작일
                    except:
                        driver.refresh()
                        WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#startDateId"))) #클릭 시작일
                    driver.find_element(By.CSS_SELECTOR, "#startDateId").click()

                    before_Ym = today_tday.strftime("%Y년 %m월")
                    before_d = str(int(today_tday.strftime("%d")))
                    yesterday_Ym = today_yday.strftime("%Y년 %m월")
                    yesterday_d = str(int(today_yday.strftime("%d")))

                    firstCal = driver.find_elements(By.XPATH, "//*[@id='ad-reporting-app']/div[2]/div/div[2]/div[1]/div[2]/div[3]/div[2]/div[2]/div[1]/div/div/div[2]/div/div/div[1]/div[2]/div[2]/div/div[2]")
                    secondCal = driver.find_elements(By.XPATH, "//*[@id='ad-reporting-app']/div[2]/div/div[2]/div[1]/div[2]/div[3]/div[2]/div[2]/div[1]/div/div/div[2]/div/div/div[1]/div[2]/div[2]/div/div[3]")

                    # 시작날짜
                    try:
                        for i in firstCal:
                            # 텍스트를 줄 단위로 나누기
                            lines = (i.text).strip().split('\n')
                            if lines[0] == today_Tday년월:
                                print("OK")
                                i.find_element(By.XPATH, f"""//*[@id='ad-reporting-app']/div[2]/div/div[2]/div[1]/div[2]/div[3]/div[2]/div[2]/div[1]/div/div/div[2]/div/div/div[1]/div[2]/div[2]/
                                div/div[2]/div/table//*[text()='{today_Tday일}']""").click()

                        for i in secondCal:
                            # 텍스트를 줄 단위로 나누기
                            lines = (i.text).strip().split('\n')
                            if lines[0] == today_Tday년월:
                                print("OK")
                                i.find_element(By.XPATH, f"//*[@id='ad-reporting-app']/div[2]/div/div[2]/div[1]/div[2]/div[3]/div[2]/div[2]/div[1]/div/div/div[2]/div/div/div[1]/div[2]/div[2]/div/div[3]/div/table//*[text()='{today_Tday일}']").click()

                    except: pass
                        

                            
                    time.sleep(0.1)

                    # 종료날짜
                    try:
                        for i in firstCal:
                            # 텍스트를 줄 단위로 나누기
                            lines = (i.text).strip().split('\n')
                            if lines[0] == today_Yday년월:
                                print("OK")
                                i.find_element(By.XPATH, f"//*[@id='ad-reporting-app']/div[2]/div/div[2]/div[1]/div[2]/div[3]/div[2]/div[2]/div[1]/div/div/div[2]/div/div/div[1]/div[2]/div[2]/div/div[2]/div/table//*[text()='{today_Yday일}']").click()

                        for i in secondCal:
                            # 텍스트를 줄 단위로 나누기
                            lines = (i.text).strip().split('\n')
                            if lines[0] == today_Yday년월:
                                print("OK")
                                i.find_element(By.XPATH, f"//*[@id='ad-reporting-app']/div[2]/div/div[2]/div[1]/div[2]/div[3]/div[2]/div[2]/div[1]/div/div/div[2]/div/div/div[1]/div[2]/div[2]/div/div[3]/div/table//*[text()='{today_Yday일}']").click()

                    except: pass
                    element = driver.find_element(By.CSS_SELECTOR, '#ad-reporting-app > div.self-service-ad-reporting-ui > div > div.sc-11l2gxs-0.fcpsUc > div.sc-ipia07-0.iCqAxH > div.panel-options > div.sc-19odvm9-0.kgfJLF > div.select-date-group')#기간 구분
                    element.click() 
                    ActionChains(driver).move_to_element_with_offset(element,5,75).click().perform() #클릭 일별
                    time.sleep(0.3)

                    driver.find_element(By.CSS_SELECTOR,'#ad-reporting-app > div.self-service-ad-reporting-ui > div > div.sc-11l2gxs-0.fcpsUc > div.sc-ipia07-0.iCqAxH > div.panel-options > div.sc-1jpf51e-0.hSjByk > div > div.campaign-picker-container > div > button > span.text').click() #캠페인 선택
                    time.sleep(0.5)
                    checkbox = driver.find_element(By.CSS_SELECTOR, "body > div.sc-1jpf51e-1.jljGiJ.popper > div > div.select-all-campaigns > label > span.ant-checkbox > input[type='checkbox']")
                    if not checkbox.is_selected():
                        checkbox.click()  # 체크박스가 체크되어 있지 않다면 클릭하여 체크합니다.
                    driver.find_element(By.CSS_SELECTOR, "body > div.sc-1jpf51e-1.jljGiJ.popper > div > div.button-container > button.ant-btn.ant-btn-primary > span").click()
                    time.sleep(0.3)

### 보고서 생성 실패하면 페이지 다시 로딩 후 생성
                    try:
                        try:
                            driver.find_element(By.CSS_SELECTOR, "#generateReport > span").click() #보고서 생성

                        except:
                            driver.find_element(By.CSS_SELECTOR,'#ad-reporting-app > div.self-service-ad-reporting-ui > div > div.sc-11l2gxs-0.fcpsUc > div.sc-ipia07-0.iCqAxH > div.panel-options > div.sc-1jpf51e-0.hSjByk > div > div.campaign-picker-container > div > button > span.text').click() #캠페인 선택
                            time.sleep(0.5)
                            checkbox = driver.find_element(By.CSS_SELECTOR, "body > div.sc-1jpf51e-1.jljGiJ.popper > div > div.select-all-campaigns > label > span.ant-checkbox > input[type='checkbox']")

                            if not checkbox.is_selected():
                                checkbox.click()  # 체크박스가 체크되어 있지 않다면 클릭하여 체크합니다.

                            driver.find_element(By.CSS_SELECTOR, "body > div.sc-1jpf51e-1.jljGiJ.popper > div > div.button-container > button.ant-btn.ant-btn-primary > span").click()
                            driver.find_element(By.CSS_SELECTOR, "#generateReport > span").click() #보고서 생성

                        time.sleep(1.5)

                        if driver.find_element(By.CSS_SELECTOR, "#rc-tabs-0-panel-requestedReport > div > div.react-grid-Container > div > div > div:nth-child(2) > div > div > div:nth-child(2) > div:nth-child(1) > div:nth-child(5) > div > div > span > div").text == "생성 실패":
                            driver.find_element(By.CSS_SELECTOR, "#generateReport > span").click() 
                            time.sleep(5)

                        

                        # 보고서 다운로드
                        element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#rc-tabs-0-panel-requestedReport > div > div.react-grid-Container > div > div > div:nth-child(2) > div > div > div:nth-child(2) > div:nth-child(1) > div:nth-child(6) > div > div > span > div > div:nth-child(2) > button > span"))) 

                        # 다운로드 확인
                        cnt = 1
                        current_file_count1 = count_files(download_folder)
                        while cnt < 10:
                            element.click()
                            time.sleep(3)
                            current_file_count2 = count_files(download_folder)
                            if current_file_count1 != current_file_count2:
                                break
                            elif cnt == 300:
                                break

                            cnt += 1

                        # check_download()
                        time.sleep(1)
                        try:
                            WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "body > div.MuiDialog-root.sc-852clq-0.efPzRF > div.MuiDialog-container.MuiDialog-scrollPaper > div > div:nth-child(3) > button"))).click()
                        except: pass

                    except:
                        driver.get(url_coupang_daily)

                        try:
                            WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#startDateId"))) #클릭 시작일
                        except:
                            driver.refresh()
                            WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#startDateId"))) #클릭 시작일
                        driver.find_element(By.CSS_SELECTOR, "#startDateId").click()

                        before_Ym = today_tday.strftime("%Y년 %m월")
                        before_d = str(int(today_tday.strftime("%d")))
                        yesterday_Ym = today_yday.strftime("%Y년 %m월")
                        yesterday_d = str(int(today_yday.strftime("%d")))

                        firstCal = driver.find_elements(By.XPATH, "//*[@id='ad-reporting-app']/div[2]/div/div[2]/div[1]/div[2]/div[3]/div[2]/div[2]/div[1]/div/div/div[2]/div/div/div[1]/div[2]/div[2]/div/div[2]")
                        secondCal = driver.find_elements(By.XPATH, "//*[@id='ad-reporting-app']/div[2]/div/div[2]/div[1]/div[2]/div[3]/div[2]/div[2]/div[1]/div/div/div[2]/div/div/div[1]/div[2]/div[2]/div/div[3]")

                        # 시작날짜
                        try:
                            for i in firstCal:
                                # 텍스트를 줄 단위로 나누기
                                lines = (i.text).strip().split('\n')
                                if lines[0] == today_Tday년월:
                                    print("OK")
                                    i.find_element(By.XPATH, f"""//*[@id='ad-reporting-app']/div[2]/div/div[2]/div[1]/div[2]/div[3]/div[2]/div[2]/div[1]/div/div/div[2]/div/div/div[1]/div[2]/div[2]/
                                    div/div[2]/div/table//*[text()='{today_Tday일}']""").click()

                            for i in secondCal:
                                # 텍스트를 줄 단위로 나누기
                                lines = (i.text).strip().split('\n')
                                if lines[0] == today_Tday년월:
                                    print("OK")
                                    i.find_element(By.XPATH, f"//*[@id='ad-reporting-app']/div[2]/div/div[2]/div[1]/div[2]/div[3]/div[2]/div[2]/div[1]/div/div/div[2]/div/div/div[1]/div[2]/div[2]/div/div[3]/div/table//*[text()='{today_Tday일}']").click()

                        except: pass
                            

                                
                        time.sleep(0.1)

                        # 종료날짜
                        try:
                            for i in firstCal:
                                # 텍스트를 줄 단위로 나누기
                                lines = (i.text).strip().split('\n')
                                if lines[0] == today_Yday년월:
                                    print("OK")
                                    i.find_element(By.XPATH, f"//*[@id='ad-reporting-app']/div[2]/div/div[2]/div[1]/div[2]/div[3]/div[2]/div[2]/div[1]/div/div/div[2]/div/div/div[1]/div[2]/div[2]/div/div[2]/div/table//*[text()='{today_Yday일}']").click()

                            for i in secondCal:
                                # 텍스트를 줄 단위로 나누기
                                lines = (i.text).strip().split('\n')
                                if lines[0] == today_Yday년월:
                                    print("OK")
                                    i.find_element(By.XPATH, f"//*[@id='ad-reporting-app']/div[2]/div/div[2]/div[1]/div[2]/div[3]/div[2]/div[2]/div[1]/div/div/div[2]/div/div/div[1]/div[2]/div[2]/div/div[3]/div/table//*[text()='{today_Yday일}']").click()

                        except: pass
                        element = driver.find_element(By.CSS_SELECTOR, '#ad-reporting-app > div.self-service-ad-reporting-ui > div > div.sc-11l2gxs-0.fcpsUc > div.sc-ipia07-0.iCqAxH > div.panel-options > div.sc-19odvm9-0.kgfJLF > div.select-date-group')#기간 구분
                        element.click() 
                        ActionChains(driver).move_to_element_with_offset(element,5,75).click().perform() #클릭 일별
                        time.sleep(0.3)

                        driver.find_element(By.CSS_SELECTOR,'#ad-reporting-app > div.self-service-ad-reporting-ui > div > div.sc-11l2gxs-0.fcpsUc > div.sc-ipia07-0.iCqAxH > div.panel-options > div.sc-1jpf51e-0.hSjByk > div > div.campaign-picker-container > div > button > span.text').click() #캠페인 선택
                        time.sleep(0.5)
                        checkbox = driver.find_element(By.CSS_SELECTOR, "body > div.sc-1jpf51e-1.jljGiJ.popper > div > div.select-all-campaigns > label > span.ant-checkbox > input[type='checkbox']")
                        if not checkbox.is_selected():
                            checkbox.click()  # 체크박스가 체크되어 있지 않다면 클릭하여 체크합니다.
                        driver.find_element(By.CSS_SELECTOR, "body > div.sc-1jpf51e-1.jljGiJ.popper > div > div.button-container > button.ant-btn.ant-btn-primary > span").click()
                        time.sleep(0.3)

                        driver.find_element(By.CSS_SELECTOR, "#generateReport > span").click() #보고서 생성

                        # 보고서 다운로드
                        element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#rc-tabs-0-panel-requestedReport > div > div.react-grid-Container > div > div > div:nth-child(2) > div > div > div:nth-child(2) > div:nth-child(1) > div:nth-child(6) > div > div > span > div > div:nth-child(2) > button > span"))) 

                        # 다운로드 확인
                        cnt = 1
                        while cnt < 10:
                            current_file_count1 = count_files(download_folder)
                            element.click()
                            time.sleep(3)
                            current_file_count2 = count_files(download_folder)
                            if current_file_count1 != current_file_count2:
                                break

                            cnt += 1

                        # check_download()
                        time.sleep(1)
                        # 알림창 제거
                        try:
                            WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "body > div.MuiDialog-root.sc-852clq-0.efPzRF > div.MuiDialog-container.MuiDialog-scrollPaper > div > div:nth-child(3) > button"))).click()
                        except: pass



                def coupangC(url, url_sheet, name_sheet, brand):

                    # 로그인 시작
                    driver.get(url)  

                    target_days = target_days_input
                    
                    while target_days > 0:

                        dayx = datetime.timedelta(days=target_days)
                        today_tday_str = (today-dayx).strftime('%Y-%m-%d')

                        try:
                            time.sleep(1)
                            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#search-filter-panel > div:nth-child(1) > span.sc-common-date-group > span:nth-child(1)"))) # 시작 날짜
                            time.sleep(1)
                            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#search-filter-panel > div:nth-child(1) > span.sc-common-date-group > span:nth-child(1)"))).click() # 시작 날짜
                        except:
                            driver.get(coupC_url)
                            time.sleep(1)
                            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#search-filter-panel > div:nth-child(1) > span.sc-common-date-group > span:nth-child(1)"))) # 시작 날짜
                            time.sleep(1)
                            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#search-filter-panel > div:nth-child(1) > span.sc-common-date-group > span:nth-child(1)"))).click() # 시작 날짜


                        input_field = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#dateEnd")))
                        time.sleep(0.1)
                        input_field.send_keys(Keys.CONTROL + "a")
                        input_field.send_keys(Keys.BACKSPACE)
                        input_field.send_keys(today_tday_str)
                        
                        input_field = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#dateStart")))
                        time.sleep(0.1)
                        input_field.send_keys(Keys.CONTROL + "a")
                        input_field.send_keys(Keys.BACKSPACE)
                        input_field.send_keys(today_tday_str)


                        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#date_range"))) # 날짜변경확인

                        time.sleep(1)

                        # 다운로드 확인
                        cnt = 1
                        while cnt < 10:
                            current_file_count1 = count_files(download_folder)
                            time.sleep(3)
                            try:
                                WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#download-product-info"))).click() # 다운로드
                                driver.find_element(By.CSS_SELECTOR, "#download-product-info").click()
                            except:
                                WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#download-product-info"))).click() # 다운로드
                                driver.find_element(By.CSS_SELECTOR, "#download-product-info").click()
                            current_file_count2 = count_files(download_folder)
                            if current_file_count1 != current_file_count2:
                                break

                            cnt += 1
                            # check_download()
                        time.sleep(1)

                        xlsx_file = get_latest_file(download_folder)

                        df_uploaded_new = pd.read_excel(xlsx_file)
                        # '러브슬라임'이라는 단어가 포함된 모든 행을 '옵션명' 열을 기준으로 필터링합니다.
                        filtered_rows_with_loveslime = df_uploaded_new[df_uploaded_new['옵션명'].astype(str).str.contains(brand)]

                        # 필터링된 행들의 데이터를 리스트로 변환합니다.
                        rows_list_with_loveslime = filtered_rows_with_loveslime.values.tolist()


                        # 두 번째 값만 정수형으로 변환한 후 문자열로 변환하여 업데이트하는 과정
                        updated_data_list = []
                        for row in rows_list_with_loveslime:
                            new_row = row.copy()  # 원본 데이터의 복사본 생성
                            if len(row) > 1:  # 두 번째 값이 존재하는지 확인
                                new_row[1] = str(int(row[1]))  # 두 번째 값을 정수형으로 변환 후 문자열로 변환
                            updated_data_list.append(new_row)

                        # 결과 출력
                        print(updated_data_list)

                        # 서비스 계정 키 파일 경로
                        credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                        # gspread 클라이언트 초기화
                        client = gspread.service_account(filename=credential_file)

                        # Google 시트 열기
                        spreadsheet = client.open_by_url(url_sheet)

                        # 첫 번째 시트 선택
                        sheet = spreadsheet.worksheet(name_sheet)

                        last_row = len(sheet.get_all_values())
                        print(last_row)
                        next_row = last_row + 1  # 다음 행 번호
                            
                        # Google 시트에 데이터 쓰기
                        if len(updated_data_list) > 1:
                            i = 0
                            while i < len(updated_data_list):
                                range_to_write = f'B{next_row+i}:N{next_row+i}'
                                sheet.update([updated_data_list[i]], range_to_write)
                                sheet.update([[today_tday_str]], f'A{next_row+i}')
                                i += 1
                        else:
                            range_to_write = f'B{next_row}:N{next_row}'
                            sheet.update([updated_data_list[0]], range_to_write)
                            sheet.update([[today_tday_str]], f'A{next_row}')

                        target_days -= 1
                    
                    # 쿠팡 로그아웃
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#top-header-menu > div.top-header-control.etc-buttons > ul > li.my-user-menu > div'))).click()
                    time.sleep(0.2)
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#wing-top-header-user-action-layer > div > div.my-user-menu-bottom > span'))).click()

                    time.sleep(2)


                def coupang_rawdata(sheet_url, sheet_nameR, options, sheet_nameC):

                    xlsx_file = get_latest_file(download_folder)

                    df_uploaded_new = pd.read_excel(xlsx_file)
                    # '러브슬라임'이라는 단어가 포함된 모든 행을 '옵션명' 열을 기준으로 필터링합니다.
                    filtered_rows_with_loveslime = df_uploaded_new[df_uploaded_new['입찰유형'].astype(str).str.contains("cpc")]


                    # 필터링된 행들의 데이터를 리스트로 변환합니다.
                    rows_list_with_loveslime = filtered_rows_with_loveslime.values.tolist()

                    excel_dates = []
                    for i in rows_list_with_loveslime:
                        num = rows_list_with_loveslime.index(i)
                        excel_dates.append((rows_list_with_loveslime[num])[0])
                    print(excel_dates)

                    formatted_excel_dates = [datetime.datetime.strptime(str(date), "%Y%m%d").strftime("%Y-%m-%d") for date in excel_dates]

                    print(formatted_excel_dates)

                    updated_data_list = []
                    for row in rows_list_with_loveslime:
                        new_row = row.copy()  # 원본 데이터의 복사본 생성
                        if len(row) > 1:  # 두 번째 값이 존재하는지 확인
                            new_row[1] = str(row[1])  # 두 번째 값을 정수형으로 변환 후 문자열로 변환
                        updated_data_list.append(new_row)

                    print(updated_data_list)


                    # 서비스 계정 키 파일 경로
                    credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                    # gspread 클라이언트 초기화
                    client = gspread.service_account(filename=credential_file)

                    # Google 시트 열기
                    spreadsheet = client.open_by_url(sheet_url)

                    # 첫 번째 시트 선택
                    sheet = spreadsheet.worksheet(sheet_nameR)

                    
                    last_row = len(sheet.get_all_values())
                    print(last_row)
                    next_row = last_row + 1  # 다음 행 번호

                    formatted_date = today_yday.strftime("%Y-%m-%d")
                    # Google 시트에 데이터 쓰기

                    if len(updated_data_list) > 1:
                        i = 0

                        while i < len(updated_data_list):
                            print((updated_data_list[i])[1:])
                            range_to_write = f'B{next_row+i}:AI{next_row+i}'
                            sheet.update([(updated_data_list[i])[1:-1]], range_to_write)
                            sheet.update([[formatted_excel_dates[i]]], f'A{next_row+i}')
                            i += 1

                    else:
                        range_to_write = f'B{next_row}:AI{next_row}'
                        sheet.update([(updated_data_list[0])[1:-1]], range_to_write)
                        sheet.update([[formatted_date]], f'A{next_row}')

                    
                # 쿠팡 하엔
                coupC_url = "https://wing.coupang.com/seller/notification/metrics/dashboard"
                coup_report_url = 'https://advertising.coupang.com/marketing-reporting/billboard/reports/pa'
                sheet_url_coupC = 'https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=374561563'

                
                if self.chk_coup_haen.isChecked() == True:
                    
                    coupang_id_haen = self.login_info("COUP_HAEN_ID")
                    coupang_pw_haen = self.login_info("COUP_HAEN_PW")
                    sheet_url_haen_all = 'https://docs.google.com/spreadsheets/d/1V8b3FRe_8witwHXQceekgm-BAvTQLwwkcyuaW-mIi30/edit#gid=1338112098'
                    sheet_name_haenR = '하엔 쿠팡 R'
                    sheet_name_haenC = '하엔C'
                    options = "하엔"

                    coupang(coup_report_url, coupang_id_haen, coupang_pw_haen, coupC_url)
                    coupang_rawdata(sheet_url_haen_all, sheet_name_haenR, options, sheet_name_haenC)
                    coupangC(coupC_url, sheet_url_coupC, sheet_name_haenC, options)

                # 쿠팡 러블로
                if self.chk_coup_lovl.isChecked() == True:

                    coupang_id_lovelo = self.login_info("COUP_LOVE_ID")
                    coupang_pw_lovelo = self.login_info("COUP_LOVE_PW")
                    sheet_url_love_all = 'https://docs.google.com/spreadsheets/d/1NVnVJsCj0Ap_o2xabua9ANUw_1IUREVMJKteY_O1yks/edit#gid=392530415'
                    sheet_name_loveR = '러블로 쿠팡 R'
                    sheet_name_loveC = '러블로C'
                    options = "러브슬라임"

                    coupang(coup_report_url, coupang_id_lovelo, coupang_pw_lovelo, coupC_url)
                    coupang_rawdata(sheet_url_love_all, sheet_name_loveR, options, sheet_name_loveC)
                    coupangC(coupC_url, sheet_url_coupC, sheet_name_loveC, options)

                # 쿠팡 노마셀
                if self.chk_coup_know.isChecked() == True:
                    coupang_id_knowmycell = self.login_info("COUP_KNOW_ID")
                    coupang_pw_knowmycell = self.login_info("COUP_KNOW_PW")
                    sheet_url_know_all = 'https://docs.google.com/spreadsheets/d/12FWmZMuznsxOY_IDbBWeSis-EW1Ds1f9TB6X7K6TFBc/edit#gid=1042061913'
                    sheet_name_knowR = '노마셀 쿠팡 R'
                    sheet_name_knowC = '노마셀C'
                    options = "노마셀"

                    coupang(coup_report_url, coupang_id_knowmycell, coupang_pw_knowmycell, coupC_url)
                    coupang_rawdata(sheet_url_know_all, sheet_name_knowR, options, sheet_name_knowC)
                    coupangC(coupC_url, sheet_url_coupC, sheet_name_knowC, options)



                #########메타로데이터##########
                def meta_rawdata(sheet_url, sheet_name, know_TF):

                    xlsx_file = get_latest_file(download_folder)
                    wb = load_workbook(xlsx_file)
                    ws = wb.active

                    # 서비스 계정 키 파일 경로
                    credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                    # gspread 클라이언트 초기화
                    client = gspread.service_account(filename=credential_file)

                    # Google 시트 열기
                    spreadsheet = client.open_by_url(sheet_url)

                    # 첫 번째 시트 선택
                    sheet = spreadsheet.worksheet(sheet_name)

                    # if know_TF == 0:
                    #     data_to_paste = []

                    #     targe = ws['4']
                        
                    #     for cell in targe:
                    #         data_to_paste.append(cell.value)
                    #     data_to_paste = data_to_paste[2:]
                    #     print(data_to_paste)

                    #     today = datetime.date.today().strftime("%Y-%m-%d")
                    #     column_values = sheet.col_values(1)
                    #     for idx, cell_value in enumerate(column_values, start=1):  # start=1로 설정하여 행 번호를 1부터 시작
                    #         if cell_value == today:
                    #             print(cell_value)
                    #             print(gspread.utils.rowcol_to_a1(idx, 1))
                    #             cell_addr = gspread.utils.rowcol_to_a1(idx, 1)
                    #             # return f"{gspread.utils.rowcol_to_a1(idx, 1)}"  # 셀 주소 반환
                            
                    #     (start_row, start_col) = gspread.utils.a1_to_rowcol(cell_addr)

                    #     # Google 시트에 데이터 쓰기
                    #     range_to_write = f'B{start_row-1}:P{start_row-1}'
                    #     sheet.update([data_to_paste], range_to_write)
                    
                    # else:
                    data_to_paste = []
                    data_to_pasteDay = []
                    today_tdayTemp = today_tday

                    # 8. 메타 n일전 데이터 불러오기(데이터 없으면 더미데이터 입력)
                    # 두 번째 행이 비어있는지 확인
                    second_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
                    if second_row and all(cell is None for cell in second_row[0]):
                        while today_tdayTemp != today:
                            metaDataEmpty = [str(today_tdayTemp), '-', 0, 0, 0, 0, 0, 0, 0, 0, '-', 0, 0, 0, 0, str(today_tdayTemp), str(today_tdayTemp)]

                            last_row = len(sheet.get_all_values())
                            print(last_row)
                            next_row = last_row + 1  # 다음 행 번호

                            # 데이터 추가
                            range_to_write = f'A{next_row}:Q{next_row}'
                            sheet.update([metaDataEmpty], range_to_write) #한줄

                            today_tdayTemp += timedelta(days=1)
                            

                    else:
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            data_to_paste.append(list(row))
                        data_to_paste.reverse()
                        print(data_to_paste)
                        for i in data_to_paste:
                            data_to_pasteDay.append(i[0])

                        print(data_to_pasteDay)

                        print(today_tdayTemp)
                        print(today_yday)

                        while today_tdayTemp != today_yday:
                            if str(today_tdayTemp) in data_to_pasteDay:
                                num = data_to_pasteDay.index(str(today_tdayTemp))
                                data_to_paste[num]

                                last_row = len(sheet.get_all_values())
                                print(last_row)
                                next_row = last_row + 1  # 다음 행 번호

                                # 데이터 추가
                                range_to_write = f'A{next_row}:Q{next_row}'
                                sheet.update([data_to_paste[num]], range_to_write) #한줄

                            today_tdayTemp += timedelta(days=1)

                    targe1 = ws['2']
                    targe2 = ws['3']
                    data_to_paste_know1 = []
                    data_to_paste_know2 = []
                    for cell in targe1:
                        data_to_paste_know1.append(cell.value)
                    for cell in targe2:
                        data_to_paste_know2.append(cell.value)
                    print(data_to_paste_know1)
                    print(data_to_paste_know2)

                    last_row = len(sheet.get_all_values())
                    print(last_row)
                    next_row = last_row + 1  # 다음 행 번호

                    # 데이터 추가
                    range_to_write = f'A{next_row}:Q{next_row}'
                    sheet.update([data_to_paste_know1], range_to_write) #한줄
                    
                    # K 값 지정
                    K_value = sheet.acell(f'K{next_row}').value
                    K_value_previous = sheet.acell(f'K{next_row-1}').value

                    if sheet.acell(f'A{next_row}').value is not None and K_value is None: #웹사이트 구매 빈칸일 때
                        sheet.update([[K_value_previous]], f'K{next_row}')

                    range_to_write = f'A{next_row+1}:Q{next_row+1}'
                    sheet.update([data_to_paste_know2], range_to_write) #두줄

                    time.sleep(2)
                    K_value = sheet.acell(f'K{next_row+1}').value
                    if sheet.acell(f'A{next_row+1}').value is not None and K_value is None: #웹사이트 구매 빈칸일 때
                        sheet.update([[K_value_previous]], f'K{next_row+1}')

                # 메타
                def meta(url_meta, know_TF):

                    driver.get(url_meta)

                    # meta_id = 'healer10@kakao.com'
                    # meta_pw = 'fhdifxmfl1305!!'

                    # '비밀번호를' 텍스트를 포함하는 요소 찾기
                    time.sleep(2)

                    try:
                        WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "비밀번호를")]')))
                        pw = driver.find_element(By.XPATH, '//*[contains(text(), "비밀번호를")]')

                        if pw:
                            print("pw 만족")
                            # 이전 형제 요소 찾기
                            parent_element = pw.find_element(By.XPATH, '..')
                            previous_sibling = parent_element.find_element(By.XPATH, 'preceding-sibling::*[1]')
                            print("Previous sibling found:", previous_sibling.text)
                            print
                            previous_sibling.click()
                        
                        else:
                            print("요소를 찾을 수 없습니다.")

                        time.sleep(1)
                        driver.get(url_meta)

                    except:
                        pass
                    #알림 제거
                    try:
                        body = driver.find_element(By.CSS_SELECTOR, 'body')
                        ActionChains(driver).move_to_element(body).click().perform()
                    except: pass

                    # if know_TF == 0:  
                        # try:
                        #     WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#js_1v")))
                        #     checkbox = driver.find_element(By.CSS_SELECTOR, "#js_1v")

                        #     if checkbox.is_selected():
                        #         checkbox.click()  # 체크박스가 체크되어 있다면 클릭하여 체크 해제합니다.#캠페인이름
                        #     time.sleep(0.3)

                        # except:
                        #     WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#left_rail_nux_target_node > div._5jln > div > div > div.x1ye3gou.xn6708d.xz9dl7a.xjkvuk6 > div:nth-child(1) > div > div > div.x19kh74d > div > div > div > div:nth-child(1)")))
                        #     checkbox = driver.find_element(By.CSS_SELECTOR, "#left_rail_nux_target_node > div._5jln > div > div > div.x1ye3gou.xn6708d.xz9dl7a.xjkvuk6 > div:nth-child(1) > div > div > div.x19kh74d > div > div > div > div:nth-child(1)")
                        #     if know_TF == 0:
                        #         if checkbox.is_selected():
                        #             checkbox.click()  # 체크박스가 체크되어 있다면 클릭하여 체크 해제합니다.#캠페인이름
                        #         time.sleep(0.3)

                        # try:
                        #     checkbox = driver.find_element(By.CSS_SELECTOR, "#js_3w")
                        #     if checkbox.is_selected():
                        #         checkbox.click()  # 체크박스가 체크되어 있다면 클릭하여 체크 해제합니다.# 일

                        # except:
                        #     checkbox = driver.find_element(By.CSS_SELECTOR, "#left_rail_nux_target_node > div._5jln > div > div > div.x1ye3gou.xn6708d.xz9dl7a.xjkvuk6 > div:nth-child(1) > div > div > div.x19kh74d > div > div > div > div:nth-child(13)")
                        #     if checkbox.is_selected():
                        #         checkbox.click()  # 체크박스가 체크되어 있다면 클릭하여 체크 해제합니다.# 일




                    # 달력 열기
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="PNG_EXPORT"]/div/div[3]/div[1]/div[1]/div/div/div/div[1]/div[2]/div[2]/div/div/div/span')))
                    driver.find_element(By.XPATH, '//*[@id="PNG_EXPORT"]/div/div[3]/div[1]/div[1]/div/div/div/div[1]/div[2]/div[2]/div/div/div/span').click()

                    # 오늘 선택하기
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "오늘")]'))).click()

                    # 달력 열기
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="PNG_EXPORT"]/div/div[3]/div[1]/div[1]/div/div/div/div[1]/div[2]/div[2]/div/div/div/span')))
                    driver.find_element(By.XPATH, '//*[@id="PNG_EXPORT"]/div/div[3]/div[1]/div[1]/div/div/div/div[1]/div[2]/div[2]/div/div/div/span').click()

                    WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, f'//*[contains(text(), "{Tday_month월}")]')))
                    mon = driver.find_element(By.XPATH, f'//*[contains(text(), "{Tday_month월}")]')

                    print(Tday_month월)
                    print(mon.text)

                    if mon.text == Tday_month월:
                        parent_element = mon.find_element(By.XPATH, "..")
                        next_sibling = parent_element.find_element(By.XPATH, 'following-sibling::*[1]')
                        
                    
                        # 시작날짜선택
                        d = next_sibling.find_elements(By.XPATH, ".//div")
                        for i in d:
                            if i.text == today_Tday일:
                                i.click()
                                break
                        # 종료날짜선택
                        for i in d:
                            if i.text == today_Yday일:
                                i.click()
                                break

                        updateElements = driver.find_elements(By.XPATH, '//*[contains(text(), "업데이트")]')

                        if len(updateElements) > 1:  # 요소가 두 개 이상 있는지 확인
                            
                            updateElements[2].click()
                        else:
                            print("두 번째 요소를 찾을 수 없습니다.")
                        

                    elif mon.text != Tday_month월:
                        prevmonth = driver.find_element(By.XPATH, '//*[contains(text(), "이전 달")]')
                        prevmonth.find_element(By.XPATH, "..").click()
                        time.sleep(0.5)

                        mon = driver.find_element(By.XPATH, f'//*[contains(text(), "{Tday_month월}")]')

                        if mon.text == Tday_month월:
                            parent_element = mon.find_element(By.XPATH, "..")
                            next_sibling = parent_element.find_element(By.XPATH, 'following-sibling::*[1]')
                        
                    
                            # 시작날짜선택
                            d = next_sibling.find_elements(By.XPATH, ".//div")
                            for i in d:
                                if i.text == today_Tday일:
                                    i.click()
                                    break

                            # 종료날짜선택
                            if mon.text == Yday_month월:
                                for i in d:
                                    if i.text == today_Tday일:
                                        i.click()
                                        break
                            
                            else:
                                mon = driver.find_element(By.XPATH, f'//*[contains(text(), "{Yday_month월}")]')

                                parent_element = mon.find_element(By.XPATH, "..")
                                next_sibling = parent_element.find_element(By.XPATH, 'following-sibling::*[1]')
                            
                        
                                # 텍스트 값 출력
                                d = next_sibling.find_elements(By.XPATH, ".//div")
                                for i in d:
                                    if i.text == today_Tday일:
                                        i.click()
                                        break

                        updateElements = driver.find_elements(By.XPATH, '//*[contains(text(), "업데이트")]')

                        if len(updateElements) > 1:  # 요소가 두 개 이상 있는지 확인
                            print("요소가 두개입니다.")
                            updateElements[2].click()
                        else:
                            print("두 번째 요소를 찾을 수 없습니다.")
                        

                    else:
                        print("요소를 찾을 수 없습니다.")  

                    
                    element = driver.find_element(By.CSS_SELECTOR, "#export_button > div > div > span > div > div.xeuugli.x2lwn1j.x6s0dn4.x78zum5.x1q0g3np.x1iyjqo2.xozqiw3.x19lwn94.x1hc1fzr.x13dflua.x6o7n8i.xxziih7.x12w9bfk.xl56j7k.xh8yej3 > div > div")
                    driver.find_element(By.CSS_SELECTOR, "#export_button > div > div > span > div > div.xeuugli.x2lwn1j.x6s0dn4.x78zum5.x1q0g3np.x1iyjqo2.xozqiw3.x19lwn94.x1hc1fzr.x13dflua.x6o7n8i.xxziih7.x12w9bfk.xl56j7k.xh8yej3 > div > div").click() #내보내기
                    time.sleep(1.5)
                    ActionChains(driver).move_to_element_with_offset(element,-579,497).click().perform() #다운로드

                    
                    while True:
                        check_download()
                        if check == 0:
                                driver.find_element(By.CSS_SELECTOR, "#export_button > div > div > span > div > div.xeuugli.x2lwn1j.x6s0dn4.x78zum5.x1q0g3np.x1iyjqo2.xozqiw3.x19lwn94.x1hc1fzr.x13dflua.x6o7n8i.xxziih7.x12w9bfk.xl56j7k.xh8yej3 > div > div").click() #내보내기
                                time.sleep(1)
                                ActionChains(driver).move_to_element_with_offset(element,-579,497).click().perform() #다운로드
                        else: break

                        
                    time.sleep(3)
                
                #메타 하엔
                if self.chk_meta_haen.isChecked() == True:
                    url_meta_haen = 'https://adsmanager.facebook.com/adsmanager/reporting/view?act=774078054299392&business_id=341660836507461&selected_report_id=120202962853720679' #하엔
                    sheet_url_haen_all = 'https://docs.google.com/spreadsheets/d/1V8b3FRe_8witwHXQceekgm-BAvTQLwwkcyuaW-mIi30/edit#gid=168246212'
                    sheet_name_haen = '하엔 페이스북 R'
                    know_TF = 0

                    meta(url_meta_haen, know_TF)
                    meta_rawdata(sheet_url_haen_all, sheet_name_haen, know_TF)

                #메타 러블로
                if self.chk_meta_lovl.isChecked() == True:
                    url_meta_lovelo = 'https://adsmanager.facebook.com/adsmanager/reporting/view?act=1913234209031352&business_id=267018165996779&selected_report_id=120200964749160675' #러블로
                    sheet_url_love_all = 'https://docs.google.com/spreadsheets/d/1NVnVJsCj0Ap_o2xabua9ANUw_1IUREVMJKteY_O1yks/edit#gid=1607702031'
                    sheet_name_love = '러블로 페이스북 R'
                    know_TF = 0

                    meta(url_meta_lovelo, know_TF)
                    meta_rawdata(sheet_url_love_all, sheet_name_love, know_TF)

                #메타 노마셀
                if self.chk_meta_know.isChecked() == True:
                    url_meta_knowmycell = 'https://adsmanager.facebook.com/adsmanager/reporting/view?act=238068255778220&business_id=635001998695042&selected_report_id=120200841324100083' #노마셀
                    sheet_url_know_all = 'https://docs.google.com/spreadsheets/d/12FWmZMuznsxOY_IDbBWeSis-EW1Ds1f9TB6X7K6TFBc/edit#gid=137297262'
                    sheet_name_know = '노마셀 페이스북 R'
                    know_TF = 1

                    meta(url_meta_knowmycell, know_TF)
                    meta_rawdata(sheet_url_know_all, sheet_name_know, know_TF)
                    
                # 구글ads n일 전 데이터 추출 기능 추가
                def google_rawdata(sheet_url, sheet_name, brand):

                    target_days = target_days_input
                    dayx = timedelta(days=target_days)
                    day1 = timedelta(days=1)

                    # 오늘 날짜 구하기
                    today_yday = today-day1
                    today_tday = today-dayx
                    
                    csv_file = get_latest_file(download_folder)

                    # 서비스 계정 키 파일 경로
                    credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                    # gspread 클라이언트 초기화
                    client = gspread.service_account(filename=credential_file)

                    # Google 시트 열기
                    spreadsheet = client.open_by_url(sheet_url)

                    # 첫 번째 시트 선택
                    sheet = spreadsheet.worksheet(sheet_name)

                    selected_rows = []

                    with open(csv_file, newline='', encoding='utf-16') as csvfile:
                        reader = csv.reader(csvfile)
                        for i, row in enumerate(reader):
                            if 3 <= i <= 100:  # 범위 내 행 적용

                                # 데이터를 올바르게 파싱하기 위해 먼저 전체 문자열을 하나로 합친다
                                full_data = "".join(row)

                                # 탭(\t)으로 데이터를 분리한다
                                parsed_data = full_data.split('\t')

                                cleaned_data = [item.replace('"', '') for item in parsed_data]
                                selected_rows.append(cleaned_data)
                        print(selected_rows)

                    updates = []
                    formats = []

                    while today_tday != today:

                        last_row = len(sheet.col_values(1))
                        next_row = int(last_row) + 1

                        print(today_tday, "검색 시작")
                        for i in selected_rows:
                            if str(today_tday) in i:
                                new_selected_rows = []
                                result = []
                                print(today_tday, "찾음!")

                                for item in i:
                                    if isinstance(item, str) and '%' in item:
                                        result.append(float(item.strip('%')) / 100)
                                    elif isinstance(item, str) and ',' in item and '.' in item.replace(',', ''):
                                        result.append(float(item.replace(',', '')))
                                    elif isinstance(item, str) and ',' in item:
                                        result.append(int(item.replace(',', '')))
                                    elif isinstance(item, str) and item.replace('.', '', 1).isdigit() and item.count('.') == 1:
                                        result.append(float(item))
                                    elif isinstance(item, str) and item.isdigit():
                                        result.append(int(item))
                                    else:
                                        result.append(item)

                                new_selected_rows.append(result)
                                print(new_selected_rows)

                                range_to_write = f'A{next_row}:M{next_row}'
                                updates.append({'range': range_to_write, 'values': new_selected_rows})

                                next_row += 1

                                if brand == "하엔":
                                    print("하엔 입력 시작")

                                    # range_to_write = f'A{next_row}:M{next_row}'
                                    # updates.append({'range': range_to_write, 'values': new_selected_rows})
                                    # sheet.update(new_selected_rows, range_to_write)

                                    # # 셀 포맷 설정
                                    # sheet.format(f"H{next_row}", {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}})
                                    # sheet.format(f'I{next_row}:L{next_row}', {"numberFormat": {"type":'NUMBER'}})
                                    # time.sleep(1)
                                    # 셀 포맷 설정
                                    formats.append({
                                        'range': f"H{next_row}",
                                        'format': {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}}
                                    })
                                    formats.append({
                                        'range': f'I{next_row}:L{next_row}',
                                        'format': {"numberFormat": {"type": 'NUMBER'}}
                                    })

                                if brand == "노마셀":
                                    print("노마셀 입력 시작")
                                    print(new_selected_rows)

                                    # range_to_write = f'A{next_row}:K{next_row}'
                                    # updates.append({'range': range_to_write, 'values': new_selected_rows})

                                    # sheet.update(new_selected_rows, range_to_write)

                                    # # 셀 포맷 설정
                                    # sheet.format(f"G{next_row}", {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}})
                                    # sheet.format(f'H{next_row}:J{next_row}', {"numberFormat": {"type":'NUMBER'}})
                                    # time.sleep(1)

                                    # 셀 포맷 설정
                                    formats.append({
                                        'range': f"G{next_row}",
                                        'format': {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}}
                                    })
                                    formats.append({
                                        'range': f'H{next_row}:J{next_row}',
                                        'format': {"numberFormat": {"type": 'NUMBER'}}
                                    })

                        if updates:
                            print(updates)
                            sheet.batch_update(updates)
                            for fmt in formats:
                                sheet.format(fmt['range'], fmt['format'])
                            updates.clear()
                            formats.clear()
                            time.sleep(0.5)  # 각 배치 요청 사이에 지연 시간을 추가
                            print("Batch update and format applied.")

                        today_tday += timedelta(days=1)

                
                def google(url_google):
                    driver.get(url_google)
                    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CLASS_NAME, 'button-text')))
                    calOpen = driver.find_element(By.CLASS_NAME, 'button-text')

                    time.sleep(1)
                    # 달력 열기
                    calOpen.click()

                    # 날짜 선택
                    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CLASS_NAME, 'visible-month')))
                    time.sleep(1.5)
                    elements = driver.find_elements(By.CSS_SELECTOR, 'material-select-item')
                    for element in elements:
                        if '지난 7일(어제까지)' in element.text:
                            print("Element found:", element.text)
                            element.click()
                            break
                    # WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "지난 7일(어제까지)")]')))
                    # weekElement = driver.find_element(By.XPATH, '//*[contains(text(), "지난 7일(어제까지)")]')
                    # weekElement.click()

                    time.sleep(1)
                    # 11. 구글 다운로드 클릭, CSS선택자 방식에서 XPATH - contains, text() 방식으로 변경(CSS선택자 매번바뀜)
                    schedule = driver.find_element(By.XPATH, "//*[contains(text(), '일정')]")
                    #다운
                    # 부모의 부모의 부모의 부모의 부모의 이전 요소 찾기 및 클릭
                    previous_sibling = schedule.find_element(By.XPATH, "ancestor::*[4]/preceding-sibling::*[1]")
                    previous_sibling.click()


                    # Excel .csv 선택
                    lists = driver.find_element(By.CLASS_NAME, "download-dropdown").find_elements(By.CLASS_NAME, "item")

                    for item in lists:
                        print(item.text)
                        if item.text == "Excel .csv":
                            item.click()
                            break

## 구글 다운로드 실패 시 재시도 적용
                    try:
                        check_download()
                    except:
                        schedule = driver.find_element(By.XPATH, "//*[contains(text(), '일정')]")
                        #다운
                        # 부모의 부모의 부모의 부모의 부모의 이전 요소 찾기 및 클릭
                        previous_sibling = schedule.find_element(By.XPATH, "ancestor::*[4]/preceding-sibling::*[1]")
                        previous_sibling.click()


                        # Excel .csv 선택
                        lists = driver.find_element(By.CLASS_NAME, "download-dropdown").find_elements(By.CLASS_NAME, "item")

                        for item in lists:
                            print(item.text)
                            if item.text == "Excel .csv":
                                item.click()
                                break
                        check_download()


                    time.sleep(2)

                
                # 구글
                if self.chk_goog_haen.isChecked() == True:
                    url_ads_haen = 'https://ads.google.com/aw/reporteditor/view?ocid=1181720304&workspaceId=0&reportId=927965366&euid=1114690018&__u=8943315282&uscid=1181720304&__c=5821258096&authuser=0'
                    sheet_url_goog = "https://docs.google.com/spreadsheets/d/1V8b3FRe_8witwHXQceekgm-BAvTQLwwkcyuaW-mIi30/edit#gid=1966867512"
                    sheet_name_goog = "하엔 구글 R"
                    brand = "하엔"
                    google(url_ads_haen)
                    google_rawdata(sheet_url_goog, sheet_name_goog, brand)


                if self.chk_goog_know.isChecked() == True:
                    url_ads_know = 'https://ads.google.com/aw/reporteditor/view?ocid=1379143590&workspaceId=-1615213561&reportId=928192574&euid=1114690018&__u=8943315282&uscid=1379143590&__c=4267857910&authuser=0'
                    sheet_url_goog = "https://docs.google.com/spreadsheets/d/12FWmZMuznsxOY_IDbBWeSis-EW1Ds1f9TB6X7K6TFBc/edit#gid=1001228164"
                    sheet_name_goog = "노마셀 구글 R"
                    brand = "노마셀"
                    google(url_ads_know)
                    google_rawdata(sheet_url_goog, sheet_name_goog, brand)


                def ssDown(brand):

                    # 날짜 구하기
                    today = date.today()

                    today_date = today.strftime("%d")
                    today_Ym = today.strftime("%Y. %m.")

                    number = target_days_input
                    dayx = datetime.timedelta(days=number)
                    dayy = datetime.timedelta(days=1)
                    day1 = datetime.timedelta(days=1)

                    today_yday = today-day1
                    startday = today-dayx
                    endday = today-dayy
                    tday_Ym = startday.strftime("%Y. %m.")
                    tday_d = startday.strftime("%d")

                    # EdgeOptions 객체 생성
                    edge_options = webdriver.EdgeOptions()
                    edge_options.use_chromium = True
                    edge_options.add_argument("disable-gpu")
                    edge_options.add_argument("no-sandbox")


                    # 사용자의 프로필 경로 설정
                    profile_path = 'C:\\Users\\A\\AppData\\Local\\Microsoft\\Edge\\User Data1'
                    edge_options.add_argument(f"user-data-dir={profile_path}")
                    edge_options.add_argument("--profile-directory=Default")


                    # Edge 드라이버 서비스 시작
                    edge_service = Service(executable_path='C:\\path\\to\\edgedriver_win64\\msedgedriver.exe')
                    edge_driver = webdriver.Edge(service=edge_service, options=edge_options)


                    edge_driver.get("https://bizadvisor.naver.com/shopping/product")

                    # 로그인
                    WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#wrap > div > div > div.login_box > ul > li:nth-child(1) > a"))).click()
                    try:
                        WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#root > div > div.Layout_wrap__3uDBh > div > div > div.Login_simple_box__2bfAS > button"))).click()
                    
                    except:
                        driver.find_element(By.CSS_SELECTOR, '[class^="Login_btn_more"]').click()

                        current_window_handle = driver.current_window_handle

                        new_window_handle = None
                        while not new_window_handle:
                            for handle in driver.window_handles:
                                if handle != current_window_handle:
                                    new_window_handle = handle
                                    break

                        #팝업으로 제어 변경
                        driver.switch_to.window(driver.window_handles[1])


                        WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#log\.login")))
                        
                        txtInput = edge_driver.find_element(By.CSS_SELECTOR, "#id")
                        txtInput.send_keys("wntlsqhr")
                        time.sleep(0.1)
                        txtInput = edge_driver.find_element(By.CSS_SELECTOR, "#pw")
                        txtInput.send_keys("dnflskfk00@")
                        time.sleep(0.1)
                        driver.find_element(By.CSS_SELECTOR, "#log\.login")

                        #원래 페이지로 제어 변경
                        edge_driver.switch_to.window(edge_driver.window_handles[0])


                    # 상품별 이동
                    WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#include_nav > div > div > div:nth-child(1) > ul > li:nth-child(4) > a"))).click()
                    WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#include_nav > div > div > div:nth-child(1) > ul > li.on > div > ul > li:nth-child(1) > a"))).click()

                    brandtext = edge_driver.find_element(By.CSS_SELECTOR, "#include_header > div > div.header_tit > div > div:nth-child(2) > div > div > div > span:nth-child(1)").text[:3]


                    # 브랜드 변경
                    if not brandtext == brand:
                        edge_driver.find_element(By.CSS_SELECTOR, "#include_header > div > div.header_tit > div > div:nth-child(2) > div > div > div > div > a > div > span").click()

                    while startday != today:

                        # 날짜 클릭(달력오픈)
                        WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.date_select > a.btn.select_data'))).click()


                        # 날짜 변수 지정
                        tday_Ym = startday.strftime("%Y. %m.")
                        tday_d = str(int(startday.strftime("%d")))
                        trick = (startday-day1).strftime("%Y. %m.")
                        print(startday)


                        # %Y. %m 표시(웹상)
                        DPmonthStart = edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_area > div.pick_calendar_layout > div.DayPicker.DayPicker > div > div.DayPicker-Month.rdp-caption_start")
                        DPmonthBtw = edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_area > div.pick_calendar_layout > div.DayPicker.DayPicker > div > div.DayPicker-Month.rdp-caption_between")


                        # 첫번째 단락 년,월 대조
                        if tday_Ym == DPmonthStart.text[:9]:

                                days = edge_driver.find_elements(By.XPATH, f"//*[@id='wrap']/div[1]/section/div/div[2]/div[1]/div/ul/li[2]/div/div/div[2]/div/div[1]/div[2]/div[2]/div/div[1]/table/tbody//td[not(contains(@class, 'DayPicker-Day--outside'))]")
                                for day in days:
                                    if day.text == tday_d:
                                        print("target: ", day.text)
                                        day.click()
                                        time.sleep(0.1)
                                        day.click()
                                        break

                                # 적용
                                edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_info_area > div.btn_area > a:nth-child(1)").click()
                        
                        
                        # 두번째 단락 년,월 대조
                        elif tday_Ym == DPmonthBtw.text[:9]:
                        
                                days = edge_driver.find_elements(By.XPATH, f"//*[@id='wrap']/div[1]/section/div/div[2]/div[1]/div/ul/li[2]/div/div/div[2]/div/div[1]/div[2]/div[2]/div/div[2]/table/tbody//td[not(contains(@class, 'DayPicker-Day--outside'))]")
                                for day in days:
                                    if day.text == tday_d:
                                        print("target: ", day.text)
                                        day.click()
                                        time.sleep(0.1)
                                        day.click()
                                        break
                                    
                                # 적용
                                edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_info_area > div.btn_area > a:nth-child(1)").click()

                        # 5. 네이버 스스 이전 달 날짜 선택 안되는 코드 변경
                        else:
                            # 이전 달로 이동
                            WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_area > div.pick_calendar_layout > div.DayPicker-NavBar > span.DayPicker-NavButton.DayPicker-NavButton--prev'))).click()

                            if tday_Ym == DPmonthStart.text[:9]:

                                days = edge_driver.find_elements(By.XPATH, f"//*[@id='wrap']/div[1]/section/div/div[2]/div[1]/div/ul/li[2]/div/div/div[2]/div/div[1]/div[2]/div[2]/div/div[1]/table/tbody//td[not(contains(@class, 'DayPicker-Day--outside'))]")
                                for day in days:
                                    if day.text == tday_d:
                                        print("target: ", day.text)
                                        day.click()
                                        time.sleep(0.1)
                                        day.click()
                                        break

                                # 적용
                                edge_driver.find_element(By.CSS_SELECTOR, "#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(2) > div > div > div.calendar_lypop > div > div.pick_info_area > div.btn_area > a:nth-child(1)").click()
                        
                    

                        # 다운로드 버튼
                        WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#wrap > div:nth-child(1) > section > div > div.fixed_header.on > div.tit_area > div > ul > li:nth-child(1) > span > a'))).click()

                        check_download()

                        startday += datetime.timedelta(days=1)

                def ssWrite(brandSheet, sheet_url):
                    defaultData = ["화장품/미용", "바디케어", "입욕제", "-", "러블로 러브슬라임 슬라임탕 젤 입욕제 젤탕", "9019908272",	"일반배송",	"0", "0", "0", "0.00%"]
                    # 날짜 구하기
                    today = date.today()

                    today_date = today.strftime("%d")
                    today_Ym = today.strftime("%Y. %m.")

                    number = target_days_input
                    dayx = datetime.timedelta(days=number)
                    dayy = datetime.timedelta(days=1)
                    day1 = datetime.timedelta(days=1)

                    today_yday = today-day1
                    startday = today-dayx
                    endday = today-dayy
                    tday_Ym = startday.strftime("%Y. %m.")
                    tday_d = startday.strftime("%d")

                    while number > 0:

                        # 서비스 계정 키 파일 경로
                        credential_file = 'triple-nectar-412808-da4dac0cc16e.json'
                        # gspread 클라이언트 초기화
                        client = gspread.service_account(filename=credential_file)
                        # Google 시트 열기
                        spreadsheet = client.open_by_url(sheet_url)
                        # 첫 번째 시트 선택
                        sheet = spreadsheet.worksheet(brandSheet)

                        last_row = len(sheet.get_all_values())
                        print(last_row)
                        next_row = last_row + 1  # 다음 행 번호

                        i = get_nth_latest_file(download_folder, number)

                        wb = openpyxl.load_workbook(i)
                        sheet = wb.active  # 활성 시트 선택

                        if check_data_in_second_row(i):
                            pass

                        else:
                            sheet = spreadsheet.worksheet(brandSheet)

                            # 날짜 입력
                            sheet.update([[str(startday)]], f"A{next_row}")
                            range_to_write = f'B{next_row}:L{next_row}'
                            sheet.update([defaultData], range_to_write)
                            number -= 1
                            startday += timedelta(days=1)  # 날짜 하루 증가
                            continue


                        # 원본 시트의 행을 반복하며 첫 번째 행을 제외하고 데이터가 있는 행만 복사
                        for row in sheet.iter_rows(min_row=2):  # 첫 번째 행을 제외하고 시작
                            # 각 셀에 데이터가 있는지 확인
                            data_exists = any(cell.value not in (None, '', ' ') for cell in row)  # 빈 문자열과 공백도 무시

                            # 서비스 계정 키 파일 경로
                            credential_file = 'triple-nectar-412808-da4dac0cc16e.json'
                            # gspread 클라이언트 초기화
                            client = gspread.service_account(filename=credential_file)
                            # Google 시트 열기
                            spreadsheet = client.open_by_url(sheet_url)
                            # 첫 번째 시트 선택
                            sheet = spreadsheet.worksheet(brandSheet)

                            # 날짜 입력
                            sheet.update([[str(startday)]], f"A{next_row}")
                            
                            values = []
                            for col_index, cell in enumerate(row, start=2):
                                values.append(cell.value)

                            range_to_write = f'B{next_row}:L{next_row}'
                            sheet.update([values], range_to_write)

                            
                            next_row += 1

                        startday += timedelta(days=1)  # 날짜 하루 증가
                        number -= 1

                # if self.chk_ss_haen.isChecked() == True:

                #     label = self.chk_cafe_haen
                #     url_cafe24 = "https://eclogin.cafe24.com/Shop/"
                #     url_cafe24_req_haen = "https://woo8425.cafe24.com/disp/admin/shop1/report/DailyList"
                    
                #     cafe24_id_haen = self.login_info("CAFE_HAEN_ID")
                #     cafe24_pw_haen = self.login_info("CAFE_HAEN_PW")

                #     sheet_haenR_url = 'https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=1894651086'
                #     sheet_haenR = '하엔R'
                #     sheet_haenD = "하엔D"

                sheet_url = "https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=374561563"
                if self.chk_ss_lovl.isChecked() == True:

                    brand = "러브슬"
                    brandSheet = "러블로N"

                    ssDown(brand)
                    ssWrite(brandSheet, sheet_url)


                    label = self.chk_cafe_haen
                    url_cafe24 = "https://eclogin.cafe24.com/Shop/"
                    url_cafe24_req_haen = "https://woo8425.cafe24.com/disp/admin/shop1/report/DailyList"
                    
                    cafe24_id_haen = self.login_info("CAFE_HAEN_ID")
                    cafe24_pw_haen = self.login_info("CAFE_HAEN_PW")

                    sheet_haenR_url = 'https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=1894651086'
                    sheet_haenR = '하엔R'
                    sheet_haenD = "하엔D"

                if self.chk_ss_know.isChecked() == True:

                    brand = "노마셀"
                    brandSheet = "노마셀N"

                    ssDown(brand)
                    ssWrite(brandSheet, sheet_url)


                    label = self.chk_cafe_haen
                    url_cafe24 = "https://eclogin.cafe24.com/Shop/"
                    url_cafe24_req_haen = "https://woo8425.cafe24.com/disp/admin/shop1/report/DailyList"
                    
                    cafe24_id_haen = self.login_info("CAFE_HAEN_ID")
                    cafe24_pw_haen = self.login_info("CAFE_HAEN_PW")

                    sheet_haenR_url = 'https://docs.google.com/spreadsheets/d/145lVmBVqp87AwsRK9KCclE-Dgkh0B7jbwsfaHKmwOz0/edit#gid=1894651086'
                    sheet_haenR = '하엔R'
                    sheet_haenD = "하엔D"

                def naverad(url):

                    # EdgeOptions 객체 생성
                    edge_options = webdriver.EdgeOptions()
                    edge_options.use_chromium = True
                    edge_options.add_argument("disable-gpu")
                    edge_options.add_argument("no-sandbox")


                    # 사용자의 프로필 경로 설정
                    profile_path = 'C:\\Users\\A\\AppData\\Local\\Microsoft\\Edge\\User Data1'
                    edge_options.add_argument(f"user-data-dir={profile_path}")
                    edge_options.add_argument("--profile-directory=Default")


                    # Edge 드라이버 서비스 시작
                    edge_service = Service(executable_path='C:\\path\\to\\edgedriver_win64\\msedgedriver.exe')
                    edge_driver = webdriver.Edge(service=edge_service, options=edge_options)

                    edge_driver.get(url)

                    # 네이버검색광고 로그인 확인 창 제거 로직 수정(visibility_of_all_elements_located -> element_to_be_clickable)
                    try:
                        WebDriverWait(edge_driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "환영합니다")]')))
                        edge_driver.find_element(By.CLASS_NAME, "btn_name").click()
                        print("로그인확인 창 제거")
                    except: 
                        print("로그인확인 창 없음")
                        pass
                    
                    # 네이버검색광고 로그인 확인 창 제거 로직 수정(visibility_of_all_elements_located -> element_to_be_clickable)
                    try:
                        WebDriverWait(edge_driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "환영합니다")]')))
                        checkbox = edge_driver.find_element(By.ID, "chk_cls")
                        checkbox.click()

                    except: pass

                    # 캘린더 열기
                    WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-nclick="datePicker"]'))).click()

                    # 날짜 선택
                    WebDriverWait(edge_driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "지난 7일")]'))).click()

                    time.sleep(0.5)
                    element = edge_driver.find_element(By.XPATH, "//*[contains(text(), '다운로드') and not(contains(text(), '대용량 다운로드 보고서'))]")
                    time.sleep(0.5)

                    # 다운로드 확인
                    cnt = 1
                    while cnt < 10:
                        current_file_count1 = count_files(download_folder)
                        element.click()
                        time.sleep(3)
                        current_file_count2 = count_files(download_folder)
                        if current_file_count1 != current_file_count2:
                            break

                        cnt += 1
                        
                    
                    time.sleep(1)
                    edge_driver.close()

                def naveradInput(url, name):

                    target_days = target_days_input
                    dayx = datetime.timedelta(days=target_days)
                    day1 = datetime.timedelta(days=1)

                    # 오늘 날짜 구하기
                    today_yday = today-day1
                    today_tday = today-dayx
                    # CSV 파일 읽기 (첫 번째 행은 건너뛰고 두 번째 행을 열 이름으로 사용)
                    df = pd.read_csv(get_latest_file(download_folder), skiprows=1)

                    # 날짜 열 이름 추출 (A열, 즉 첫 번째 열)
                    date_column = df.columns[0]

                    # 'date' 열을 datetime 형식으로 변환
                    df[date_column] = pd.to_datetime(df[date_column], format='%Y.%m.%d.')

                    df[date_column] = df[date_column].dt.strftime('%Y-%m-%d')
                    # 필터링된 데이터프레임 출력
                    dataList = df.values.tolist()
                    print(df.values.tolist())

                    csv_file = get_latest_file(download_folder)

                    # 서비스 계정 키 파일 경로
                    credential_file = 'triple-nectar-412808-da4dac0cc16e.json'

                    # gspread 클라이언트 초기화
                    client = gspread.service_account(filename=credential_file)

                    # Google 시트 열기
                    spreadsheet = client.open_by_url(url)

                    # 첫 번째 시트 선택
                    sheet = spreadsheet.worksheet(name)

                    while today_tday != today:
                        print(today_tday, "검색 시작")
                        for i in dataList:
                            if str(today_tday) in i:
                                result = []
                                print(today_tday, "찾음!")

                                for item in i:
                                    if isinstance(item, str) and '%' in item:
                                        result.append(float(item.strip('%')) / 100)
                                    elif isinstance(item, str) and ',' in item and '.' in item.replace(',', ''):
                                        result.append(float(item.replace(',', '')))
                                    elif isinstance(item, str) and ',' in item:
                                        result.append(int(item.replace(',', '')))
                                    elif isinstance(item, str) and item.replace('.', '', 1).isdigit() and item.count('.') == 1:
                                        result.append(float(item))
                                    elif isinstance(item, str) and item.isdigit():
                                        result.append(int(item))
                                    else:
                                        result.append(item)


                                last_row = len(sheet.col_values(1))
                                next_row = int(last_row) + 1
                                range_to_write = f'A{next_row}:S{next_row}'
                                sheet.update([result], range_to_write)
                        today_tday += timedelta(days=1)

                if self.chk_nad_haen.isChecked() == True:

                    sheet_url = 'https://docs.google.com/spreadsheets/d/1V8b3FRe_8witwHXQceekgm-BAvTQLwwkcyuaW-mIi30/edit?gid=2136174248#gid=2136174248'
                    sheet_name = '하엔 네이버 R'
                    target_url = "https://manage.searchad.naver.com/customers/2621471/reports/rtt-a001-000000000650376"
                    
                    naverad(target_url)
                    naveradInput(sheet_url, sheet_name)

                if self.chk_nad_lovl.isChecked() == True:

                    sheet_url = 'https://docs.google.com/spreadsheets/d/1NVnVJsCj0Ap_o2xabua9ANUw_1IUREVMJKteY_O1yks/edit?gid=910059812#gid=910059812'
                    sheet_name = '러블로 네이버 R'
                    target_url = "https://manage.searchad.naver.com/customers/2914810/reports/rtt-a001-000000000651901"
                    
                    naverad(target_url)
                    naveradInput(sheet_url, sheet_name)

                if self.chk_nad_know.isChecked() == True:

                    sheet_url = 'https://docs.google.com/spreadsheets/d/12FWmZMuznsxOY_IDbBWeSis-EW1Ds1f9TB6X7K6TFBc/edit?gid=1997928779#gid=1997928779'
                    sheet_name = '노마셀 네이버 R'
                    target_url = "https://manage.searchad.naver.com/customers/2957190/reports/rtt-a001-000000000651985"
                    
                    naverad(target_url)
                    naveradInput(sheet_url, sheet_name)


        except UnicodeDecodeError as e:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("다운로드 경로를 확인해주세요")
            msg.setWindowTitle("에러")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.setWindowFlags(msg.windowFlags() | Qt.WindowStaysOnTopHint)
            msg.exec_()
            return
        
        except InvalidFileException as e:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("다운로드 경로를 확인해주세요")
            msg.setWindowTitle("에러")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.setWindowFlags(msg.windowFlags() | Qt.WindowStaysOnTopHint)
            msg.exec_()
            return
        
        # except SessionNotCreatedException as e:
        #     msg = QMessageBox()
        #     msg.setIcon(QMessageBox.Warning)
        #     msg.setText("크롬 종료 후 다시 실행해주세요")
        #     msg.setWindowTitle("에러")
        #     msg.setStandardButtons(QMessageBox.Ok)
        #     msg.setWindowFlags(msg.windowFlags() | Qt.WindowStaysOnTopHint)
        #     msg.exec_()
        #     return
        

        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText("추출이 완료되었습니다")
        msg.setWindowTitle("알림")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.setWindowFlags(msg.windowFlags() | Qt.WindowStaysOnTopHint)
        msg.exec_()
        return


    def saveText(self):
        text = self.path_folder.text()
        text1 = self.chrome_path_folder.text()
        with open('saved_text.txt', 'w') as file:
            file.write(text)
            file.write("\n")
            file.write(text1)
        QMessageBox.information(self,'알림','저장되었습니다.')

        with open('checkbox_state.txt', 'w') as file:
            file.write(f"{self.chk_cafe_haen.isChecked()}\n")
            file.write(f"{self.chk_cafe_lovl.isChecked()}\n")
            file.write(f"{self.chk_cafe_know.isChecked()}\n")
            file.write(f"{self.chk_cafe_ZQ.isChecked()}\n")

            file.write(f"{self.chk_coup_haen.isChecked()}\n")
            file.write(f"{self.chk_coup_lovl.isChecked()}\n")
            file.write(f"{self.chk_coup_know.isChecked()}\n")

            file.write(f"{self.chk_meta_haen.isChecked()}\n")
            file.write(f"{self.chk_meta_lovl.isChecked()}\n")
            file.write(f"{self.chk_meta_know.isChecked()}\n")

            file.write(f"{self.chk_goog_haen.isChecked()}\n")
            file.write(f"{self.chk_goog_lovl.isChecked()}\n")
            file.write(f"{self.chk_goog_know.isChecked()}\n")

            file.write(f"{self.chk_ss_haen.isChecked()}\n")
            file.write(f"{self.chk_ss_lovl.isChecked()}\n")
            file.write(f"{self.chk_ss_know.isChecked()}\n")

            file.write(f"{self.chk_pc_haen.isChecked()}\n")
            file.write(f"{self.chk_pc_lovl.isChecked()}\n")
            file.write(f"{self.chk_pc_know.isChecked()}\n")

            file.write(f"{self.chk_nad_haen.isChecked()}\n")
            file.write(f"{self.chk_nad_lovl.isChecked()}\n")
            file.write(f"{self.chk_nad_know.isChecked()}\n")


    def loadCheckboxState(self):
        try:
            with open('checkbox_state.txt', 'r') as file:
                states = file.readlines()
                self.chk_cafe_haen.setChecked(states[0].strip() == 'True')
                self.chk_cafe_lovl.setChecked(states[1].strip() == 'True')
                self.chk_cafe_know.setChecked(states[2].strip() == 'True')
                self.chk_cafe_ZQ.setChecked(states[3].strip() == 'True')

                self.chk_coup_haen.setChecked(states[4].strip() == 'True')
                self.chk_coup_lovl.setChecked(states[5].strip() == 'True')
                self.chk_coup_know.setChecked(states[6].strip() == 'True')

                self.chk_meta_haen.setChecked(states[7].strip() == 'True')
                self.chk_meta_lovl.setChecked(states[8].strip() == 'True')
                self.chk_meta_know.setChecked(states[9].strip() == 'True')

                self.chk_goog_haen.setChecked(states[10].strip() == 'True')
                self.chk_goog_lovl.setChecked(states[11].strip() == 'True')
                self.chk_goog_know.setChecked(states[12].strip() == 'True')

                self.chk_ss_haen.setChecked(states[13].strip() == 'True')
                self.chk_ss_lovl.setChecked(states[14].strip() == 'True')
                self.chk_ss_know.setChecked(states[15].strip() == 'True')

                self.chk_pc_haen.setChecked(states[16].strip() == 'True')
                self.chk_pc_lovl.setChecked(states[17].strip() == 'True')
                self.chk_pc_know.setChecked(states[18].strip() == 'True')

                self.chk_nad_haen.setChecked(states[19].strip() == 'True')
                self.chk_nad_lovl.setChecked(states[20].strip() == 'True')
                self.chk_nad_know.setChecked(states[21].strip() == 'True')
                # 나머지 체크박스도 동일하게 불러옵니다.
        except FileNotFoundError:
            pass

    def loadText(self):
        try:
            with open('saved_text.txt', 'r') as f:
                saved_text = f.read()
                texts = saved_text.split("\n")

                self.path_folder.setText(texts[0])
                self.chrome_path_folder.setText(texts[1])

                
        except FileNotFoundError:
            pass

    def login_info(self, target_word):
        try:
            with open('login_info.txt', 'r', encoding='utf-8') as f:
                lines = f.readlines()  # 파일의 모든 줄을 읽어 리스트로 저장

            # 모든 줄을 순회하면서 target_word 찾기
            for i, line in enumerate(lines):
                if target_word in line:  # 현재 줄에 target_word가 포함되어 있는지 확인
                    if i + 1 < len(lines):  # 다음 줄이 존재하는지 확인
                        print(lines[i + 1].strip())  # 다음 줄의 내용을 프린트 (공백 제거)
                        return(lines[i + 1].strip())
        except FileNotFoundError: print("cannot find login information.")

    def folderopen(self):
        fname = QFileDialog.getExistingDirectory(self,'폴더선택','')
        self.path_folder.setText(fname)
    
    def chromefolderopen(self):
        fname = QFileDialog.getExistingDirectory(self,'폴더선택','')
        self.chrome_path_folder.setText(fname)

    def my_exception_hook(exctype, value, traceback):
        # Print the error and traceback
        print(exctype, value, traceback)
        # Call the normal Exception hook after
        sys._excepthook(exctype, value, traceback)
        # sys.exit(1)

    # Back up the reference to the exceptionhook
    sys._excepthook = sys.excepthook

    # Set the exception hook to our wrapping function
    sys.excepthook = my_exception_hook

if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = Rawdata_extractor()
    win.show()
    sys.exit(app.exec_())
